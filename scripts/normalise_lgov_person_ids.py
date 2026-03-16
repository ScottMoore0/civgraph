"""
Normalise PersonIDs in lgov-modern-wikipedia.stvfix.xlsx

- Strips ‡ from all name fields and re-parses First/Last names
- Merges 29 confirmed duplicate pairs
- Replaces large hash IDs (>999999) with new sequential 6-digit IDs
- Zero-pads all small IDs to 6 digits
- Stores all PersonIDs as strings to preserve leading zeros
- Saves to lgov-modern-wikipedia.stvfix.normalised.xlsx
"""

import re
import openpyxl
from copy import copy

INPUT = r"C:\Users\scomo\boundaries-website\_tmp_xls2rar_extract\out\wiki_lgov_modern\lgov-modern-wikipedia.stvfix.xlsx"
OUTPUT = r"C:\Users\scomo\boundaries-website\_tmp_xls2rar_extract\out\wiki_lgov_modern\lgov-modern-wikipedia.stvfix.normalised.xlsx"

# Step 1: 29 confirmed duplicate pairs (keep, discard)
DUPLICATE_PAIRS = [
    (2859679412, 2393825577),   # bill keery
    (1903914006, 3074483928),   # carole howard
    (3776063797, 1215270263),   # darryn causby
    (613577180, 1019292244),    # david arthurs
    (761967322, 1007484523),    # declan boyle
    (1186157040, 4051152685),   # fergal lennon
    (495353946, 3220327211),    # garbhan mcphillips
    (597473620, 1095705866),    # graham craig
    (2670978434, 335063882),    # graham warke
    (2680996767, 3264286482),   # ivor wallace
    (2187559242, 2865165391),   # james mccorkell
    (2747246846, 2041427362),   # jason barr
    (2919348200, 3502447071),   # jim mckeever
    (1118443049, 2152263802),   # joanne donnelly
    (2331622464, 3215679580),   # john mcdermott
    (1676651428, 1161056368),   # john palmer
    (2915401292, 2757776026),   # jolene bunting
    (954017054, 3886370180),    # johnny mccarthy
    (892514497, 3654209875),    # kate mullan
    (3219319994, 535878283),    # margaret anne mckillop
    (4128906175, 1669932576),   # nathan anderson
    (3319019087, 4072193164),   # noelle robinson
    (167116203, 1122855962),    # patrick convery
    (4058284259, 4141168732),   # paul mccusker
    (3689006369, 3819456676),   # rosemarie shields
    (2515393861, 2866883373),   # sharon mckillop
    (3052633704, 364353721),    # shauna cusack
    (2116483353, 3257103599),   # wesley irvine
    (3865899751, 3238821678),   # alan lewis
]

# Existing 6-digit IDs that stay as-is
EXISTING_6DIGIT = {100001, 100002, 100003, 100006, 100007, 819405}

def collect_all_person_ids(wb):
    """Collect all unique PersonIDs from both sheets."""
    ids = set()
    ws = wb["ElectionResults"]
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 132).value
        if pid is not None:
            ids.add(to_int(pid))
        # TransferSubject1-23 (cols 16, 21, 26, ... every 5 cols)
        for ts_col in range(16, 127, 5):
            val = ws.cell(r, ts_col).value
            if val:
                for part in str(val).split(","):
                    part = part.strip()
                    if part:
                        try:
                            ids.add(int(float(part)))
                        except ValueError:
                            pass

    ws2 = wb["Transfers"]
    for r in range(2, ws2.max_row + 1):
        for col in [7, 26]:  # PersonID, SourcePersonID
            pid = ws2.cell(r, col).value
            if pid is not None:
                ids.add(to_int(pid))
        val = ws2.cell(r, 15).value  # TransferSubject
        if val:
            for part in str(val).split(","):
                part = part.strip()
                if part:
                    try:
                        ids.add(int(float(part)))
                    except ValueError:
                        pass
    return ids


def to_int(val):
    if isinstance(val, float):
        return int(val)
    if isinstance(val, str):
        return int(float(val))
    return int(val)


def build_mapping(all_ids):
    """Build the complete old-ID -> new-6-digit-string mapping."""
    # Step 1: duplicate merge mapping (discard -> keep)
    discard_to_keep = {d: k for k, d in DUPLICATE_PAIRS}
    discarded_ids = set(discard_to_keep.keys())

    # Categorise all IDs
    small_ids = set()      # 1-5 digits, need zero-padding
    six_digit_ids = set()   # already 6 digits, keep as-is
    hash_ids = set()        # >999999 and not in EXISTING_6DIGIT

    for pid in all_ids:
        if pid in discarded_ids:
            continue  # will be merged
        if pid < 100000:
            small_ids.add(pid)
        elif pid in EXISTING_6DIGIT:
            six_digit_ids.add(pid)
        elif 100000 <= pid <= 999999:
            six_digit_ids.add(pid)
        else:
            hash_ids.add(pid)

    # Also include kept IDs from duplicates that are hashes
    for keep_id in [k for k, _ in DUPLICATE_PAIRS]:
        if keep_id > 999999 and keep_id not in EXISTING_6DIGIT:
            hash_ids.add(keep_id)

    print(f"Small IDs (to zero-pad): {len(small_ids)}")
    print(f"Existing 6-digit IDs (keep): {len(six_digit_ids)}")
    print(f"Hash IDs (need new 6-digit): {len(hash_ids)}")
    print(f"Discarded (merged): {len(discarded_ids)}")

    # Build the mapping: old_id (int) -> new_id (str, 6 chars)
    mapping = {}

    # Small IDs: zero-pad to 6 digits
    for pid in small_ids:
        mapping[pid] = f"{pid:06d}"

    # Existing 6-digit IDs: keep as strings
    for pid in six_digit_ids:
        mapping[pid] = f"{pid:06d}"

    # Hash IDs: assign new sequential 6-digit IDs starting from 100020
    next_id = 100020
    for pid in sorted(hash_ids):  # sort for deterministic assignment
        while next_id in EXISTING_6DIGIT or next_id == 819405:
            next_id += 1
        mapping[pid] = f"{next_id:06d}"
        next_id += 1

    print(f"New sequential IDs assigned: 100020 to {next_id - 1}")

    # Duplicate merge: discarded IDs map to the kept ID's new string
    for discard_id, keep_id in discard_to_keep.items():
        if keep_id in mapping:
            mapping[discard_id] = mapping[keep_id]
        else:
            raise ValueError(f"Kept ID {keep_id} not found in mapping!")

    return mapping


def map_single_id(val, mapping):
    """Map a single PersonID value. Returns string or None."""
    if val is None:
        return None
    pid = to_int(val)
    if pid in mapping:
        return mapping[pid]
    else:
        print(f"  WARNING: PersonID {pid} not found in mapping")
        return f"{pid:06d}"


def map_transfer_subject(val, mapping):
    """Map a comma-separated TransferSubject string. Returns string or None."""
    if val is None or str(val).strip() == "":
        return val
    parts = str(val).split(",")
    new_parts = []
    for part in parts:
        part = part.strip()
        if not part:
            new_parts.append(part)
            continue
        try:
            pid = int(float(part))
            if pid in mapping:
                new_parts.append(mapping[pid])
            else:
                print(f"  WARNING: TransferSubject ID {pid} not found in mapping")
                new_parts.append(f"{pid:06d}")
        except ValueError:
            new_parts.append(part)
    return ", ".join(new_parts)


def strip_dagger(val):
    """Remove ‡ (and any surrounding whitespace it creates) from a string."""
    if val is None or not isinstance(val, str):
        return val
    cleaned = val.replace("‡", "").strip()
    # Collapse any double spaces left behind
    cleaned = re.sub(r"  +", " ", cleaned)
    return cleaned


def parse_first_last(full_name):
    """Parse a full name into (first_name, last_name)."""
    if not full_name or not full_name.strip():
        return ("", "")
    parts = full_name.strip().split()
    if len(parts) == 1:
        return (parts[0], "")
    # Last name is the last token, first name is everything before
    return (" ".join(parts[:-1]), parts[-1])


def clean_names(wb):
    """Strip ‡ from all name fields and re-parse First/Last names."""
    # ElectionResults: col 8=Source Name, 9=Name usually known by, 10=First Name, 11=Last Name
    # Also TransferName columns: 17, 22, 27, ... every 5 up to 127
    ws = wb["ElectionResults"]
    er_fixed = 0
    for r in range(2, ws.max_row + 1):
        # Clean Source Name (8) and Name usually known by (9)
        for c in [8, 9]:
            old = ws.cell(r, c).value
            if old and isinstance(old, str) and "‡" in old:
                ws.cell(r, c).value = strip_dagger(old)
                er_fixed += 1

        # Re-parse First Name (10) and Last Name (11) from cleaned Name usually known by (9)
        name_col = ws.cell(r, 9).value
        old_first = ws.cell(r, 10).value
        old_last = ws.cell(r, 11).value
        if (old_first and isinstance(old_first, str) and "‡" in old_first) or \
           (old_last and isinstance(old_last, str) and "‡" in old_last):
            first, last = parse_first_last(strip_dagger(name_col) if name_col else None)
            ws.cell(r, 10).value = first
            ws.cell(r, 11).value = last
            er_fixed += 2

        # Clean TransferName columns (17, 22, 27, ...)
        for c in range(17, 128, 5):
            old = ws.cell(r, c).value
            if old and isinstance(old, str) and "‡" in old:
                ws.cell(r, c).value = strip_dagger(old)
                er_fixed += 1

    print(f"  ElectionResults: {er_fixed} cells cleaned")

    # Transfers: col 8=Name, 16=TransferName
    ws2 = wb["Transfers"]
    t_fixed = 0
    for r in range(2, ws2.max_row + 1):
        for c in [8, 16]:
            old = ws2.cell(r, c).value
            if old and isinstance(old, str) and "‡" in old:
                ws2.cell(r, c).value = strip_dagger(old)
                t_fixed += 1

    print(f"  Transfers: {t_fixed} cells cleaned")
    return er_fixed + t_fixed


def apply_mapping(wb, mapping):
    """Apply the mapping to both sheets."""
    # ElectionResults sheet
    ws = wb["ElectionResults"]
    print(f"\nProcessing ElectionResults ({ws.max_row - 1} rows)...")
    for r in range(2, ws.max_row + 1):
        # PersonID column (132)
        old = ws.cell(r, 132).value
        if old is not None:
            ws.cell(r, 132).value = map_single_id(old, mapping)

        # TransferSubject1-23 (cols 16, 21, 26, 31, ... up to 126)
        for ts_col in range(16, 127, 5):
            old = ws.cell(r, ts_col).value
            if old is not None and str(old).strip():
                ws.cell(r, ts_col).value = map_transfer_subject(old, mapping)

    # Transfers sheet
    ws2 = wb["Transfers"]
    print(f"Processing Transfers ({ws2.max_row - 1} rows)...")
    for r in range(2, ws2.max_row + 1):
        # PersonID column (7)
        old = ws2.cell(r, 7).value
        if old is not None:
            ws2.cell(r, 7).value = map_single_id(old, mapping)

        # TransferSubject column (15)
        old = ws2.cell(r, 15).value
        if old is not None and str(old).strip():
            ws2.cell(r, 15).value = map_transfer_subject(old, mapping)

        # SourcePersonID column (26)
        old = ws2.cell(r, 26).value
        if old is not None:
            ws2.cell(r, 26).value = map_single_id(old, mapping)


def verify(wb, mapping):
    """Verify all PersonIDs are 6-digit strings."""
    ws = wb["ElectionResults"]
    pids = set()
    issues = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 132).value
        if val is not None:
            pids.add(val)
            if not isinstance(val, str) or len(val) != 6:
                issues.append(f"ElectionResults row {r}: PersonID = {val!r} (type={type(val).__name__})")
            try:
                int(val)
            except (ValueError, TypeError):
                issues.append(f"ElectionResults row {r}: PersonID not numeric: {val!r}")

    ws2 = wb["Transfers"]
    t_pids = set()
    for r in range(2, ws2.max_row + 1):
        for col in [7, 26]:
            val = ws2.cell(r, col).value
            if val is not None:
                t_pids.add(val)
                if not isinstance(val, str) or len(val) != 6:
                    issues.append(f"Transfers row {r} col {col}: PersonID = {val!r}")

    all_pids = pids | t_pids
    # Check for any hash IDs remaining
    hash_remaining = [p for p in all_pids if isinstance(p, str) and int(p) > 999999]

    print(f"\n=== Verification ===")
    print(f"Unique PersonIDs in ElectionResults: {len(pids)}")
    print(f"Unique PersonIDs in Transfers: {len(t_pids)}")
    print(f"Combined unique: {len(all_pids)}")
    print(f"Hash IDs remaining (>999999): {len(hash_remaining)}")
    if hash_remaining:
        print(f"  Sample: {hash_remaining[:5]}")
    print(f"Issues found: {len(issues)}")
    for iss in issues[:20]:
        print(f"  {iss}")
    if len(issues) > 20:
        print(f"  ... and {len(issues) - 20} more")

    # All should be exactly 6 chars
    bad_len = [p for p in all_pids if not isinstance(p, str) or len(p) != 6]
    print(f"IDs with wrong length: {len(bad_len)}")
    if bad_len:
        print(f"  Sample: {bad_len[:10]}")

    # Verify no ‡ remains anywhere in name columns
    dagger_remaining = 0
    ws = wb["ElectionResults"]
    for r in range(2, ws.max_row + 1):
        for c in [8, 9, 10, 11] + list(range(17, 128, 5)):
            val = ws.cell(r, c).value
            if val and isinstance(val, str) and "‡" in val:
                dagger_remaining += 1
    ws2 = wb["Transfers"]
    for r in range(2, ws2.max_row + 1):
        for c in [8, 16]:
            val = ws2.cell(r, c).value
            if val and isinstance(val, str) and "‡" in val:
                dagger_remaining += 1
    print(f"Cells still containing ‡: {dagger_remaining}")


def main():
    print(f"Loading {INPUT}...")
    wb = openpyxl.load_workbook(INPUT)

    # Collect all IDs before
    all_ids_before = collect_all_person_ids(wb)
    print(f"\nTotal unique PersonIDs before: {len(all_ids_before)}")

    # Collect just ElectionResults PersonID column for unique count
    ws = wb["ElectionResults"]
    er_pids_before = set()
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 132).value
        if pid is not None:
            er_pids_before.add(to_int(pid))
    print(f"Unique PersonIDs in ElectionResults column before: {len(er_pids_before)}")

    # Clean ‡ from names and re-parse first/last
    print("\nCleaning ‡ from names...")
    total_cleaned = clean_names(wb)
    print(f"Total cells cleaned: {total_cleaned}")

    # Build mapping
    mapping = build_mapping(all_ids_before)
    print(f"Total mapping entries: {len(mapping)}")

    # Count unique output IDs (should be 29 fewer than input after merges)
    output_ids = set(mapping.values())
    print(f"Unique output IDs: {len(output_ids)}")
    print(f"Reduction from merges: {len(all_ids_before) - len(output_ids)}")

    # Apply mapping
    apply_mapping(wb, mapping)

    # Verify
    verify(wb, mapping)

    # Save
    print(f"\nSaving to {OUTPUT}...")
    wb.save(OUTPUT)
    print("Done!")


if __name__ == "__main__":
    main()
