#!/usr/bin/env python
"""Build v7 — ARK-informed splits for genuine same-name collisions."""

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl


def norm(name):
    if not name: return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def normalize_party(p):
    if not p: return ""
    l = p.lower()
    for needle, code in [("sdlp","sdlp"),("democratic unionist","dup"),("dup","dup"),
        ("ulster unionist","uup"),("uup","uup"),("sinn","sf"),("alliance","all"),
        ("independent","ind"),("green","grn"),("tuv","tuv"),("workers","wp"),
        ("pup","pup"),("labour","lab"),("conservative","con"),("socialist","soc")]:
        if needle in l: return code
    return l[:15]


# Load ARK context lookup
ark_candidates = json.load(open("_tmp_ark_candidates.json", encoding="utf-8"))
ark_ctx = {}
for c in ark_candidates:
    full = c["full_name"]
    parts = full.split()
    if len(parts) < 2: continue
    short = f"{parts[0]} {parts[-1]}"
    const = norm(c["constituency"])
    year = c["year"]
    key = (norm(short), const, year)
    ark_ctx[key] = norm(full)


def get_ark_full(name, constituency, year):
    n = norm(name)
    const_n = norm(constituency)
    return ark_ctx.get((n, const_n, year))


# PIDs confirmed as genuine collisions via ARK middle-name analysis + manual review.
# Format: (pid, split_fn) where split_fn(row) -> group_key
# Only includes cases where ARK clearly shows different middle names AND
# party/constituency data supports different people.
CONFIRMED_SPLITS = [
    # Thomas G Murphy (UUP, Coleraine) vs Thomas J Murphy (SDLP, Down)
    ("820015", lambda r: "uup" if normalize_party(r["party"]) == "uup" else "sdlp"),
    # Thomas Edward Black (DUP, Armagh) vs Thomas James Black (Socialist, Belfast)
    # Already partially split in v4 but DUP+Labour/Socialist still share PID
    ("52501", lambda r: "soc" if any(x in (r["party"] or "").lower() for x in ["socialist", "labour coalition"]) else "dup"),
    # William Dunlop Logan (UUP, Ballymoney) vs William John Logan (DUP/Ind/SDLP, various)
    ("60458", lambda r: "sdlp" if normalize_party(r["party"]) == "sdlp" else "other"),
    # John Colville Elliott (DUP, Larne) vs John Gordon Elliott (Alliance, Ballymena)
    ("820743", lambda r: "alliance" if normalize_party(r["party"]) == "all" else "dup"),
    # William Stanley Stevenson vs William Henry Stevenson (different parties/areas)
    ("820994", lambda r: "stanley" if r["year"] and int(r["year"]) <= 1977 else "henry"),
    # Samuel James Martin (DUP/Loyalist/UUP) vs Samuel T Raymond Martin (different person)
    ("820469", lambda r: "raymond" if r["year"] and int(r["year"]) >= 1993 and normalize_party(r["party"]) == "uup" and "Knockiveagh" not in (r["const"] or "") else "james"),
    # William Alexander King (UUP, South Antrim Assembly 1973) vs William Hugh King (UUP, Coleraine/Bann)
    ("85957", lambda r: "alexander" if "South Antrim" in (r["const"] or "") and r["body"] == "Northern Ireland Assembly" else "hugh"),
    # John Patrick Gallagher (SDLP, Strabane/Glenelly) vs John Leo Gallagher (VFYRDTMPH, Belfast)
    ("93383", lambda r: "leo" if "Belfast" in (r["const"] or "") else "patrick"),
    # John Graham Norris (DUP, Castlereagh) vs John Samuel Norris (UUP, Victoria)
    ("12211", lambda r: "uup_victoria" if normalize_party(r["party"]) == "uup" and "Victoria" in (r["const"] or "") else "dup"),
    # Samuel John Hanna (DUP) vs Samuel James Hanna (UUP) — different party in same area
    ("821303", lambda r: "uup" if normalize_party(r["party"]) == "uup" else "dup"),
    # John James Anderson vs John Hugh Anderson — different middle names
    ("824246", lambda r: "group_b" if r["year"] and int(r["year"]) >= 1990 else "group_a"),
    # Robert James Irvine (DUP/UUP, Enniskillen) vs Robert John Irvine (Ind, Holywood)
    # This was already split in v4/v5 — verify
    # John Henry Smyth vs John Gerard Smyth
    ("57287", lambda r: "gerard" if "Workers" in (r["party"] or "") or normalize_party(r["party"]) == "wp" else "henry"),
    # Arthur McG Templeton vs Arthur John Templeton — ARK says different
    # but checking the data, it's one person (UUP/Ind, Newtownabbey/Ballyclare/Antrim Line consistently)
    # SKIP — same person
    # William Blair Smith vs William Plum Smith vs William Woolsey Smith
    ("93842", lambda r: "blair" if r["year"] and int(r["year"]) >= 1989 else "other"),
    # Robert James Wilson vs Robert Alexander Wilson vs Robert Hugh Alexander Wilson
    ("821317", lambda r: "ha" if "Conservative" in (r["party"] or "") else
                          "alexander" if r["year"] and int(r["year"]) <= 1977 else "james"),
    # Samuel James Walker vs Samuel George Walker
    ("821021", lambda r: "george" if "Coleraine" in (r["const"] or "") else "james"),
    # John David O'Neill Wright vs John Robert Wright
    ("30094", lambda r: "robert" if normalize_party(r["party"]) == "grn" or normalize_party(r["party"]) == "ind" and r["year"] and int(r["year"]) >= 1996 else "david"),
    # Desmond Hugh Donnelly vs Desmond Arthur Donnelly
    ("822272", lambda r: "arthur" if r["year"] and int(r["year"]) >= 1997 else "hugh"),
]


def main():
    input_path = Path("Full election tables - comprehensive - personid-v7.xlsx")
    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb["ElectionResults"]

    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))

    pid_col = col["PersonID"]
    rt_col = col["ResultType"]
    name_col = col["Name usually known by"]
    party_col = col["Party Name"]
    date_col = col["Date"]
    body_col = col["ElectedBody"]
    const_col = col["Constituency"]

    max_pid = max(
        (int(str(row[pid_col])) for row in rows
         if row[pid_col] is not None and str(row[pid_col]).isdigit()),
        default=0)
    next_id = max_pid + 1

    # Apply confirmed splits
    print("\nApplying ARK-confirmed splits...")
    split_count = 0
    split_log = []
    for target_pid, key_fn in CONFIRMED_SPLITS:
        groups = defaultdict(list)
        for i, row in enumerate(rows):
            if row[rt_col] != "Candidate":
                continue
            if str(row[pid_col]) != target_pid:
                continue
            r = {
                "name": row[name_col],
                "party": row[party_col],
                "year": str(row[date_col] or "")[:4],
                "body": row[body_col],
                "const": row[const_col],
            }
            key = key_fn(r)
            groups[key].append(i)

        if len(groups) <= 1:
            continue

        group_list = sorted(groups.items())
        for gkey, indices in group_list[1:]:
            new_pid = str(next_id)
            next_id += 1
            for idx in indices:
                rows[idx][pid_col] = new_pid
            split_count += 1
            name_sample = rows[indices[0]][name_col]
            party_sample = rows[indices[0]][party_col]
            split_log.append({
                "pid": target_pid, "group": gkey, "new_pid": new_pid,
                "rows": len(indices), "sample": f"{name_sample} ({party_sample})",
            })

    print(f"  Applied {split_count} splits")
    for s in split_log:
        print(f"    PID {s['pid']} -> {s['new_pid']}: {s['sample']} ({s['rows']} rows)")

    # Update Transfers
    print("\nUpdating Transfers...")
    name_to_pid: dict[str, str] = {}
    for row in rows:
        if row[rt_col] != "Candidate": continue
        name = row[name_col]
        pid = str(row[pid_col] or "")
        if name and pid:
            n = norm(name)
            if n: name_to_pid[n] = pid

    ws_t = wb["Transfers"]
    t_h = [cell.value for cell in ws_t[1]]
    t_pid_idx = t_h.index("PersonID")
    t_name_idx = t_h.index("Name")
    t_upd = 0
    for t_row in ws_t.iter_rows(min_row=2):
        nv = t_row[t_name_idx].value
        if nv:
            n = norm(nv)
            if n in name_to_pid:
                new = name_to_pid[n]
                if str(t_row[t_pid_idx].value or "") != new:
                    t_row[t_pid_idx].value = new
                    t_upd += 1
    print(f"  Updated {t_upd} Transfers rows")

    # Write
    print(f"\nSaving {input_path}...")
    for i, rd in enumerate(rows):
        for j, val in enumerate(rd):
            ws.cell(row=i + 2, column=j + 1, value=val)
    wb.save(input_path)

    # Final stats
    pid_names = defaultdict(set)
    pid_bodies = defaultdict(set)
    pid_dates = defaultdict(set)
    total = 0
    for row in rows:
        if row[rt_col] != "Candidate": continue
        total += 1
        pid = str(row[pid_col])
        pid_names[pid].add(row[name_col] or "")
        pid_bodies[pid].add(row[body_col] or "")
        pid_dates[pid].add(str(row[date_col] or ""))

    cross = sum(1 for bs in pid_bodies.values() if len(bs) > 1)
    multi = sum(1 for ns in pid_names.values() if len(ns) > 1)
    long40 = sum(1 for pid in pid_dates
                 if len([int(d[:4]) for d in pid_dates[pid] if d[:4].isdigit()]) >= 2
                 and max(int(d[:4]) for d in pid_dates[pid] if d[:4].isdigit()) -
                     min(int(d[:4]) for d in pid_dates[pid] if d[:4].isdigit()) > 40
                 and max(sorted(int(d[:4]) for d in pid_dates[pid] if d[:4].isdigit())[i+1] -
                         sorted(int(d[:4]) for d in pid_dates[pid] if d[:4].isdigit())[i]
                         for i in range(len([int(d[:4]) for d in pid_dates[pid] if d[:4].isdigit()]) - 1)) > 25)

    body_counts = defaultdict(int)
    for row in rows:
        if row[rt_col] != "Candidate": continue
        body_counts[row[body_col]] += 1

    print(f"\n{'='*60}")
    print(f"  V7 FINAL STATE")
    print(f"{'='*60}")
    print(f"  Candidate rows: {total}")
    print(f"  Unique PersonIDs: {len(pid_names)}")
    print(f"  Cross-body people: {cross}")
    print(f"  PIDs with name variants: {multi}")
    print(f"  Suspicious long spans: {long40}")
    for body in sorted(body_counts):
        bp = len({p for p, bs in pid_bodies.items() if body in bs})
        print(f"    {body}: {body_counts[body]} candidacies, {bp} people")

    Path("personid_v7_log.json").write_text(json.dumps({
        "splits": split_log,
        "total_splits": split_count,
    }, indent=2, default=str, ensure_ascii=False), encoding="utf-8")
    print(f"\n  Log: personid_v7_log.json")


if __name__ == "__main__":
    main()
