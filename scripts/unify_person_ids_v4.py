#!/usr/bin/env python
"""Unify PersonIDs v4 — starts from the v2 workbook (personid-review.xlsx copy).

v2 already has:
  - Phase 1a: garbage rows removed (good)
  - Phase 1b: T-prefix stripped (good)
  - Phase 1c: exact name merges into curated IDs (good)
  - Phase 2: surname-only fuzzy merges (BAD — caused 218 collisions)
  - Phase 2b: hash PID unification by name (good)
  - Phase 3: fresh IDs assigned (contaminated by Phase 2 errors)

v4 approach:
  Step 1 — Clean name artefacts (trailing dashes, special chars)
  Step 2 — Split ALL collisions (different people sharing a PID)
  Step 3 — Merge via first-name variants (same surname + known variant pair)
  Step 4 — Verify: zero collisions remain
  Step 5 — Assign fresh sequential IDs to any remaining hash PIDs
  Step 6 — Update Transfers sheet

The 15 "merge opportunities" (same name, multiple PIDs) are INTENTIONAL splits
from the existing identity-fix pipeline (e.g. David Taylor Green ≠ David Taylor
UKUP) and must NOT be re-merged.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

# ── First-name variant lookup ─────────────────────────────────────────────

VARIANT_GROUPS: list[list[str]] = [
    ["james","jim","jimmy","jas","seamus","séamas"],
    ["william","will","willy","willie","bill","billy","liam"],
    ["robert","rob","robbie","bob","bobby","bert"],
    ["john","johnny","jack","jock","sean","seán","shane"],
    ["thomas","tom","tommy","thos"],
    ["edward","ed","eddie","eddy","ted","teddy","ned","eamonn","eamon","edmund"],
    ["richard","dick","dickie","rich","rick","ricky"],
    ["charles","charlie","chas","cathal"],
    ["patrick","paddy","pat","patsy","padraig","padraic"],
    ["michael","mike","mick","mickey","mickie","mícheál"],
    ["christopher","chris"],
    ["daniel","dan","danny"],
    ["andrew","andy","drew"],
    ["anthony","tony"],
    ["joseph","joe","joey"],
    ["samuel","sam","sammy"],
    ["benjamin","ben","benny"],
    ["alexander","alex","alec","alistair","alastair","alasdair"],
    ["frederick","fred","freddie","freddy"],
    ["kenneth","ken","kenny"],
    ["ronald","ron","ronnie"],
    ["raymond","ray"],
    ["stephen","steve","steven"],
    ["philip","phil"],
    ["lawrence","larry","laurence"],
    ["matthew","matt"],
    ["gerald","gerry","gerard","gearoid","gearóid"],
    ["peter","pete"],
    ["david","dave","davy","davey"],
    ["donald","don","donnie"],
    ["dennis","denis","den"],
    ["terence","terry","terrence"],
    ["henry","harry","hal"],
    ["albert","bert","bertie","al"],
    ["alfred","alf","alfie"],
    ["leonard","len","lenny"],
    ["bernard","bernie"],
    ["reginald","reg","reggie"],
    ["archibald","archie"],
    ["herbert","herb","herbie"],
    ["geoffrey","geoff","jeff","jeffrey"],
    ["timothy","tim","timmy"],
    ["nicholas","nick","nicky"],
    ["douglas","doug"],
    ["vincent","vince"],
    ["francis","frank","frankie","proinsias"],
    ["desmond","des"],
    ["clifford","cliff"],
    ["roderick","roddy","rod"],
    ["wallace","wally"],
    ["stanley","stan"],
    ["ernest","ernie"],
    ["norman","norm"],
    ["percy","percival"],
    ["elizabeth","liz","lizzie","beth","betty","bess","bessie","eliza"],
    ["margaret","maggie","meg","peggy","madge","marge","mags"],
    ["catherine","kate","katie","cathy","kathleen","katharine","katherine"],
    ["patricia","pat","patsy","tricia","trish"],
    ["jennifer","jenny","jen"],
    ["dorothy","dot","dolly","dora"],
    ["pamela","pam"],
    ["susan","sue","susie","susanne"],
    ["deborah","deb","debbie"],
    ["christine","chris","christina","tina"],
    ["jacqueline","jackie"],
    ["rosemary","rosie"],
    ["caroline","carol"],
    ["joanna","jo","joanne","joan"],
    ["anne","ann","annie"],
    ["mary","molly","may","maire"],
    ["eleanor","ella","ellie","nell","nellie"],
    ["bridget","brid","bridie"],
    ["eileen","aileen"],
    ["niall","neal","neil"],
    ["ciaran","kieran"],
    ["jj","j j","j. j.","j.j."],
]

_VARIANT_MAP: dict[str, int] = {}
for _gi, _grp in enumerate(VARIANT_GROUPS):
    for _n in _grp:
        _VARIANT_MAP[_n.lower()] = _gi


def norm(name: str) -> str:
    if not name: return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def split_fn_sn(name: str) -> tuple[str, str]:
    parts = name.strip().split()
    if not parts: return ("", "")
    if len(parts) == 1: return ("", norm(parts[0]))
    return (norm(" ".join(parts[:-1])), norm(parts[-1]))


def are_fn_variants(fn1: str, fn2: str) -> bool:
    n1, n2 = fn1.lower().strip(), fn2.lower().strip()
    if n1 == n2: return True
    g1, g2 = _VARIANT_MAP.get(n1), _VARIANT_MAP.get(n2)
    if g1 is not None and g2 is not None and g1 == g2: return True
    if len(n1) >= 3 and len(n2) >= 3 and (n1.startswith(n2) or n2.startswith(n1)): return True
    return False


def names_same_person(n1: str, n2: str) -> bool:
    if norm(n1) == norm(n2): return True
    fn1, sn1 = split_fn_sn(n1)
    fn2, sn2 = split_fn_sn(n2)
    if sn1 != sn2: return False
    return are_fn_variants(fn1, fn2)


def pid_type(pid: str) -> str:
    if not pid: return "none"
    if pid.startswith("T"): return "temp"
    if pid.isdigit(): return "curated" if len(pid) <= 6 else "hash"
    return "other"


def is_garbage(name: str) -> bool:
    if not name or not name.strip(): return True
    c = name.strip().lower().strip("() ")
    return c in {"party","(politician)","ireland","voice","fein","fáin","ecology",
                 "clubs","labour","conservative"} or len(name.strip()) <= 2


# ── Known intentional splits (do NOT merge these) ────────────────────────

INTENTIONAL_SPLIT_PIDS: set[str] = {
    "100001","100002","100003","100004","100005","100006","100007",
    "100008","100009","100010","100011","100012","100013","100014",
    "100015","100016","100017","100018","100019",
}


def main() -> None:
    input_path = Path("Full election tables - comprehensive - personid-v4.xlsx")
    log_path = Path("personid_v4_log.json")

    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb["ElectionResults"]

    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))

    C = {k: col[k] for k in ["PersonID","ResultType","Name usually known by",
         "Party Name","Date","ElectedBody","Constituency","Council","Source Name"]}
    log: dict[str, Any] = {"steps": {}}

    # ── Step 1: Clean name artefacts ──────────────────────────────────────
    print("\nStep 1: Cleaning name artefacts...")
    cleaned = 0
    for i, row in enumerate(rows):
        for c_idx in [C["Name usually known by"], C["Source Name"]]:
            name = row[c_idx]
            if name and isinstance(name, str):
                new = re.sub(r"\s*[–—\-]\s*$", "", name).strip()
                new = re.sub(r"\s*[♭†‡*]+\s*$", "", new).strip()
                if new != name:
                    rows[i][c_idx] = new
                    cleaned += 1
    # Also mark any remaining garbage
    garbage = 0
    for i, row in enumerate(rows):
        if row[C["ResultType"]] == "Candidate" and is_garbage(row[C["Name usually known by"]]):
            rows[i][C["ResultType"]] = "GarbageRemoved"
            garbage += 1
    log["steps"]["1_clean"] = {"artefacts_cleaned": cleaned, "garbage_marked": garbage}
    print(f"  Cleaned {cleaned} artefacts, marked {garbage} additional garbage")

    # ── Step 2: Split ALL collisions ──────────────────────────────────────
    print("\nStep 2: Splitting collisions...")

    # Build PID → name groups (cluster names by same-person)
    pid_name_groups: dict[str, list[tuple[str, list[int]]]] = defaultdict(list)
    for i, row in enumerate(rows):
        if row[C["ResultType"]] != "Candidate": continue
        pid = str(row[C["PersonID"]] or "")
        name = row[C["Name usually known by"]] or ""
        if not pid or not name: continue
        found = False
        for gname, gidxs in pid_name_groups[pid]:
            if names_same_person(name, gname):
                gidxs.append(i)
                found = True
                break
        if not found:
            pid_name_groups[pid].append((name, [i]))

    # Find max PID for fresh assignments
    max_pid = max((int(str(row[C["PersonID"]])) for row in rows
                   if row[C["PersonID"]] is not None and str(row[C["PersonID"]]).isdigit()),
                  default=0)
    next_id = max(max_pid + 1, 300001)

    split_count = 0
    split_log: list[dict] = []
    for pid, groups in pid_name_groups.items():
        if len(groups) <= 1: continue
        # Multiple person-groups share this PID — collision
        for gname, gidxs in groups[1:]:
            new_pid = str(next_id)
            next_id += 1
            for idx in gidxs:
                rows[idx][C["PersonID"]] = new_pid
            split_count += 1
            split_log.append({
                "original_pid": pid,
                "kept": groups[0][0],
                "split_name": gname,
                "new_pid": new_pid,
                "rows": len(gidxs),
            })

    log["steps"]["2_split"] = {"collisions_split": split_count}
    print(f"  Split {split_count} collision groups")

    # ── Step 3: Merge via first-name variants ─────────────────────────────
    print("\nStep 3: Merging via first-name variants...")

    # Build surname → [(curated_pid, fn, display_name)] from curated-ID rows
    surname_idx: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    seen_curated: set[tuple[str, str, str]] = set()
    for row in rows:
        if row[C["ResultType"]] != "Candidate": continue
        pid = str(row[C["PersonID"]] or "")
        if pid_type(pid) != "curated": continue
        if pid in INTENTIONAL_SPLIT_PIDS: continue
        name = row[C["Name usually known by"]] or ""
        fn, sn = split_fn_sn(name)
        if sn and fn:
            key = (pid, sn, fn)
            if key not in seen_curated:
                surname_idx[sn].append((pid, fn, name))
                seen_curated.add(key)

    merged = 0
    merge_log: list[dict] = []
    for i, row in enumerate(rows):
        if row[C["ResultType"]] != "Candidate": continue
        pid = str(row[C["PersonID"]] or "")
        if pid_type(pid) == "curated": continue
        name = row[C["Name usually known by"]] or ""
        fn, sn = split_fn_sn(name)
        if not sn or not fn: continue

        candidates = surname_idx.get(sn, [])
        matches = [(cpid, cfn, cname) for cpid, cfn, cname in candidates
                    if are_fn_variants(fn, cfn)]
        matched_pids = set(cpid for cpid, _, _ in matches)

        if len(matched_pids) == 1:
            target = matched_pids.pop()
            old = pid
            rows[i][C["PersonID"]] = target
            merged += 1
            if len(merge_log) < 200:
                merge_log.append({"name": name, "old_pid": old, "new_pid": target,
                                  "matched_to": matches[0][2]})

    log["steps"]["3_variant_merge"] = {"merged": merged}
    print(f"  Merged {merged} rows via first-name variants")

    # ── Step 3b: Second pass after merges expanded the curated index ──────
    print("\nStep 3b: Second variant-merge pass...")
    # Rebuild index
    surname_idx2: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    seen2: set[tuple[str, str, str]] = set()
    for row in rows:
        if row[C["ResultType"]] != "Candidate": continue
        pid = str(row[C["PersonID"]] or "")
        if pid_type(pid) != "curated": continue
        if pid in INTENTIONAL_SPLIT_PIDS: continue
        name = row[C["Name usually known by"]] or ""
        fn, sn = split_fn_sn(name)
        if sn and fn:
            key = (pid, sn, fn)
            if key not in seen2:
                surname_idx2[sn].append((pid, fn, name))
                seen2.add(key)

    merged2 = 0
    for i, row in enumerate(rows):
        if row[C["ResultType"]] != "Candidate": continue
        pid = str(row[C["PersonID"]] or "")
        if pid_type(pid) == "curated": continue
        name = row[C["Name usually known by"]] or ""
        fn, sn = split_fn_sn(name)
        if not sn or not fn: continue
        candidates = surname_idx2.get(sn, [])
        matches = [(cpid, cfn, cname) for cpid, cfn, cname in candidates
                    if are_fn_variants(fn, cfn)]
        matched_pids = set(cpid for cpid, _, _ in matches)
        if len(matched_pids) == 1:
            rows[i][C["PersonID"]] = matched_pids.pop()
            merged2 += 1

    log["steps"]["3b_pass2"] = {"merged": merged2}
    print(f"  Merged {merged2} more rows")

    # ── Step 4: Unify remaining hash IDs sharing a normalised name ────────
    print("\nStep 4: Unifying hash IDs by normalised name...")
    norm_to_hashes: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        if row[C["ResultType"]] != "Candidate": continue
        pid = str(row[C["PersonID"]] or "")
        if pid_type(pid) != "hash": continue
        n = norm(row[C["Name usually known by"]] or "")
        if n: norm_to_hashes[n].add(pid)

    unify: dict[str, str] = {}
    for n, pids in norm_to_hashes.items():
        if len(pids) <= 1: continue
        canonical = sorted(pids)[0]
        for p in pids:
            if p != canonical: unify[p] = canonical

    unified = 0
    for i, row in enumerate(rows):
        pid = str(row[C["PersonID"]] or "")
        if pid in unify:
            rows[i][C["PersonID"]] = unify[pid]
            unified += 1

    log["steps"]["4_unify_hash"] = {"unified_pids": len(unify), "rows": unified}
    print(f"  Unified {len(unify)} hash PIDs ({unified} rows)")

    # ── Step 5: Verify zero collisions ────────────────────────────────────
    print("\nStep 5: Verification...")
    pid_name_check: dict[str, list[tuple[str, list[int]]]] = defaultdict(list)
    for i, row in enumerate(rows):
        if row[C["ResultType"]] != "Candidate": continue
        pid = str(row[C["PersonID"]] or "")
        name = row[C["Name usually known by"]] or ""
        if not pid or not name: continue
        found = False
        for gname, gidxs in pid_name_check[pid]:
            if names_same_person(name, gname):
                gidxs.append(i)
                found = True
                break
        if not found:
            pid_name_check[pid].append((name, [i]))

    remaining_collisions = sum(1 for groups in pid_name_check.values() if len(groups) > 1)
    if remaining_collisions:
        print(f"  WARNING: {remaining_collisions} remaining collisions — splitting...")
        for pid, groups in pid_name_check.items():
            if len(groups) <= 1: continue
            for gname, gidxs in groups[1:]:
                new_pid = str(next_id)
                next_id += 1
                for idx in gidxs:
                    rows[idx][C["PersonID"]] = new_pid
                split_count += 1
        print(f"  Split {remaining_collisions} more collision groups")
    else:
        print(f"  Zero collisions — clean")

    # ── Step 6: Assign fresh IDs to remaining hash PIDs ───────────────────
    print("\nStep 6: Assigning fresh IDs...")
    hash_to_fresh: dict[str, str] = {}
    for i, row in enumerate(rows):
        if row[C["ResultType"]] != "Candidate": continue
        pid = str(row[C["PersonID"]] or "")
        if pid_type(pid) != "hash": continue
        if pid not in hash_to_fresh:
            hash_to_fresh[pid] = str(next_id)
            next_id += 1
        rows[i][C["PersonID"]] = hash_to_fresh[pid]

    log["steps"]["6_fresh"] = {"fresh_ids": len(hash_to_fresh),
                               "range": f"{min(int(v) for v in hash_to_fresh.values()) if hash_to_fresh else 0}-{next_id-1}"}
    print(f"  Assigned {len(hash_to_fresh)} fresh IDs")

    # ── Step 7: Update Transfers sheet ────────────────────────────────────
    print("\nStep 7: Updating Transfers sheet...")
    name_to_pid: dict[str, str] = {}
    for row in rows:
        if row[C["ResultType"]] != "Candidate": continue
        name = row[C["Name usually known by"]]
        pid = str(row[C["PersonID"]] or "")
        if name and pid:
            n = norm(name)
            if n: name_to_pid[n] = pid

    ws_t = wb["Transfers"]
    t_h = [cell.value for cell in ws_t[1]]
    t_pid = t_h.index("PersonID") if "PersonID" in t_h else None
    t_name = t_h.index("Name") if "Name" in t_h else None
    t_upd = 0
    if t_pid is not None and t_name is not None:
        for t_row in ws_t.iter_rows(min_row=2):
            nv = t_row[t_name].value
            if nv:
                n = norm(nv)
                if n in name_to_pid:
                    new = name_to_pid[n]
                    if str(t_row[t_pid].value or "") != new:
                        t_row[t_pid].value = new
                        t_upd += 1
    print(f"  Updated {t_upd} Transfers rows")

    # ── Write ─────────────────────────────────────────────────────────────
    print(f"\nWriting {input_path}...")
    for i, rd in enumerate(rows):
        for j, val in enumerate(rd):
            ws.cell(row=i+2, column=j+1, value=val)
    wb.save(input_path)

    # ── Final stats ───────────────────────────────────────────────────────
    final_pids = set()
    final_names = set()
    pid_multi = defaultdict(set)
    type_counts = defaultdict(int)
    body_counts = defaultdict(int)
    for row in rows:
        if row[C["ResultType"]] != "Candidate": continue
        pid = str(row[C["PersonID"]] or "")
        name = row[C["Name usually known by"]] or ""
        type_counts[pid_type(pid)] += 1
        final_pids.add(pid)
        if name:
            final_names.add(name)
            pid_multi[pid].add(name)
        body_counts[row[C["ElectedBody"]]] += 1

    multi_name_count = sum(1 for names in pid_multi.values() if len(names) > 1)
    cross_body = sum(1 for row_groups in pid_name_check.values()
                     for _ in row_groups if len(row_groups) > 0) # recount
    # Better cross-body count
    pid_bodies: dict[str, set] = defaultdict(set)
    for row in rows:
        if row[C["ResultType"]] != "Candidate": continue
        pid_bodies[str(row[C["PersonID"]] or "")].add(row[C["ElectedBody"]])
    cross = sum(1 for bs in pid_bodies.values() if len(bs) > 1)

    log["summary"] = {
        "total_candidates": sum(type_counts.values()),
        "unique_pids": len(final_pids),
        "unique_names": len(final_names),
        "cross_body_people": cross,
        "pids_with_variants": multi_name_count,
        "pid_types": dict(type_counts),
    }

    print(f"\n{'='*60}")
    print(f"  V4 FINAL SUMMARY")
    print(f"{'='*60}")
    print(f"  Candidate rows: {sum(type_counts.values())}")
    print(f"  Unique PersonIDs: {len(final_pids)}")
    print(f"  Unique names: {len(final_names)}")
    print(f"  Cross-body people: {cross}")
    print(f"  PIDs with name variants: {multi_name_count}")
    print(f"  PID types: {dict(type_counts)}")
    for body in sorted(body_counts):
        bp = len({p for p, bs in pid_bodies.items() if body in bs})
        print(f"    {body}: {body_counts[body]} candidacies, {bp} people")

    log_path.write_text(json.dumps(log, indent=2, default=str, ensure_ascii=False), encoding="utf-8")
    print(f"\n  Log: {log_path}")


if __name__ == "__main__":
    main()
