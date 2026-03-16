#!/usr/bin/env python
"""Build STV-only ElectionResults / Transfers workbooks from raw spreadsheets."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import xlrd


ELECTION_RESULTS_HEADERS = [
    "Date",
    "Event",
    "EventType",
    "ElectedBody",
    "Source Party Name",
    "Deduplicated Party Name",
    "Wikipedia Party Name",
    "ResultType",
    "Party Name",
    "Source Name",
    "Name usually known by",
    "First Name",
    "Last Name",
    "Constituency",
    "Council",
]

for i in range(1, 24):
    ELECTION_RESULTS_HEADERS.extend(
        [
            f"Votes{i}",
            f"Transfers{i}",
            f"TransferSubject{i}",
            f"TransferName{i}",
            f"TransferParty{i}",
        ]
    )

ELECTION_RESULTS_HEADERS.extend(
    [
        "Outcome",
        "%ValidShare",
        "%ElectorateShare",
        "PersonID",
        "DevolvedInstance",
        "WMInstance",
        "EUInstance",
        "TotalInstance",
        "Gender",
        "ElectionKey",
    ]
)

TRANSFERS_HEADERS = [
    "Date",
    "Event",
    "Constituency",
    "Council",
    "ElectedBody",
    "ResultType",
    "PersonID",
    "Name",
    "Party",
    "Deduplicated Party Name",
    "Wikipedia Party Name",
    "Count",
    "Votes",
    "Transfers",
    "TransferSubject",
    "TransferName",
    "TransferParty",
    "TransferPct",
    "EliminatedThisRound",
    "ElectedThisRound",
    "TransferPartyRelation",
    "RemainingCandidateIDsDesc",
    "RemainingCandidateNamesInIDOrder",
    "RemainingCandidatePartiesAZ",
    "RemainingCandidatePartiesInIDOrder",
    "SourcePersonID",
]

LGOV_DATE_FALLBACKS = {
    "1973": "1973-05-30",
    "1977": "1977-05-18",
    "1981": "1981-05-20",
    "1985": "1985-05-15",
    "1989": "1989-05-17",
    "1993": "1993-05-19",
    "1997": "1997-05-21",
    "2001": "2001-06-07",
    "2005": "2005-05-05",
    "2011": "2011-05-05",
}

LGOV_COUNCIL_BY_CODE = {
    "ANT": "Antrim",
    "ARD": "Ards",
    "ARM": "Armagh",
    "BMA": "Ballymena",
    "BMY": "Ballymoney",
    "BRG": "Banbridge",
    "BT": "Belfast",
    "CAR": "Carrickfergus",
    "CAS": "Castlereagh",
    "COL": "Coleraine",
    "COO": "Cookstown",
    "DE": "Derry",
    "DOW": "Down",
    "DUN": "Dungannon and South Tyrone",
    "FER": "Fermanagh",
    "LAR": "Larne",
    "LIM": "Limavady",
    "LIS": "Lisburn",
    "MAG": "Magherafelt",
    "MOY": "Moyle",
    "NAM": "Newry and Mourne",
    "NEW": "Newtownabbey",
    "NOD": "North Down",
    "OMA": "Omagh",
    "STR": "Strabane",
    "CRA": "Craigavon",
}

EURO_DATE_FALLBACKS = {
    "1979": "1979-06-07",
    "1984": "1984-06-14",
    "1989": "1989-06-15",
    "1994": "1994-06-09",
    "1999": "1999-06-10",
    "2004": "2004-06-10",
    "2009": "2009-06-04",
}

ASBY_DATE_FALLBACKS = {
    "1973": "1973-06-28",
    "1982": "1982-10-20",
    "1998": "1998-06-25",
    "2003": "2003-11-26",
    "2007": "2007-03-07",
    "2011": "2011-05-05",
}

CONV_DATE_FALLBACKS = {"1975": "1975-05-01"}


@dataclass
class Stage:
    index: int
    raw_header: str


@dataclass
class Candidate:
    marker: str
    number: str
    raw_name: str
    source_party: str
    party: str
    first_pref: float | None
    deltas: list[float | None] = field(default_factory=list)
    totals: list[float | None] = field(default_factory=list)
    person_id: int | None = None


@dataclass
class Contest:
    family: str
    date: str
    event: str
    event_type: str
    elected_body: str
    constituency: str
    council: str | None
    election_key: str
    seats: int | None
    electorate: float | None
    votes_polled: float | None
    valid_votes: float | None
    invalid_votes: float | None
    quota: float | None
    percent_poll: float | None
    stages: list[Stage]
    candidates: list[Candidate]
    nontransferable: Candidate | None
    source_path: str


@dataclass
class PartyNormalizationLookup:
    by_raw: dict[str, tuple[str | None, str | None]] = field(default_factory=dict)
    by_canonical: dict[str, tuple[str | None, str | None]] = field(default_factory=dict)

    def resolve(self, raw_party: str | None) -> tuple[str | None, str | None]:
        if not raw_party:
            return (None, None)
        raw_key = normalize_space(raw_party)
        if raw_key in self.by_raw:
            return self.by_raw[raw_key]
        canonical_key = canonical_label(raw_party)
        return self.by_canonical.get(canonical_key, (None, None))


def normalize_space(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\u2019", "'").replace("\u2018", "'").replace("\u2014", "-")
    return re.sub(r"\s+", " ", text).strip()


def canonical_label(text: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", normalize_space(text).lower()).strip()


def value_is_blank(value: Any) -> bool:
    if value is None:
        return True
    text = normalize_space(value)
    return text in {"", "-", "\u2014", "\u2013", "—", "–"}


def parse_numeric(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = normalize_space(value)
    if text in {"", "-", "\u2014", "\u2013", "—", "–"}:
        return None
    text = text.replace(",", "").replace("%", "").replace("'", "")
    try:
        return float(text)
    except ValueError:
        return None


def excel_serial_to_date(value: float) -> dt.date | None:
    if value is None:
        return None
    if value < 20000 or value > 60000:
        return None
    base = dt.datetime(1899, 12, 30)
    try:
        return (base + dt.timedelta(days=float(value))).date()
    except OverflowError:
        return None


def coerce_date_string(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    numeric = parse_numeric(value)
    date_value = excel_serial_to_date(numeric) if numeric is not None else None
    if date_value:
        return date_value.isoformat()
    text = normalize_space(value)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def family_for_path(path: Path) -> str:
    parts = {part.lower() for part in path.parts}
    for family in ("asby", "conv", "euro", "lgov"):
        if family in parts:
            return family
    raise ValueError(f"Unable to determine family from path: {path}")


def family_metadata(family: str) -> tuple[str, str]:
    if family in {"asby", "conv"}:
        return ("DevolvedElection", "GeneralElection")
    if family == "euro":
        return ("EuropeanElection", "GeneralElection")
    if family == "lgov":
        return ("LocalGovernmentElection", "GeneralElection")
    raise ValueError(f"Unsupported family: {family}")


def elected_body_for_family(family: str) -> str:
    return {
        "asby": "Northern Ireland Assembly",
        "conv": "Northern Ireland Constitutional Convention",
        "euro": "European Parliament",
        "lgov": "Local Government",
    }[family]


def fallback_date_for_path(path: Path, family: str) -> str:
    text = str(path)
    year_match = re.search(r"(19\d{2}|20\d{2})", text)
    year = year_match.group(1) if year_match else None
    if year is None:
        stem = path.stem.lower()
        short_match = None
        if family == "euro":
            short_match = re.search(r"eu(\d{2})", stem)
        elif family == "asby":
            short_match = re.search(r"as(\d{2})", stem)
        elif family == "conv":
            short_match = re.search(r"cc(\d{2})", stem)
        elif family == "lgov":
            short_match = re.search(r"lg(\d{2})", stem)
        if short_match:
            yy = int(short_match.group(1))
            year = str(1900 + yy if yy >= 70 else 2000 + yy)
    if family == "lgov" and year in LGOV_DATE_FALLBACKS:
        return LGOV_DATE_FALLBACKS[year]
    if family == "euro" and year in EURO_DATE_FALLBACKS:
        return EURO_DATE_FALLBACKS[year]
    if family == "asby" and year in ASBY_DATE_FALLBACKS:
        return ASBY_DATE_FALLBACKS[year]
    if family == "conv" and year in CONV_DATE_FALLBACKS:
        return CONV_DATE_FALLBACKS[year]
    if year:
        return f"{year}-01-01"
    raise ValueError(f"No fallback date available for {path}")


def source_year_hint(path: Path) -> int | None:
    text = str(path)
    year_match = re.search(r"(19\d{2}|20\d{2})", text)
    if year_match:
        return int(year_match.group(1))
    stem = path.stem.lower()
    for prefix in ("eu", "as", "cc", "lg"):
        match = re.search(rf"{prefix}(\d{{2}})", stem)
        if match:
            yy = int(match.group(1))
            return 1900 + yy if yy >= 70 else 2000 + yy
    return None


def load_sheet_rows(path: Path) -> list[list[Any]]:
    if path.suffix.lower() == ".xlsx":
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        return [[cell.value for cell in row] for row in ws.iter_rows()]
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    return [sh.row_values(i, 0, sh.ncols) for i in range(sh.nrows)]


def file_preference_score(path: Path) -> tuple[int, int]:
    stem = path.stem.lower()
    return (
        1 if stem.endswith("-corrected") else 0,
        1 if path.suffix.lower() == ".xlsx" else 0,
    )


def preferred_stv_files(root: Path, families: Iterable[str]) -> list[Path]:
    families = set(families)
    chosen: dict[tuple[str, str], Path] = {}
    for family in families:
        for path in sorted((root / family).rglob("*")):
            if path.suffix.lower() not in {".xls", ".xlsx"}:
                continue
            if "overall-results" in path.stem.lower():
                continue
            stem = path.stem.lower().replace("-corrected", "")
            key = (family, stem)
            current = chosen.get(key)
            if current is None or file_preference_score(path) > file_preference_score(current):
                chosen[key] = path
    return sorted(chosen.values())


def find_header_row(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows[:8]):
        joined = " | ".join(normalize_space(v) for v in row if not value_is_blank(v)).lower()
        if (
            "stage 1" in joined
            or "stage no. 1" in joined
            or "stage no 1" in joined
            or "1 stage" in joined
            or "first stage" in joined
            or "1st stage" in joined
            or "1st pref" in joined
        ):
            return idx
        if "candidates" in joined and "description" in joined:
            return idx
    raise ValueError("Unable to find STV header row")


def find_first_stage_col(rows: list[list[Any]], header_row: int) -> int:
    row = rows[header_row]
    for idx, cell in enumerate(row):
        text = normalize_space(cell).lower()
        if (
            "stage 1" in text
            or "stage no. 1" in text
            or "stage no 1" in text
            or "1 stage" in text
            or "first stage" in text
            or "1st stage" in text
            or text == "1st pref"
        ):
            return idx
    second = rows[header_row + 1] if header_row + 1 < len(rows) else []
    for idx, cell in enumerate(second):
        text = normalize_space(cell).lower()
        if "first preference" in text or "1st pref" in text:
            return idx
    return 4 if len(row) > 4 else max(0, len(row) - 1)


def extract_stages(rows: list[list[Any]], header_row: int, first_stage_col: int) -> list[Stage]:
    stages = [Stage(index=1, raw_header="First Preference Votes")]
    first_header = normalize_space(rows[header_row][first_stage_col] if first_stage_col < len(rows[header_row]) else "")
    if not first_header and (header_row + 1 >= len(rows) or value_is_blank(rows[header_row + 1][first_stage_col] if first_stage_col < len(rows[header_row + 1]) else None)):
        return stages
    idx = first_stage_col + 1
    stage_index = 2
    while idx + 1 < len(rows[header_row]):
        raw_left = normalize_space(rows[header_row][idx] if idx < len(rows[header_row]) else "")
        raw_right = normalize_space(rows[header_row][idx + 1] if idx + 1 < len(rows[header_row]) else "")
        raw_header = raw_left or raw_right
        if not raw_header:
            break
        stages.append(Stage(index=stage_index, raw_header=raw_header))
        stage_index += 1
        idx += 2
    return stages


def flatten_row(row: list[Any], width: int) -> list[Any]:
    if len(row) < width:
        return row + [None] * (width - len(row))
    return row[:width]


def find_metadata_start(rows: list[list[Any]], header_row: int, first_stage_col: int) -> int:
    width = max(len(row) for row in rows) if rows else 0
    for idx in range(header_row + 3, len(rows)):
        row = flatten_row(rows[idx], width)
        row_text = " | ".join(normalize_space(v) for v in row if not value_is_blank(v)).lower()
        if any(
            token in row_text
            for token in (
                "eligible electorate",
                "electorate",
                "no of electors",
                "number to be elected",
                "no. to be elected",
                "quota",
                "votes polled",
                "valid votes",
                "invalid votes",
                "% poll",
                "constituency of",
                "district electorial area",
                "district electoral area",
            )
        ):
            return idx
    return len(rows)


def compact_candidate_rows(rows: list[list[Any]], header_row: int, metadata_start: int, first_stage_col: int) -> list[list[Any]]:
    width = max(len(row) for row in rows) if rows else 0
    compacted: list[list[Any]] = []
    last_candidate_idx: int | None = None
    for idx in range(header_row + 2, metadata_start):
        row = flatten_row(rows[idx], width)
        marker = normalize_space(row[0] if len(row) > 0 else "")
        number = row[1] if len(row) > 1 else None
        name = normalize_space(row[2] if len(row) > 2 else "")
        desc = normalize_space(row[3] if len(row) > 3 else "")
        numeric_cells = [parse_numeric(cell) for cell in row[first_stage_col:]]
        has_numeric = any(value is not None for value in numeric_cells)
        label_blob = " ".join(filter(None, [name, desc])).lower()
        if "non-transferable" in label_blob or "non transferable" in label_blob or label_blob == "totals" or label_blob == "total":
            compacted.append(row)
            last_candidate_idx = None
            continue
        if name and not has_numeric and not desc and last_candidate_idx is not None:
            base = compacted[last_candidate_idx]
            base[2] = normalize_space(f"{base[2]} {name}")
            continue
        if not name and not desc and not has_numeric and not marker and value_is_blank(number):
            continue
        if name or has_numeric or desc:
            compacted.append(row)
            if name and desc != "Non-transferable":
                last_candidate_idx = len(compacted) - 1
    return compacted


def normalize_party(value: str) -> str:
    text = normalize_space(value)
    lowered = text.lower()
    if not text:
        return ""
    mappings = [
        ("social democratic and labour party", "SDLP"),
        ("democratic unionist party", "DUP"),
        ("dermocratic unionist", "DUP"),
        ("official unionist party", "UUP"),
        ("ulster unionist party", "UUP"),
        ("ulster conservative and unionists - new force", "UUP"),
        ("sinn fein", "Sinn Féin"),
        ("sinn féin", "Sinn Féin"),
        ("alliance party", "Alliance"),
        ("green party", "Green / Ecology"),
        ("green / ecology", "Green / Ecology"),
        ("independent", "Independent"),
        ("traditional unionist voice-tuv", "TUV"),
        ("traditional unionist voice", "TUV"),
        ("tuv", "TUV"),
    ]
    for needle, replacement in mappings:
        if needle in lowered:
            return replacement
    return text


def split_name(raw_name: str) -> tuple[str, str, str]:
    text = normalize_space(raw_name)
    if not text:
        return ("", "", "")
    if "," in text:
        last, rest = [normalize_space(part) for part in text.split(",", 1)]
        first = normalize_space(rest)
        display = normalize_space(f"{first} {last}")
        return (display, first, last)
    parts = text.split()
    if len(parts) == 1:
        return (text, "", text)
    return (text, " ".join(parts[:-1]), parts[-1])


class PersonRegistry:
    def __init__(self) -> None:
        self._ids: dict[str, int] = {}

    def get(self, raw_name: str) -> int:
        key = normalize_space(raw_name).lower()
        if key not in self._ids:
            digest = hashlib.md5(key.encode("utf-8")).hexdigest()
            self._ids[key] = int(digest[:8], 16)
        return self._ids[key]


def extract_metadata(rows: list[list[Any]], metadata_start: int, path: Path, family: str) -> dict[str, Any]:
    meta: dict[str, Any] = {
        "date": None,
        "constituency": None,
        "council": None,
        "electorate": None,
        "votes_polled": None,
        "seats": None,
        "valid_votes": None,
        "invalid_votes": None,
        "quota": None,
        "percent_poll": None,
    }
    scan_start = max(0, metadata_start - 2)
    for row in rows[scan_start:]:
        text_cells = [normalize_space(cell) for cell in row if not value_is_blank(cell)]
        if not text_cells:
            continue
        row_norm = [normalize_space(cell) for cell in row]
        row_canon = [canonical_label(cell) for cell in row]
        if meta["date"] is None:
            for cell in row[:3]:
                maybe_date = coerce_date_string(cell)
                if maybe_date:
                    meta["date"] = maybe_date
                    break
        joined = " | ".join(text_cells)
        lowered = joined.lower()
        if meta["constituency"] is None:
            for cell in row_norm[:4]:
                low = canonical_label(cell)
                if low.startswith("constituency of "):
                    meta["constituency"] = normalize_space(cell.split("of", 1)[1])
                    break
                if low.startswith("constituency -"):
                    meta["constituency"] = normalize_space(cell.split("-", 1)[1])
                    break
                if "district electorial area:" in low or "district electoral area:" in low:
                    meta["constituency"] = normalize_space(cell.split(":", 1)[1])
                    break
            if "constituency -" in lowered:
                match = re.search(r"constituency\s*-\s*(.+)", joined, re.I)
                if match:
                    meta["constituency"] = normalize_space(match.group(1).split("|", 1)[0])
            elif "constituency of" in lowered:
                match = re.search(r"constituency of\s+(.+)", joined, re.I)
                if match:
                    meta["constituency"] = normalize_space(match.group(1).split("|", 1)[0])
            elif "district electorial area" in lowered or "district electoral area" in lowered:
                match = re.search(r"district elect\w* area:\s*(.+)", joined, re.I)
                if match:
                    meta["constituency"] = normalize_space(match.group(1).split("|", 1)[0])
        label_value_pairs = [
            ("electorate", ("eligible electorate", "eligible electroate", "no of electors", "number of electors"), 4),
            ("votes_polled", ("votes polled", "total votes polled"), 4),
            ("seats", ("number to be elected", "number of members to be elected", "no. to be elected"), 4),
            ("valid_votes", ("total valid votes", "valid votes", "vaild votes"), 4),
            ("invalid_votes", ("invalid votes",), 4),
            ("quota", ("quota", "electoral quota"), 4),
        ]
        for meta_key, tokens, numeric_col in label_value_pairs:
            if meta[meta_key] is not None:
                continue
            label_index = None
            canon_tokens = [canonical_label(token) for token in tokens]
            for idx, low in enumerate(row_canon):
                if any(token in low for token in canon_tokens):
                    label_index = idx
                    break
            if label_index is not None:
                candidate_value_index = min(label_index + 1, len(row) - 1) if row else 0
                if len(row) > candidate_value_index:
                    meta[meta_key] = parse_numeric(row[candidate_value_index])
                if meta[meta_key] is None and len(row) > numeric_col:
                    meta[meta_key] = parse_numeric(row[numeric_col])
                if meta[meta_key] is None:
                    for cell in reversed(row):
                        numeric = parse_numeric(cell)
                        if numeric is not None:
                            meta[meta_key] = numeric
                            break
        if meta["percent_poll"] is None:
            percent_labels = {"poll", "percent poll"}
            label_index = next((idx for idx, low in enumerate(row_canon) if low in percent_labels), None)
            if label_index is not None:
                candidate_value_index = min(label_index + 1, len(row) - 1) if row else 0
                if len(row) > candidate_value_index:
                    meta["percent_poll"] = parse_numeric(row[candidate_value_index])
    if meta["date"] is None:
        meta["date"] = fallback_date_for_path(path, family)
    hinted_year = source_year_hint(path)
    if hinted_year is not None:
        try:
            parsed_year = int(str(meta["date"])[:4])
        except (TypeError, ValueError):
            parsed_year = hinted_year
        if abs(parsed_year - hinted_year) > 1:
            meta["date"] = fallback_date_for_path(path, family)
    if meta["constituency"] is None:
        meta["constituency"] = infer_constituency_from_path(path, family)
    if meta["council"] is None:
        meta["council"] = infer_council_from_path(path, family)
    if meta["votes_polled"] is not None and meta["valid_votes"] is not None and meta["votes_polled"] < meta["valid_votes"]:
        if meta["invalid_votes"] is not None:
            meta["votes_polled"] = meta["valid_votes"] + meta["invalid_votes"]
    if meta["percent_poll"] is not None and meta["percent_poll"] > 100 and meta["electorate"] and meta["votes_polled"] is not None:
        meta["percent_poll"] = meta["votes_polled"] / meta["electorate"] * 100
    if meta["percent_poll"] is None and meta["votes_polled"] is not None and meta["electorate"]:
        meta["percent_poll"] = meta["votes_polled"] / meta["electorate"] * 100
    return meta


def infer_constituency_from_path(path: Path, family: str) -> str:
    stem = path.stem.replace("-corrected", "")
    if family == "euro":
        return "Northern Ireland"
    if family == "lgov":
        stem = re.sub(r"^lg\d{2,4}-[A-Z]{3}-", "", stem, flags=re.I)
        return normalize_space(stem.replace("-", " "))
    if family == "asby":
        stem = re.sub(r"^as\d{2}-", "", stem, flags=re.I)
        return normalize_space(stem.replace("-", " "))
    if family == "conv":
        stem = re.sub(r"^cc\d{2}-", "", stem, flags=re.I)
        return normalize_space(stem.replace("-", " "))
    return normalize_space(stem)


def infer_council_from_path(path: Path, family: str) -> str | None:
    if family != "lgov":
        return None
    match = re.search(r"^lg\d{2,4}-([A-Za-z]{2,3})-", path.stem)
    if not match:
        return None
    code = match.group(1).upper()
    return LGOV_COUNCIL_BY_CODE.get(code, code)


def build_contest(path: Path, registry: PersonRegistry) -> Contest:
    family = family_for_path(path)
    rows = load_sheet_rows(path)
    header_row = find_header_row(rows)
    first_stage_col = find_first_stage_col(rows, header_row)
    stages = extract_stages(rows, header_row, first_stage_col)
    metadata_start = find_metadata_start(rows, header_row, first_stage_col)
    compacted = compact_candidate_rows(rows, header_row, metadata_start, first_stage_col)
    metadata = extract_metadata(rows, metadata_start, path, family)
    event, event_type = family_metadata(family)
    body = elected_body_for_family(family)

    candidates: list[Candidate] = []
    nontransferable: Candidate | None = None
    for row in compacted:
        marker = normalize_space(row[0] if len(row) > 0 else "")
        number = normalize_space(row[1] if len(row) > 1 else "")
        raw_name = normalize_space(row[2] if len(row) > 2 else "")
        desc = normalize_space(row[3] if len(row) > 3 else "")
        label_blob = " ".join(filter(None, [raw_name, desc])).lower()
        if label_blob in {"total", "totals"}:
            continue
        first_pref = parse_numeric(row[first_stage_col] if len(row) > first_stage_col else None)
        deltas: list[float | None] = []
        totals: list[float | None] = []
        col = first_stage_col + 1
        for _ in stages[1:]:
            delta = parse_numeric(row[col] if len(row) > col else None)
            total = parse_numeric(row[col + 1] if len(row) > col + 1 else None)
            deltas.append(delta)
            totals.append(total)
            col += 2
        candidate = Candidate(
            marker=marker,
            number=number,
            raw_name=raw_name,
            source_party=desc,
            party=normalize_party(desc),
            first_pref=first_pref,
            deltas=deltas,
            totals=totals,
        )
        if desc.lower() in {"non-transferable", "non -transferable", "non transferable"}:
            candidate.raw_name = "NonTransferable"
            candidate.party = ""
            nontransferable = candidate
        elif raw_name:
            candidate.person_id = registry.get(raw_name)
            candidates.append(candidate)

    election_key = f"{metadata['date']}|{body}|{metadata['constituency']}"
    return Contest(
        family=family,
        date=metadata["date"],
        event=event,
        event_type=event_type,
        elected_body=body,
        constituency=metadata["constituency"],
        council=metadata["council"],
        election_key=election_key,
        seats=int(metadata["seats"]) if metadata["seats"] is not None else None,
        electorate=metadata["electorate"],
        votes_polled=metadata["votes_polled"],
        valid_votes=metadata["valid_votes"],
        invalid_votes=metadata["invalid_votes"],
        quota=metadata["quota"],
        percent_poll=metadata["percent_poll"],
        stages=stages,
        candidates=candidates,
        nontransferable=nontransferable,
        source_path=str(path),
    )


def donor_bundle_for_stage(contest: Contest, stage_offset: int) -> tuple[str | None, str | None, str | None, float | None, int | None]:
    donor_candidates = []
    donor_total = 0.0
    for candidate in contest.candidates:
        if stage_offset >= len(candidate.deltas):
            continue
        delta = candidate.deltas[stage_offset]
        if delta is not None and delta < 0:
            donor_candidates.append(candidate)
            donor_total += abs(delta)
    if not donor_candidates:
        return (None, None, None, None, None)
    donor_ids = ",".join(str(candidate.person_id) for candidate in donor_candidates if candidate.person_id is not None)
    donor_names = ", ".join(to_display_name(candidate.raw_name) for candidate in donor_candidates)
    donor_parties = ", ".join(candidate.party or candidate.source_party for candidate in donor_candidates)
    source_person_id = donor_candidates[0].person_id if len(donor_candidates) == 1 else None
    return (donor_ids or None, donor_names or None, donor_parties or None, donor_total or None, source_person_id)


def to_display_name(raw_name: str) -> str:
    display, _, _ = split_name(raw_name)
    return display


def outcome_for_candidate(candidate: Candidate) -> str | None:
    marker = normalize_space(candidate.marker).upper()
    if marker.startswith("E"):
        return "Elected"
    if any(delta is not None and delta < 0 for delta in candidate.deltas):
        return "Excluded"
    return None


def load_party_normalization_lookup(path: Path | None) -> PartyNormalizationLookup:
    lookup = PartyNormalizationLookup()
    if path is None:
        return lookup
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            raw_name = normalize_space(row.get("source_party_name"))
            dedup_name = normalize_space(row.get("deduplicated_party_name")) or None
            wiki_name = normalize_space(row.get("wikipedia_party_name")) or None
            if not raw_name:
                continue
            lookup.by_raw[raw_name] = (dedup_name, wiki_name)
            canonical_key = canonical_label(raw_name)
            existing = lookup.by_canonical.get(canonical_key)
            if existing is None:
                lookup.by_canonical[canonical_key] = (dedup_name, wiki_name)
    return lookup


def build_election_results_rows(contest: Contest, party_lookup: PartyNormalizationLookup | None = None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for candidate in contest.candidates + ([contest.nontransferable] if contest.nontransferable else []):
        if candidate is None:
            continue
        display_name, first_name, last_name = split_name(candidate.raw_name)
        dedup_name, wiki_name = party_lookup.resolve(candidate.source_party) if party_lookup else (None, None)
        row: dict[str, Any] = {header: None for header in ELECTION_RESULTS_HEADERS}
        row.update(
            {
                "Date": contest.date,
                "Event": contest.event,
                "EventType": contest.event_type,
                "ElectedBody": contest.elected_body,
                "Source Party Name": candidate.source_party or None,
                "Deduplicated Party Name": dedup_name,
                "Wikipedia Party Name": wiki_name,
                "ResultType": "NonTransferable" if candidate.raw_name == "NonTransferable" else "Candidate",
                "Party Name": candidate.party or None,
                "Source Name": candidate.raw_name if candidate.raw_name != "NonTransferable" else None,
                "Name usually known by": display_name if candidate.raw_name != "NonTransferable" else None,
                "First Name": first_name if candidate.raw_name != "NonTransferable" else None,
                "Last Name": last_name if candidate.raw_name != "NonTransferable" else None,
                "Constituency": contest.constituency,
                "Council": contest.council,
                "Outcome": outcome_for_candidate(candidate),
                "PersonID": candidate.person_id if candidate.raw_name != "NonTransferable" else None,
                "ElectionKey": contest.election_key,
            }
        )
        row["Votes1"] = candidate.first_pref
        prev_total = candidate.first_pref
        for idx, _stage in enumerate(contest.stages[1:], start=1):
            donor_ids, donor_names, donor_parties, _donor_total, _source_id = donor_bundle_for_stage(contest, idx - 1)
            raw_delta = candidate.deltas[idx - 1] if idx - 1 < len(candidate.deltas) else None
            raw_total = candidate.totals[idx - 1] if idx - 1 < len(candidate.totals) else None
            delta = raw_delta if raw_delta is not None else 0
            next_total = raw_total if raw_total is not None else (prev_total + delta if prev_total is not None else None)
            if prev_total is None:
                break
            row[f"Transfers{idx}"] = delta
            row[f"TransferSubject{idx}"] = donor_ids
            row[f"TransferName{idx}"] = donor_names
            row[f"TransferParty{idx}"] = donor_parties
            if idx + 1 <= 23:
                row[f"Votes{idx + 1}"] = next_total
            prev_total = next_total
        if candidate.raw_name != "NonTransferable" and contest.valid_votes:
            row["%ValidShare"] = round((candidate.first_pref or 0) / contest.valid_votes * 100, 12)
        if candidate.raw_name != "NonTransferable" and contest.electorate:
            row["%ElectorateShare"] = round((candidate.first_pref or 0) / contest.electorate * 100, 12)
        rows.append(row)

    summary_rows = [
        ("Electorate", contest.electorate),
        ("Quota", contest.quota),
        ("Spoiled", contest.invalid_votes),
        ("Did not vote", (contest.electorate - contest.votes_polled) if contest.electorate is not None and contest.votes_polled is not None else None),
    ]
    for result_type, value in summary_rows:
        if value is None:
            continue
        row = {header: None for header in ELECTION_RESULTS_HEADERS}
        row.update(
            {
                "Date": contest.date,
                "Event": contest.event,
                "EventType": contest.event_type,
                "ElectedBody": contest.elected_body,
                "ResultType": result_type,
                "Constituency": contest.constituency,
                "Council": contest.council,
                "Votes1": value,
                "ElectionKey": contest.election_key,
            }
        )
        rows.append(row)
    return rows


def build_transfer_rows(contest: Contest, party_lookup: PartyNormalizationLookup | None = None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    actors = contest.candidates + ([contest.nontransferable] if contest.nontransferable else [])
    for candidate in actors:
        if candidate is None:
            continue
        display_name, _, _ = split_name(candidate.raw_name)
        dedup_name, wiki_name = party_lookup.resolve(candidate.source_party) if party_lookup else (None, None)
        prev_total = candidate.first_pref if candidate.first_pref is not None else 0
        for idx, _stage in enumerate(contest.stages[1:], start=1):
            donor_ids, donor_names, donor_parties, donor_total, source_person_id = donor_bundle_for_stage(contest, idx - 1)
            raw_delta = candidate.deltas[idx - 1] if idx - 1 < len(candidate.deltas) else None
            raw_total = candidate.totals[idx - 1] if idx - 1 < len(candidate.totals) else None
            delta = raw_delta if raw_delta is not None else 0
            next_total = raw_total if raw_total is not None else (prev_total + delta if prev_total is not None else None)
            if prev_total is None:
                break
            relation = "NonTransferable" if candidate.raw_name == "NonTransferable" else "Different party"
            if candidate.raw_name != "NonTransferable":
                if delta < 0 and source_person_id == candidate.person_id:
                    relation = "Outgoing"
                elif donor_parties and len({part.strip() for part in donor_parties.split(",") if part.strip()}) == 1 and normalize_space(donor_parties) == candidate.party:
                    relation = "Same party"
            transfer_pct = None
            if donor_total and donor_total > 0 and delta is not None:
                transfer_pct = round(abs(delta) / donor_total * 100, 12)
            row = {header: None for header in TRANSFERS_HEADERS}
            row.update(
                {
                    "Date": contest.date,
                    "Event": contest.event,
                    "Constituency": contest.constituency,
                    "Council": contest.council,
                    "ElectedBody": contest.elected_body,
                    "ResultType": "NonTransferable" if candidate.raw_name == "NonTransferable" else "Candidate",
                    "PersonID": candidate.person_id if candidate.raw_name != "NonTransferable" else None,
                    "Name": None if candidate.raw_name == "NonTransferable" else display_name,
                    "Party": candidate.party or None,
                    "Deduplicated Party Name": dedup_name,
                    "Wikipedia Party Name": wiki_name,
                    "Count": idx,
                    "Votes": prev_total,
                    "Transfers": delta,
                    "TransferSubject": donor_ids,
                    "TransferName": donor_names,
                    "TransferParty": donor_parties,
                    "TransferPct": transfer_pct,
                    "EliminatedThisRound": bool(delta < 0 and candidate.raw_name != "NonTransferable" and (outcome_for_candidate(candidate) == "Excluded")),
                    "ElectedThisRound": bool(delta < 0 and candidate.raw_name != "NonTransferable" and (outcome_for_candidate(candidate) == "Elected")),
                    "TransferPartyRelation": relation,
                    "SourcePersonID": source_person_id,
                }
            )
            rows.append(row)
            prev_total = next_total
    return rows


def write_workbook(election_rows: list[dict[str, Any]], transfer_rows: list[dict[str, Any]], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws_results = wb.active
    ws_results.title = "ElectionResults"
    ws_results.append(ELECTION_RESULTS_HEADERS)
    for row in election_rows:
        ws_results.append([row.get(header) for header in ELECTION_RESULTS_HEADERS])
    ws_transfers = wb.create_sheet("Transfers")
    ws_transfers.append(TRANSFERS_HEADERS)
    for row in transfer_rows:
        ws_transfers.append([row.get(header) for header in TRANSFERS_HEADERS])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def benchmark_against_reference(contests: list[Contest], election_rows: list[dict[str, Any]], transfer_rows: list[dict[str, Any]], reference_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(reference_path, data_only=True, read_only=True)
    ws_results = wb["ElectionResults"]
    ws_transfers = wb["Transfers"]
    results_headers = [cell.value for cell in next(ws_results.iter_rows(min_row=1, max_row=1))]
    transfers_headers = [cell.value for cell in next(ws_transfers.iter_rows(min_row=1, max_row=1))]
    results_idx = {name: idx for idx, name in enumerate(results_headers)}
    transfers_idx = {name: idx for idx, name in enumerate(transfers_headers)}
    report: dict[str, Any] = {"contests": []}
    generated_results_by_key: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    generated_transfers_by_key: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for row in election_rows:
        key = (row["Date"], row["ElectedBody"], row["Constituency"])
        generated_results_by_key.setdefault(key, []).append(row)
    for row in transfer_rows:
        key = (row["Date"], row["ElectedBody"], row["Constituency"])
        generated_transfers_by_key.setdefault(key, []).append(row)
    for contest in contests:
        key = (contest.date, contest.elected_body, contest.constituency)
        reference_results = [
            record
            for record in ws_results.iter_rows(min_row=2, values_only=True)
            if str(record[results_idx["Date"]])[:10] == contest.date
            and record[results_idx["ElectedBody"]] == contest.elected_body
            and record[results_idx["Constituency"]] == contest.constituency
        ]
        reference_transfers = [
            record
            for record in ws_transfers.iter_rows(min_row=2, values_only=True)
            if str(record[transfers_idx["Date"]])[:10] == contest.date
            and record[transfers_idx["ElectedBody"]] == contest.elected_body
            and record[transfers_idx["Constituency"]] == contest.constituency
        ]
        generated_result_rows = generated_results_by_key.get(key, [])
        generated_transfer_rows = generated_transfers_by_key.get(key, [])
        candidate_first_pref_sum = sum((row.get("Votes1") or 0) for row in generated_result_rows if row.get("ResultType") == "Candidate")
        reference_first_pref_sum = sum((record[results_idx["Votes1"]] or 0) for record in reference_results if record[results_idx["ResultType"]] == "Candidate")
        report["contests"].append(
            {
                "key": "|".join(key),
                "source_path": contest.source_path,
                "generated_result_rows": len(generated_result_rows),
                "reference_result_rows": len(reference_results),
                "generated_transfer_rows": len(generated_transfer_rows),
                "reference_transfer_rows": len(reference_transfers),
                "generated_candidate_first_pref_sum": candidate_first_pref_sum,
                "reference_candidate_first_pref_sum": reference_first_pref_sum,
                "generated_quota": contest.quota,
                "reference_quota": next((record[results_idx["Votes1"]] for record in reference_results if record[results_idx["ResultType"]] == "Quota"), None),
            }
        )
    return report


def contests_to_rows(contests: list[Contest], party_lookup: PartyNormalizationLookup | None = None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    election_rows: list[dict[str, Any]] = []
    transfer_rows: list[dict[str, Any]] = []
    for contest in contests:
        election_rows.extend(build_election_results_rows(contest, party_lookup))
        transfer_rows.extend(build_transfer_rows(contest, party_lookup))
    return election_rows, transfer_rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", required=True, help="Root folder containing asby/conv/euro/lgov directories")
    parser.add_argument("--families", nargs="+", required=True, choices=["asby", "conv", "euro", "lgov"])
    parser.add_argument("--output", required=True, help="Output XLSX path")
    parser.add_argument("--party-normalization-csv", help="Optional CSV with source_party_name, deduplicated_party_name, wikipedia_party_name")
    parser.add_argument("--reference-workbook", help="Optional Full election tables.xlsx path for benchmark reporting")
    parser.add_argument("--benchmark-report", help="Optional JSON path for benchmark report")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = Path(args.root)
    registry = PersonRegistry()
    files = preferred_stv_files(root, args.families)
    contests = [build_contest(path, registry) for path in files]
    party_lookup = load_party_normalization_lookup(Path(args.party_normalization_csv)) if args.party_normalization_csv else PartyNormalizationLookup()
    election_rows, transfer_rows = contests_to_rows(contests, party_lookup)
    write_workbook(election_rows, transfer_rows, Path(args.output))
    if args.reference_workbook and args.benchmark_report:
        report = benchmark_against_reference(contests, election_rows, transfer_rows, Path(args.reference_workbook))
        Path(args.benchmark_report).parent.mkdir(parents=True, exist_ok=True)
        Path(args.benchmark_report).write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "files": len(files),
                "contests": len(contests),
                "election_rows": len(election_rows),
                "transfer_rows": len(transfer_rows),
                "output": str(Path(args.output)),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
