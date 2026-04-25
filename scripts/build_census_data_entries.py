#!/usr/bin/env python
"""Extract NISRA Census 2021 XLSX sheets into clean CSVs for the data-entry runtime.

NISRA's Main Statistics XLSXs follow a consistent pattern: each sheet has
~5 metadata rows at the top (table name, population, geographic level,
source, blurb), then a header row whose first three columns are
'Geography', 'Geography code', and the actual statistic, then data rows.
Many sheets contain *two* sub-tables stacked vertically: an "(a) count"
sub-table on top and a "(b) area percentage" sub-table below, separated
by a blank row and a section label. Both sub-tables use an identical
column layout but the percentage table stores values as fractions in
[0,1].

Usage: build_census_data_entries.py — runs the manifest below.
"""
import csv
from pathlib import Path
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
SRC_DIR = REPO / "data" / "census" / "2021"
OUT_DIR = REPO / "data" / "census" / "derived"
OUT_DIR.mkdir(parents=True, exist_ok=True)

PHASE_DIRS = [
    SRC_DIR / "census-2021-main-statistics-for-northern-ireland-phase-1-all-tables (2)",
    SRC_DIR / "census-2021-main-statistics-for-northern-ireland-phase-2-all-tables (1)",
    SRC_DIR / "census-2021-main-statistics-for-northern-ireland-phase-3-all-tables (1)",
    SRC_DIR / "census-2021-main-statistics-for-northern-ireland-supplemental-all-tables (1)",
]

def find_xlsx(filename: str) -> Path | None:
    for d in PHASE_DIRS:
        p = d / filename
        if p.exists():
            return p
    return None


def _kv_count(in_name, out_name):
    return {"section": "count", "in": in_name, "out": out_name}

def _kv_pct(in_name, out_name):
    return {"section": "percent", "in": in_name, "out": out_name}

def _kv_count_sum(in_names, out_name):
    """Sum N count columns into one output column."""
    return {"section": "count", "in": in_names, "out": out_name, "op": "sum"}

def _kv_pct_sum(in_names, out_name):
    """Sum N percentage columns into one output column. Disjoint NISRA
    sub-categories sum cleanly because their fractions all share the same
    denominator."""
    return {"section": "percent", "in": in_names, "out": out_name, "op": "sum"}

# Manifest entries are dicts:
#   {xlsx, sheet, csv, values: [{section, in, out}, ...]}
# `section` is "count" (first sub-table) or "percent" (second sub-table).
# Sheets that do not have a paired percent sub-table simply omit those rows.
MANIFEST = []

def _add(xlsx, csv_name, values, sheets):
    for sh in sheets:
        # Filenames stay flat: ms-XX-{geog}.csv
        geog_slug = sh.lower()
        out = csv_name.replace("{geog}", geog_slug)
        MANIFEST.append({
            "xlsx": xlsx, "sheet": sh, "csv": out, "values": values,
        })

# === Tier 1 (single-value, no percentages) =================================
_add("census-2021-ms-a01.xlsx", "ms-a01-{geog}.csv",
     [_kv_count("All usual residents", "AllUsualResidents")],
     ["DZ", "SDZ", "Settlement", "Ward", "DEA"])  # LGD has legacy header, leave alone

_add("census-2021-ms-a14.xlsx", "ms-a14-{geog}.csv",
     [_kv_count("Population density (number of usual residents per hectare)", "PopulationDensity")],
     ["DZ", "SDZ", "DEA", "LGD"])

_add("census-2021-ms-e01.xlsx", "ms-e01-{geog}.csv",
     [_kv_count("All households", "AllHouseholds")],
     ["DZ", "SDZ", "Settlement", "Ward", "DEA", "LGD"])

_add("census-2021-ms-e02.xlsx", "ms-e02-{geog}.csv",
     [_kv_count("Average household size\n[note 1]", "AverageHouseholdSize")],
     ["DZ", "SDZ", "Settlement", "Ward", "DEA", "LGD"])

# === Tier 2 (count + percent, paired sub-tables) ===========================

# MS-A07 — Sex
_add("census-2021-ms-a07.xlsx", "ms-a07-female-{geog}.csv",
     [_kv_count("All usual residents", "Total"),
      _kv_count("Female", "Female"),
      _kv_pct("Female", "Female_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-A16 — Country of birth (basic). Headline metric: % born in NI.
# NISRA nests the geography column as "Europe: \nUnited Kingdom:\n Northern Ireland".
_NI_BORN = "Europe: \nUnited Kingdom:\n Northern Ireland"
_add("census-2021-ms-a16.xlsx", "ms-a16-born-in-ni-{geog}.csv",
     [_kv_count("All usual residents", "Total"),
      _kv_count(_NI_BORN, "BornInNI"),
      _kv_pct(_NI_BORN, "BornInNI_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-B05 — Knowledge of Irish. Headline: % with some Irish (any ability).
_add("census-2021-ms-b05.xlsx", "ms-b05-irish-{geog}.csv",
     [_kv_count("All usual residents aged 3 and over", "Total"),
      _kv_count("Some ability in Irish", "SomeIrish"),
      _kv_pct("Some ability in Irish", "SomeIrish_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-B08 — Knowledge of Ulster-Scots
_add("census-2021-ms-b08.xlsx", "ms-b08-ulster-scots-{geog}.csv",
     [_kv_count("All usual residents aged 3 and over", "Total"),
      _kv_count("Some ability in Ulster-Scots", "SomeUlsterScots"),
      _kv_pct("Some ability in Ulster-Scots", "SomeUlsterScots_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-B19 — Religion. Headline: % Catholic (single denomination).
_add("census-2021-ms-b19.xlsx", "ms-b19-catholic-{geog}.csv",
     [_kv_count("All usual residents", "Total"),
      _kv_count("Catholic \n[note 2]", "Catholic"),
      _kv_pct("Catholic \n[note 2]", "Catholic_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-B23 — Religion or religion brought up in. Headline: % Catholic background.
_add("census-2021-ms-b23.xlsx", "ms-b23-catholic-background-{geog}.csv",
     [_kv_count("All usual residents", "Total"),
      _kv_count("Catholic \n[note 2]", "CatholicBackground"),
      _kv_pct("Catholic \n[note 2]", "CatholicBackground_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-D02 — Long-term limiting condition. Headline: % "limited a lot or a little"
# (the standard LLTI definition), summed from the two NISRA sub-categories.
_LIMITED_A_LOT    = "All usual residents:\nDay-to-day activities limited a lot"
_LIMITED_A_LITTLE = "All usual residents:\nDay-to-day activities limited a little"
_add("census-2021-ms-d02.xlsx", "ms-d02-limiting-condition-{geog}.csv",
     [_kv_count("All usual residents", "Total"),
      _kv_count_sum([_LIMITED_A_LOT, _LIMITED_A_LITTLE], "LimitingCondition"),
      _kv_pct_sum([_LIMITED_A_LOT, _LIMITED_A_LITTLE], "LimitingCondition_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-D17 — Provision of unpaid care. Headline: % providing any unpaid care
# (sum of the four hour-band columns).
_CARE_BANDS = [
    "All usual residents aged 5 and over:\nProvides 1-19 hours unpaid care per week",
    "All usual residents aged 5 and over:\nProvides 20-34 hours unpaid care per week",
    "All usual residents aged 5 and over:\nProvides 35-49 hours unpaid care per week",
    "All usual residents aged 5 and over:\nProvides 50+ hours unpaid care per week",
]
_add("census-2021-ms-d17.xlsx", "ms-d17-unpaid-care-{geog}.csv",
     [_kv_count("All usual residents aged 5 and over", "Total"),
      _kv_count_sum(_CARE_BANDS, "ProvidesUnpaidCare"),
      _kv_pct_sum(_CARE_BANDS, "ProvidesUnpaidCare_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-E10 — Car or van availability. Headline: % no car/van available.
_add("census-2021-ms-e10.xlsx", "ms-e10-no-car-{geog}.csv",
     [_kv_count("All households", "Total"),
      _kv_count("No cars or vans available", "NoCar"),
      _kv_pct("No cars or vans available", "NoCar_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-E15 — Tenure (households). Three separate metrics:
#   Owner-occupied  = Owns outright + Owns with a mortgage or loan
#   Social rented   = NIHE + Housing association
#   Private rented  = the five "Private rented:" sub-categories
_OWN = ["Owner occupied:\n Owns outright", "Owner occupied:\n Owns with a mortgage or loan"]
_SOCIAL = ["Social rented:\n Northern Ireland Housing Executive",
           "Social rented:\n Housing association or charitable trust"]
_PRIVATE = [
    "Private rented: \nPrivate landlord",
    "Private rented: \nLetting agency",
    "Private rented: \nEmployer of a household member",
    # NISRA's spelling differs between sheets — "of a household member"
    # in some, "of household member" in others. Both supplied; deduped.
    "Private rented:\n Relative or friend of a household member",
    "Private rented:\n Relative or friend of household member",
    "Private rented: \nOther",
]
_add("census-2021-ms-e15.xlsx", "ms-e15-owner-occupied-{geog}.csv",
     [_kv_count("All households", "Total"),
      _kv_count_sum(_OWN, "OwnerOccupied"),
      _kv_pct_sum(_OWN, "OwnerOccupied_pct")],
     ["Settlement", "Ward", "LGD"])
_add("census-2021-ms-e15.xlsx", "ms-e15-social-rented-{geog}.csv",
     [_kv_count("All households", "Total"),
      _kv_count_sum(_SOCIAL, "SocialRented"),
      _kv_pct_sum(_SOCIAL, "SocialRented_pct")],
     ["Settlement", "Ward", "LGD"])
_add("census-2021-ms-e15.xlsx", "ms-e15-private-rented-{geog}.csv",
     [_kv_count("All households", "Total"),
      _kv_count_sum(_PRIVATE, "PrivateRented"),
      _kv_pct_sum(_PRIVATE, "PrivateRented_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-G01 — Highest qualification level. Two metrics: % no qualifications and
# % Level 4+.
_add("census-2021-ms-g01.xlsx", "ms-g01-no-quals-{geog}.csv",
     [_kv_count("All usual residents aged 16 and over", "Total"),
      _kv_count("No qualifications [note 1]", "NoQuals"),
      _kv_pct("No qualifications [note 1]", "NoQuals_pct")],
     ["Settlement", "Ward", "LGD"])
_add("census-2021-ms-g01.xlsx", "ms-g01-level-4-plus-{geog}.csv",
     [_kv_count("All usual residents aged 16 and over", "Total"),
      _kv_count("Level 4 qualifications and above [note 6]", "Level4Plus"),
      _kv_pct("Level 4 qualifications and above [note 6]", "Level4Plus_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-H02 — Economic activity by sex. Headline: % unemployed (from the
# 'All usual residents aged 16 and over: Economically active: Unemployed'
# column).
_UNEMP = "Usual residents aged 16 and over: Economically active: Unemployed"
_add("census-2021-ms-h02.xlsx", "ms-h02-unemployed-{geog}.csv",
     [_kv_count("All usual residents aged 16 and over", "Total"),
      _kv_count(_UNEMP, "Unemployed"),
      _kv_pct(_UNEMP, "Unemployed_pct")],
     ["Settlement", "Ward", "LGD"])

# MS-I01 — Method of travel to work. Headline: % work-mainly-from-home.
_add("census-2021-ms-i01.xlsx", "ms-i01-work-from-home-{geog}.csv",
     [_kv_count("All usual residents aged 16 and over (excluding full-time students) in employment", "Total"),
      _kv_count("Work mainly at or from home", "WorkFromHome"),
      _kv_pct("Work mainly at or from home", "WorkFromHome_pct")],
     ["Settlement", "Ward", "LGD"])


# ============================================================================
# Extraction
# ============================================================================

def _norm(s):
    """Normalise a header for lenient matching."""
    return " ".join(str(s).split()).strip().lower()

def _find_col(header, target):
    """Find target in header; tolerant to whitespace/newline differences."""
    if target in header:
        return header.index(target)
    nt = _norm(target)
    for j, h in enumerate(header):
        if _norm(h) == nt:
            return j
    return None

def _find_subsections(rows):
    """Locate stacked 'Geography'-headed sub-tables. Returns list of
    (section_name, header_row_idx, header_list, end_row_idx).
    Sub-tables are named 'count' (first) and 'percent' (second) by convention.
    """
    headers = []
    for i, r in enumerate(rows):
        if r and r[0] is not None and str(r[0]).strip().lower() == "geography":
            headers.append(i)
    if not headers:
        return []
    section_names = ["count", "percent", "section3", "section4"]
    out = []
    for k, idx in enumerate(headers):
        end = headers[k + 1] if k + 1 < len(headers) else len(rows)
        hdr = [str(c).strip() if c is not None else "" for c in rows[idx]]
        out.append((section_names[k] if k < len(section_names) else f"section{k+1}",
                    idx, hdr, end))
    return out

def extract_sheet(xlsx_path: Path, sheet: str, out_csv: Path, values: list):
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        print(f"  ! {sheet} not in {xlsx_path.name}; available: {wb.sheetnames}")
        return 0
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))

    sections = _find_subsections(rows)
    if not sections:
        print(f"  ! no header row in {xlsx_path.name}/{sheet}")
        return 0
    by_name = {s[0]: s for s in sections}

    # accum: code -> {"Geography": ..., "GeographyCode": ..., **value_outs}
    accum = {}
    code_order = []  # preserve count-section order for stable CSV output

    for vs in values:
        section, in_spec, out_name = vs["section"], vs["in"], vs["out"]
        op = vs.get("op")
        if section not in by_name:
            print(f"  ! sheet {sheet} has no '{section}' sub-table; "
                  f"have {[s[0] for s in sections]} — skipping {out_name}")
            continue
        _, hdr_idx, hdr, end_idx = by_name[section]

        # Resolve column indices. For sum operations, in_spec is a list of
        # candidate column names; multiple spellings can be supplied to
        # tolerate NISRA inconsistencies between sheets ("of a household"
        # vs "of household"). Indices are deduped so a column matched by
        # two synonyms is summed only once.
        if op == "sum":
            cols_set = []
            missing = []
            for n in in_spec:
                ci = _find_col(hdr, n)
                if ci is None:
                    missing.append(n)
                elif ci not in cols_set:
                    cols_set.append(ci)
            cols = cols_set
            if not cols:
                print(f"  ! sheet {sheet} {section}: no sum target columns matched; missing {missing}; have {hdr}")
                continue
            if missing:
                # Soft warning — proceed with the columns that did match.
                print(f"  · sheet {sheet} {section}: missing variants {missing} (using {len(cols)} matched cols)")
        else:
            ci = _find_col(hdr, in_spec)
            if ci is None:
                print(f"  ! sheet {sheet} {section}: no column '{in_spec}'; have {hdr}")
                continue
            cols = [ci]

        for r in rows[hdr_idx + 1:end_idx]:
            if not r: continue
            name = r[0]
            if name is None: continue
            sname = str(name).strip()
            if sname.lower() == "geography": break
            if sname.startswith(("Geographic", "Note", "Source", "MS-")): continue
            code = r[1]
            if code is None: continue
            scode = str(code).strip()
            entry = accum.get(scode)
            if entry is None:
                entry = {"Geography": sname, "GeographyCode": scode}
                accum[scode] = entry
                if section == "count":
                    code_order.append(scode)
            # Read value(s)
            raw_vals = [r[c] for c in cols]
            try:
                if op == "sum":
                    nums = [float(x) for x in raw_vals if x is not None]
                    if not nums:
                        entry[out_name] = ""
                        continue
                    total = sum(nums)
                else:
                    if raw_vals[0] is None:
                        entry[out_name] = ""
                        continue
                    total = float(raw_vals[0])
            except (TypeError, ValueError):
                entry[out_name] = str(raw_vals[0]).strip() if raw_vals[0] is not None else ""
                continue
            if section == "percent":
                entry[out_name] = f"{total * 100:.2f}"
            elif op == "sum":
                # Counts are integers; preserve as such when possible.
                entry[out_name] = str(int(round(total))) if abs(total - round(total)) < 1e-9 else f"{total:.4f}"
            else:
                entry[out_name] = str(raw_vals[0]).strip()

    # Write CSV
    if not code_order:
        # Some entries lacked any 'count' section — fall back to insertion order.
        code_order = list(accum.keys())

    out_cols = ["Geography", "GeographyCode"] + [v["out"] for v in values]
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(out_cols)
        for code in code_order:
            row = accum[code]
            w.writerow([row.get(c, "") for c in out_cols])
    return len(code_order)

def main():
    for spec in MANIFEST:
        src = find_xlsx(spec["xlsx"])
        out = OUT_DIR / spec["csv"]
        if src is None:
            print(f"  ! source missing: {spec['xlsx']} (searched all phase dirs)")
            continue
        n = extract_sheet(src, spec["sheet"], out, spec["values"])
        print(f"  {spec['xlsx']} [{spec['sheet']}] -> {out.name}  ({n} rows)")

if __name__ == "__main__":
    main()
