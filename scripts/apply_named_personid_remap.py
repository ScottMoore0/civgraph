"""Apply explicit name -> PersonID remaps to the modern local workbook."""

from __future__ import annotations

import argparse
from pathlib import Path
import re

from openpyxl import load_workbook


APPROVED_NAME_TO_ID = {
    "Andrew McMurray": 64846,
    "Andrew Muir": 3999,
    "Aaron Callan": 8594,
    "Adrian McQuillan": 33610,
    "Angela Mulholland": 54895,
    "Barry McElduff": 35846,
    "Brian Tierney": 34682,
    "Cara Hunter": 38585,
    "Carla Lockhart": 20746,
    "Cadogan Enright": 38721,
    "Carl McClean": 46273,
    "Cathal Mallaghan": 56738,
    "Cathal Ó hOisín": 3349,
    "Cathy Mason": 42486,
    "Charlotte Carson": 100006,
    "Chris McCaw": 21001,
    "Clare Bailey": 6660,
    "Claire Hanna": 40948,
    "Colin McGrath": 46652,
    "Connie Egan": 4984,
    "Darryl Wilson": 77984,
    "Darrin Foster": 90156,
    "David Harding": 20313,
    "David Jones": 83170,
    "Danny Baker": 97753,
    "Danny Donnelly": 51217,
    "Denise Mullen": 17344,
    "Derek Hussey": 21520,
    "Geraldine Rice": 87910,
    "Glyn Hanna": 46908,
    "Harold McKee": 47476,
    "Henry Reilly": 86150,
    "J. J. Magee": 91406,
    "Jenny Palmer": 15541,
    "Jordan Doran": 100007,
    "Jim Rodgers": 68780,
    "John Kyle": 41035,
    "Josephine Deehan": 65441,
    "Liz Kimmins": 44021,
    "Maurice Devenney": 84453,
    "Patsy Kelly": 16975,
    "Paul Michael": 33120,
    "Ryan McCready": 44861,
    "Simon Lee": 61216,
    "Sorcha Eastwood": 4701,
    "Sorcha McAnespy": 66418,
    "Stephanie Quigley": 89162,
    "Stephen Cooper": 85704,
    "Stephen Dunne": 100002,
    "Stephen Moutray": 11666,
    "Vasundhara Kamble": 12847,
    "Willie Clarke": 3524,
    "William McCandless": 65776,
}


def normalize_lookup_name(value: str) -> str:
    text = str(value or "").strip()
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"[‡]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def remap_scalar(value, mapping: dict[int, int]):
    if value in (None, ""):
        return value, False
    try:
        current = int(value)
    except (TypeError, ValueError):
        return value, False
    new_value = mapping.get(current)
    if new_value is None or new_value == current:
        return value, False
    return new_value, True


def remap_id_list(value, mapping: dict[int, int]):
    if value in (None, ""):
        return value, False
    parts = [part.strip() for part in str(value).split(",")]
    changed = False
    remapped: list[str] = []
    for part in parts:
        if not part:
            remapped.append(part)
            continue
        try:
            current = int(part)
        except ValueError:
            remapped.append(part)
            continue
        new_value = mapping.get(current, current)
        if new_value != current:
            changed = True
        remapped.append(str(new_value))
    return ", ".join(remapped), changed


def remap_subject_ids_by_names(subject_value, name_value, approved_name_to_id: dict[str, int]):
    if subject_value in (None, "") or name_value in (None, ""):
        return subject_value, False
    subject_parts = [part.strip() for part in str(subject_value).split(",")]
    name_parts = [normalize_lookup_name(part) for part in str(name_value).split(",")]
    if len(subject_parts) != len(name_parts):
        return subject_value, False
    changed = False
    remapped: list[str] = []
    for subject_part, name_part in zip(subject_parts, name_parts):
        target_id = approved_name_to_id.get(name_part)
        if not subject_part:
            remapped.append(subject_part)
            continue
        try:
            current_id = int(subject_part)
        except ValueError:
            remapped.append(subject_part)
            continue
        if target_id is not None and current_id != target_id:
            remapped.append(str(target_id))
            changed = True
        else:
            remapped.append(subject_part)
    return ", ".join(remapped), changed


def build_existing_id_mapping(workbook_path: Path) -> tuple[dict[int, int], dict[str, list[int]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["ElectionResults"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: pos for pos, header in enumerate(headers)}

    found_by_name: dict[str, set[int]] = {name: set() for name in APPROVED_NAME_TO_ID}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = normalize_lookup_name(row[idx["Name usually known by"]])
        if name not in found_by_name:
            continue
        person_id = row[idx["PersonID"]]
        if person_id in (None, ""):
            continue
        found_by_name[name].add(int(person_id))

    missing = [name for name, ids in found_by_name.items() if not ids]
    if missing:
        raise ValueError(f"Approved names not found in modern workbook: {missing}")

    id_mapping: dict[int, int] = {}
    for name, ids in found_by_name.items():
        target_id = APPROVED_NAME_TO_ID[name]
        for existing_id in ids:
            prior = id_mapping.get(existing_id)
            if prior is not None and prior != target_id:
                raise ValueError(
                    f"Existing modern ID {existing_id} would map to conflicting full IDs {prior} and {target_id}"
                )
            id_mapping[existing_id] = target_id

    return id_mapping, {name: sorted(ids) for name, ids in found_by_name.items()}


def apply_mapping(input_workbook: Path, output_workbook: Path, mapping: dict[int, int]) -> dict[str, int]:
    wb = load_workbook(input_workbook)
    stats = {
        "electionresults_personid_updates": 0,
        "electionresults_transfersubject_updates": 0,
        "transfers_personid_updates": 0,
        "transfers_transfersubject_updates": 0,
        "transfers_sourcepersonid_updates": 0,
        "transfers_remainingcandidateids_updates": 0,
    }

    ws_results = wb["ElectionResults"]
    result_headers = [cell.value for cell in next(ws_results.iter_rows(min_row=1, max_row=1))]
    result_idx = {header: pos + 1 for pos, header in enumerate(result_headers)}
    result_transfer_subject_headers = [
        header for header in result_headers if isinstance(header, str) and header.startswith("TransferSubject")
    ]
    for row_num in range(2, ws_results.max_row + 1):
        person_cell = ws_results.cell(row=row_num, column=result_idx["PersonID"])
        new_value, changed = remap_scalar(person_cell.value, mapping)
        if changed:
            person_cell.value = new_value
            stats["electionresults_personid_updates"] += 1
        for header in result_transfer_subject_headers:
            cell = ws_results.cell(row=row_num, column=result_idx[header])
            new_value, changed = remap_scalar(cell.value, mapping)
            if changed:
                cell.value = new_value
                stats["electionresults_transfersubject_updates"] += 1
            paired_name_header = header.replace("TransferSubject", "TransferName")
            if paired_name_header in result_idx:
                paired_name_value = ws_results.cell(row=row_num, column=result_idx[paired_name_header]).value
                new_value, changed = remap_subject_ids_by_names(cell.value, paired_name_value, APPROVED_NAME_TO_ID)
                if changed:
                    cell.value = new_value
                    stats["electionresults_transfersubject_updates"] += 1

    ws_transfers = wb["Transfers"]
    transfer_headers = [cell.value for cell in next(ws_transfers.iter_rows(min_row=1, max_row=1))]
    transfer_idx = {header: pos + 1 for pos, header in enumerate(transfer_headers)}
    for row_num in range(2, ws_transfers.max_row + 1):
        for header, stat_key in (
            ("PersonID", "transfers_personid_updates"),
            ("TransferSubject", "transfers_transfersubject_updates"),
            ("SourcePersonID", "transfers_sourcepersonid_updates"),
        ):
            cell = ws_transfers.cell(row=row_num, column=transfer_idx[header])
            new_value, changed = remap_scalar(cell.value, mapping)
            if changed:
                cell.value = new_value
                stats[stat_key] += 1

        transfer_subject_cell = ws_transfers.cell(row=row_num, column=transfer_idx["TransferSubject"])
        transfer_name_value = ws_transfers.cell(row=row_num, column=transfer_idx["TransferName"]).value
        new_value, changed = remap_subject_ids_by_names(transfer_subject_cell.value, transfer_name_value, APPROVED_NAME_TO_ID)
        if changed:
            transfer_subject_cell.value = new_value
            stats["transfers_transfersubject_updates"] += 1

        remaining_cell = ws_transfers.cell(row=row_num, column=transfer_idx["RemainingCandidateIDsDesc"])
        new_value, changed = remap_id_list(remaining_cell.value, mapping)
        if changed:
            remaining_cell.value = new_value
            stats["transfers_remainingcandidateids_updates"] += 1

    wb.save(output_workbook)
    return stats


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--modern-workbook",
        default=r"_tmp_xls2rar_extract\out\wiki_lgov_modern\lgov-modern-wikipedia.stvfix.xlsx",
    )
    parser.add_argument("--output-workbook", default=None)
    args = parser.parse_args()

    workbook_path = Path(args.modern_workbook)
    output_path = Path(args.output_workbook) if args.output_workbook else workbook_path
    mapping, found = build_existing_id_mapping(workbook_path)
    stats = apply_mapping(workbook_path, output_path, mapping)
    print(
        {
            "approved_name_count": len(APPROVED_NAME_TO_ID),
            "existing_id_mapping_count": len(mapping),
            "name_to_existing_ids": found,
            **stats,
        }
    )


if __name__ == "__main__":
    main()
