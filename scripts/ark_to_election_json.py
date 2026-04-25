#!/usr/bin/env python
"""Convert ARK Local Government election XLS spreadsheets into the per-DEA
JSON schema used by election-viewer-package/data/elections/local-government/.

Usage:
    python scripts/ark_to_election_json.py YEAR [DATE]

For each lg{YR}-{COUNTY}-{DEA}.xls in _tmp_xls2rar_extract/xls/lgov/{YEAR}/,
emit election-viewer-package/data/elections/local-government/{DATE}/{dea}.json.

The ARK XLS layout (1985-2005) is:
    row 0  : "E' for Elected" | Number | Name of Candidates | Description | 1st Stage | Stage 2 | ... |
    row 1  : (sub-header — "First Preference Votes", "Result", ...)
    rows 2+: candidate rows
    bottom : Non-transferable, Totals, then 5-7 metadata rows
             with [date|district|"District Electoral Area..."|<key>|<value>]
             — keys include 'Eligible Electorate', 'Number of members to be Elected',
             'Votes Polled', 'Invalid Votes', 'Total Valid Votes', 'Electoral Quota'

1973/1977 use 'Candidates' instead of 'Name of Candidates' and may have
no first-name in the candidate cell (just surname). 2011 is .xlsx.
"""
import os
import re
import sys
import json
import unicodedata
from pathlib import Path

import xlrd
try:
    from openpyxl import load_workbook as _openpyxl_load
except ImportError:
    _openpyxl_load = None

REPO_ROOT = Path(__file__).resolve().parent.parent
ARK_DIR   = REPO_ROOT / "_tmp_xls2rar_extract" / "xls" / "lgov"
OUT_BASE  = REPO_ROOT / "election-viewer-package" / "data" / "elections" / "local-government"

# Approximate election dates per year (2nd Wed of May, mostly)
DEFAULT_DATES = {
    "1973": "1973-05-30", "1977": "1977-05-18", "1981": "1981-05-20",
    "1985": "1985-05-15", "1989": "1989-05-17", "1993": "1993-05-19",
    "1997": "1997-05-21", "2001": "2001-06-07", "2005": "2005-05-05",
    "2011": "2011-05-05",
}

PARTY_CANONICAL = {
    # 1970s-80s
    "off. un.": "Ulster Unionist Party",
    "official ulster unionist": "Ulster Unionist Party",
    "ulster unionist": "Ulster Unionist Party",
    "ulster unionist party": "Ulster Unionist Party",
    "uup": "Ulster Unionist Party",
    "u.u.p.": "Ulster Unionist Party",
    "d.u.p.": "Democratic Unionist Party",
    "dup": "Democratic Unionist Party",
    "democratic unionist": "Democratic Unionist Party",
    "democratic unionist d.u.p": "Democratic Unionist Party",
    "democratic unionist d.u.p.": "Democratic Unionist Party",
    "democratic unionist party d.u.p.": "Democratic Unionist Party",
    "democratic unionist party": "Democratic Unionist Party",
    "sdlp": "SDLP",
    "s.d.l.p.": "SDLP",
    "soc. dem. & lab.": "SDLP",
    "social democratic and labour": "SDLP",
    "social democratic and labour party": "SDLP",
    "sf": "Sinn Féin",
    "s.f.": "Sinn Féin",
    "sinn fein": "Sinn Féin",
    "sinn féin": "Sinn Féin",
    "alliance": "Alliance Party",
    "all.": "Alliance Party",
    "alliance party": "Alliance Party",
    "alliance party of n.i.": "Alliance Party",
    "indp.": "Independent",
    "ind.": "Independent",
    "independent": "Independent",
    "n.i.l.p.": "NI Labour",
    "nilp": "NI Labour",
    "n.i. labour": "NI Labour",
    "northern ireland labour": "NI Labour",
    "labour": "Labour",
    "p.u.p.": "PUP",
    "pup": "PUP",
    "progressive unionist": "PUP",
    "u.k.u.p.": "UKUP",
    "ukup": "UKUP",
    "uk unionist": "UKUP",
    "vanguard": "Vanguard",
    "v.u.p.p.": "Vanguard",
    "tuv": "TUV",
    "traditional unionist voice": "TUV",
    "green": "Green Party",
    "green party": "Green Party",
    "ww": "Workers' Party",
    "wp": "Workers' Party",
    "the workers' party": "Workers' Party",
    "workers' party": "Workers' Party",
    "workers party": "Workers' Party",
}

def norm_party(raw):
    if not raw: return "Independent"
    k = re.sub(r"\s+", " ", str(raw)).strip().lower().rstrip(".")
    if k in PARTY_CANONICAL: return PARTY_CANONICAL[k]
    # Strip trailing punct + retry
    k2 = re.sub(r"[^\w\s'\-éÉ]", " ", k).strip()
    k2 = re.sub(r"\s+", " ", k2)
    if k2 in PARTY_CANONICAL: return PARTY_CANONICAL[k2]
    return raw

def parse_name(raw):
    s = re.sub(r"\s+", " ", str(raw or "")).strip()
    if "," in s:
        last, first = s.split(",", 1)
        return first.strip(), last.strip()
    parts = s.split()
    if len(parts) >= 2:
        return " ".join(parts[:-1]), parts[-1]
    return "", s

def slugify(s):
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")
    return s

def stable_id(*parts):
    h = 0
    s = "|".join(str(p) for p in parts)
    for ch in s:
        h = (h * 131 + ord(ch)) & 0x7FFFFFFFFFFFFFFF
    return str(h)

COUNTY_NAMES = {
    # ARK 26-district council codes (case-insensitive lookup applied)
    "ANT": "Antrim",       "ARD": "Ards",         "ARM": "Armagh",
    "BMA": "Ballymena",    "BMY": "Ballymoney",   "BRG": "Banbridge",
    "BT":  "Belfast",      "CAR": "Carrickfergus","CAS": "Castlereagh",
    "COL": "Coleraine",    "COO": "Cookstown",    "CRA": "Craigavon",
    "DE":  "Derry",        "DOW": "Down",         "DUN": "Dungannon",
    "FER": "Fermanagh",    "LAR": "Larne",        "LIM": "Limavady",
    "LIS": "Lisburn",      "MAG": "Magherafelt",  "MOY": "Moyle",
    "NAM": "Newry and Mourne", "NEW": "Newtownabbey", "NOD": "North Down",
    "OMA": "Omagh",        "STR": "Strabane",
}

class _OpenpyxlAdapter:
    """Wrap an openpyxl worksheet to look enough like xlrd for our reader."""
    def __init__(self, ws):
        self._ws = ws
        self.nrows = ws.max_row or 0
        self.ncols = ws.max_column or 0
        self._rows = list(ws.iter_rows(values_only=True))
    def cell_value(self, r, c):
        if r < len(self._rows) and c < len(self._rows[r]):
            v = self._rows[r][c]
            return "" if v is None else v
        return ""

class _OpenpyxlBookAdapter:
    def __init__(self, wb):
        self._wb = wb
    def sheet_by_index(self, i):
        ws = self._wb[self._wb.sheetnames[i]]
        return _OpenpyxlAdapter(ws)

def convert_workbook(path: Path, year: str):
    """Read one ARK XLS/XLSX file -> return dict matching the JSON schema."""
    if path.suffix.lower() == ".xlsx":
        if _openpyxl_load is None:
            print(f"  fail {path.name}: openpyxl not installed")
            return None
        try:
            wb = _OpenpyxlBookAdapter(_openpyxl_load(str(path), data_only=True, read_only=True))
        except Exception as e:
            print(f"  fail {path.name}: {e}")
            return None
    else:
        try:
            wb = xlrd.open_workbook(str(path), encoding_override="cp1252")
        except Exception:
            try:
                wb = xlrd.open_workbook(str(path))
            except Exception as e:
                print(f"  fail {path.name}: {e}")
                return None
    sh = wb.sheet_by_index(0)

    # Build candidate list and detect end-of-candidates
    candidates = []
    headers = [str(sh.cell_value(0, c)) for c in range(sh.ncols)]
    # Find first-pref column (usually col 4)
    fp_col = next((i for i, h in enumerate(headers)
                  if "1st Stage" in h or "Stage No. 1" in h or "First Preference" in h), 4)

    for r in range(2, sh.nrows):
        try:
            elected_marker = str(sh.cell_value(r, 0)).strip()
            number       = sh.cell_value(r, 1)
            name_raw     = str(sh.cell_value(r, 2)).strip()
            party_raw    = str(sh.cell_value(r, 3)).strip()
            fp           = sh.cell_value(r, fp_col)
        except IndexError:
            continue
        if not isinstance(number, float) or not name_raw: continue
        if name_raw.lower().startswith(("non-transferable", "totals", "total ")): continue
        first, last = parse_name(name_raw)
        try:
            fp_votes = float(fp) if fp not in (None, "", "-") else 0.0
        except (ValueError, TypeError):
            fp_votes = 0.0
        candidates.append({
            "Constituency_Number": "",
            "Candidate_Id": stable_id(year, path.stem, last, first),
            "Count_Number": "1",
            "Firstname": first,
            "Surname":   last,
            "Candidate_First_Pref_Votes": f"{fp_votes:.2f}",
            "Party": norm_party(party_raw),
            "Elected": bool(elected_marker.startswith("E"))
        })

    # Metadata rows at bottom — looking for known keys in column 3
    meta = {}
    for r in range(sh.nrows - 10, sh.nrows):
        if r < 0: continue
        try:
            key = str(sh.cell_value(r, 3)).strip()
            val = sh.cell_value(r, 4)
        except IndexError: continue
        kl = key.lower().rstrip()
        if "eligible electorate" in kl or "electorate" in kl:           meta["electorate"] = val
        elif "members to be elected" in kl or "number of members" in kl: meta["seats"] = val
        elif "votes polled" in kl:                                       meta["polled"] = val
        elif "invalid votes" in kl:                                      meta["invalid"] = val
        elif "total valid votes" in kl or "valid votes" in kl:           meta["valid"] = val
        elif "electoral quota" in kl:                                    meta["quota"] = val

    def num(v, default=""):
        try:
            f = float(v)
            return f"{int(f)}" if f == int(f) else f"{f:.2f}"
        except (ValueError, TypeError):
            return default

    # DEA name from filename: lg{YR}-{COUNTY}-{DEA}.xls
    # Combine county + DEA so cross-council collisions ("Northwest", "Town")
    # become unambiguous: e.g. "Antrim Northwest", "Ballymena Town".
    stem = path.stem
    m = re.match(r"^lg\d+-([A-Z]+)-(.+)$", stem)
    if m:
        county_code = m.group(1).upper()
        dea_raw = m.group(2).replace("_", " ").replace("-", " ").strip()
        dea_raw = re.sub(r"([a-z])([A-Z])", r"\1 \2", dea_raw)
        dea_raw = re.sub(r"\s+", " ", dea_raw)
        county = COUNTY_NAMES.get(county_code, county_code)
        # Prefix the council name when the DEA name alone isn't unique:
        #   - identical to the council
        #   - a generic compass / town / area / numbered designation
        #   - "Area A" / "Area B" / etc. (1973-1981 convention)
        # Distinctive names like "Newtownards", "Bangor", "Lurgan" stand
        # alone since they're already unambiguous across the 26 councils.
        dl = dea_raw.lower()
        is_generic = (
            dl == county.lower() or
            dl in ("town", "city", "central", "east", "west", "north", "south",
                   "northwest", "northeast", "southwest", "southeast",
                   "area", "district") or
            re.match(r"^area\s+[a-z0-9]+(\s+corrected)?$", dl) or
            re.match(r"^(north|south|east|west|central|town)\s+[a-z]+$", dl)
        )
        if is_generic:
            dea_name = f"{county} {dea_raw}"
        else:
            dea_name = dea_raw
    else:
        dea_name = stem
    council_name = COUNTY_NAMES.get(m.group(1).upper(), m.group(1)) if m else ""

    return {
        "Constituency": {
            "countInfo": {
                "Constituency_Name": dea_name,
                "Constituency_Number": "",
                "Number_Of_Seats":   num(meta.get("seats")),
                "Quota":             num(meta.get("quota")),
                "Total_Electorate":  num(meta.get("electorate")),
                "Total_Poll":        num(meta.get("polled")),
                "Valid_Poll":        num(meta.get("valid")),
                "Spoiled":           num(meta.get("invalid")),
            },
            "countGroup": candidates
        }
    }

def main():
    if len(sys.argv) < 2:
        print("usage: ark_to_election_json.py YEAR [DATE]")
        sys.exit(1)
    year = sys.argv[1]
    date = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_DATES.get(year, f"{year}-05-15")

    src = ARK_DIR / year
    out = OUT_BASE / date
    out.mkdir(parents=True, exist_ok=True)

    files = sorted(list(src.glob(f"lg{year[-2:]}-*.xls")) + list(src.glob(f"lg{year[-2:]}-*.xlsx")))
    print(f"{year} -> {date}: {len(files)} XLS files")
    ok, fail = 0, 0
    deas = []
    council_map = {}   # slug -> {dea, council}
    for f in files:
        data = convert_workbook(f, year)
        if not data:
            fail += 1
            continue
        dea = data["Constituency"]["countInfo"]["Constituency_Name"]
        slug = slugify(dea)
        out_path = out / f"{slug}.json"
        out_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        # Derive council from filename prefix lg{YR}-{COUNTY}-...
        # County code is case-mixed in some files (e.g. NaM, NoD) — uppercase before lookup.
        m = re.match(r"^[Ll]g\d+-([A-Za-z]+)-", f.stem)
        if m:
            code = m.group(1).upper()
            council = COUNTY_NAMES.get(code, code)
            council_map[slug] = {"dea": dea, "council": council}
        deas.append(dea)
        ok += 1
    (out / "_council_map.json").write_text(
        json.dumps(council_map, indent=2, ensure_ascii=False), encoding="utf-8")
    n_councils = len(set(v['council'] for v in council_map.values()))
    print(f"  done: {ok} ok, {fail} fail -> {out}")
    print(f"  council map: {len(council_map)} DEAs across {n_councils} councils")

if __name__ == "__main__":
    main()
