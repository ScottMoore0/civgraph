"""Split hard-collision PersonIDs across the full workbook and downstream election JSON data."""

from __future__ import annotations

import argparse
import json
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class SplitAssignment:
    old_id: int
    new_id: int
    keep_name: str
    split_name: str


ASSIGNMENTS: list[SplitAssignment] = [
    SplitAssignment(11666, 100001, "Stephen Moutray", "Donal O'Cofaigh"),
    SplitAssignment(16781, 100002, "Paul McGlinchey", "Stephen Dunne"),
    SplitAssignment(18492, 100003, "Stuart Deignan", "Gavin Malone"),
    SplitAssignment(20545, 100004, "Alister Black", "John Lindsay"),
    SplitAssignment(33653, 100005, "Paddy Meehan", "Amy Doherty"),
    SplitAssignment(57432, 100006, "Martina McIlkenny", "Charlotte Carson"),
    SplitAssignment(62838, 100007, "Roisin McMackin", "Jordan Doran"),
]

DIRECT_ID_HEADERS = {
    "PersonID",
    "CandidateID",
    "SourcePersonID",
    "SourceCandidateID",
    "DestCandidateID",
    "RecipientCandidateID",
    "FromCandidateID",
    "ToCandidateID",
    "Candidate_Id",
    "TransferSubject",
}

GROUP_ID_HEADERS = {
    "SourceGroupID",
    "ParentGroupID",
    "ChildGroupID",
}

JSON_ID_LIST_HEADERS = {
    "RemainingCandidateIDsDesc",
    "MemberCandidateIDsJSON",
    "FromCombinationIDsJSON",
    "SourceCandidateIDsJSON",
}

GROUP_REF_TEXT_HEADERS = {
    "TopSourceGroupIDsJSON",
    "TopSourceGroupsJSON",
    "SourceGroupLabel",
    "FromGroupLabel",
    "GroupLabel",
}


def normalize_space(value: Any) -> str:
    return " ".join(str(value or "").split())


def format_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def derived_election_key_from_row(row_values: dict[str, Any]) -> str | None:
    election_key = normalize_space(row_values.get("ElectionKey"))
    if election_key:
        return election_key
    if {"Date", "Event", "Constituency"}.issubset(row_values):
        date_text = format_date(row_values["Date"])
        event = normalize_space(row_values["Event"])
        constituency = normalize_space(row_values["Constituency"])
        if date_text and event and constituency:
            return f"{date_text}|{event}|{constituency}"
    return None


def replace_exact_id_token(text: str, old_id: int, new_id: int) -> tuple[str, bool]:
    old = str(old_id)
    new = str(new_id)
    updated = []
    changed = False
    token = ""
    for ch in text:
        if ch.isdigit():
            token += ch
            continue
        if token:
            if token == old:
                updated.append(new)
                changed = True
            else:
                updated.append(token)
            token = ""
        updated.append(ch)
    if token:
        if token == old:
            updated.append(new)
            changed = True
        else:
            updated.append(token)
    return "".join(updated), changed


def replace_cell_value(value: Any, old_id: int, new_id: int) -> tuple[Any, bool]:
    if value is None or value == "":
        return value, False
    if isinstance(value, bool):
        return value, False
    if isinstance(value, int):
        return (new_id, True) if value == old_id else (value, False)
    if isinstance(value, float):
        if value.is_integer() and int(value) == old_id:
            return float(new_id), True
        return value, False
    if isinstance(value, str):
        return replace_exact_id_token(value, old_id, new_id)
    return value, False


def is_split_name_row(row_values: dict[str, Any], assignment: SplitAssignment) -> bool:
    names_to_check = [
        "Full Name usually known by",
        "Name usually known by",
        "Name",
        "SourceCandidateName",
        "DestCandidateName",
        "RecipientCandidateName",
        "ToCandidateName",
        "candidateName",
        "Firstname",
        "Surname",
    ]
    for header in names_to_check:
        if header not in row_values:
            continue
        value = row_values.get(header)
        if header in {"Firstname", "Surname"}:
            combined = normalize_space(f"{row_values.get('Firstname', '')} {row_values.get('Surname', '')}")
            if combined == assignment.split_name:
                return True
        elif normalize_space(value) == assignment.split_name:
            return True
    return False


def row_matches_assignment(
    row_values: dict[str, Any],
    assignment: SplitAssignment,
    target_contexts: dict[str, set[str]],
) -> bool:
    if is_split_name_row(row_values, assignment):
        return True
    key = derived_election_key_from_row(row_values)
    return bool(key and key in target_contexts.get(assignment.split_name, set()))


def backup_file(src: Path, backup_root: Path) -> None:
    backup_root.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, backup_root / src.name)


def validate_workbook(path: Path) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    _ = wb.sheetnames
    wb.close()


def load_target_contexts(full_workbook: Path) -> dict[str, set[str]]:
    wb = load_workbook(full_workbook, read_only=True, data_only=True)
    ws = wb["ElectionResults"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: pos for pos, header in enumerate(headers)}
    contexts: dict[str, set[str]] = defaultdict(set)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_values = {header: row[pos] for header, pos in idx.items()}
        name = normalize_space(row_values.get("Name usually known by"))
        key = derived_election_key_from_row(row_values)
        if not key:
            continue
        for assignment in ASSIGNMENTS:
            if name == assignment.split_name:
                contexts[assignment.split_name].add(key)
    wb.close()
    return contexts


def collect_referenced_group_ids(
    wb,
    target_contexts: dict[str, set[str]],
) -> dict[str, set[int]]:
    group_refs: dict[str, set[int]] = defaultdict(set)
    for sheet_name, headers_of_interest in {
        "EventEdges": ["SourceGroupID"],
        "CandidateSnapshots": ["TopSourceGroupIDsJSON"],
        "LocalCompositions": ["ParentGroupID", "ChildGroupID"],
        "CandidateStatePerCount_v2": ["TopSourceGroupsJSON"],
    }.items():
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        idx = {header: pos + 1 for pos, header in enumerate(headers)}
        for row_num in range(2, ws.max_row + 1):
            row_values = {header: ws.cell(row=row_num, column=col).value for header, col in idx.items()}
            key = derived_election_key_from_row(row_values)
            if not key:
                continue
            for assignment in ASSIGNMENTS:
                if key not in target_contexts.get(assignment.split_name, set()):
                    continue
                for header in headers_of_interest:
                    value = row_values.get(header)
                    if value is None or value == "":
                        continue
                    if header in GROUP_ID_HEADERS:
                        try:
                            group_refs[assignment.split_name].add(int(value))
                        except (TypeError, ValueError):
                            pass
                    else:
                        for token in "".join(ch if str(value) and ch.isdigit() else " " for ch in str(value)).split():
                            group_refs[assignment.split_name].add(int(token))
    return group_refs


def duplicate_sourcegroups_for_splits(wb, group_refs: dict[str, set[int]], stats: defaultdict[str, int]) -> dict[tuple[str, int], int]:
    ws = wb["SourceGroups"]
    headers = [cell.value for cell in ws[1]]
    idx = {header: pos + 1 for pos, header in enumerate(headers)}
    existing_rows: dict[int, int] = {}
    max_group_id = 0
    for row_num in range(2, ws.max_row + 1):
        gid = ws.cell(row=row_num, column=idx["GroupID"]).value
        if isinstance(gid, int):
            existing_rows[gid] = row_num
            max_group_id = max(max_group_id, gid)

    new_group_ids: dict[tuple[str, int], int] = {}
    for assignment in ASSIGNMENTS:
        for old_gid in sorted(group_refs.get(assignment.split_name, set())):
            row_num = existing_rows.get(old_gid)
            if not row_num:
                continue
            row_values = {header: ws.cell(row=row_num, column=col).value for header, col in idx.items()}
            touched = False
            for header in ("CandidateID", "MemberCandidateIDsJSON", "GroupLabel"):
                updated, changed = replace_cell_value(row_values.get(header), assignment.old_id, assignment.new_id)
                if changed:
                    row_values[header] = updated
                    touched = True
            if not touched:
                continue
            max_group_id += 1
            row_values["GroupID"] = max_group_id
            ws.append([row_values.get(header) for header in headers])
            new_group_ids[(assignment.split_name, old_gid)] = max_group_id
            stats["sourcegroups_rows_duplicated"] += 1
    return new_group_ids


def patch_sheet_rows(
    wb,
    target_contexts: dict[str, set[str]],
    new_group_ids: dict[tuple[str, int], int],
    stats: defaultdict[str, int],
) -> None:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        idx = {header: pos + 1 for pos, header in enumerate(headers)}
        for row_num in range(2, ws.max_row + 1):
            row_values = {header: ws.cell(row=row_num, column=col).value for header, col in idx.items()}
            for assignment in ASSIGNMENTS:
                if not row_matches_assignment(row_values, assignment, target_contexts):
                    continue
                for header, col in idx.items():
                    if header == "GroupID":
                        continue
                    cell = ws.cell(row=row_num, column=col)
                    value = cell.value
                    if header in GROUP_ID_HEADERS:
                        try:
                            old_gid = int(value)
                        except (TypeError, ValueError):
                            continue
                        new_gid = new_group_ids.get((assignment.split_name, old_gid))
                        if new_gid is not None and new_gid != old_gid:
                            cell.value = new_gid
                            stats[f"{sheet_name.lower()}_{header.lower()}_updates"] += 1
                        continue
                    if header in GROUP_REF_TEXT_HEADERS:
                        text = str(value or "")
                        changed = False
                        for (split_name, old_gid), new_gid in new_group_ids.items():
                            if split_name != assignment.split_name:
                                continue
                            text, repl = replace_exact_id_token(text, old_gid, new_gid)
                            changed = changed or repl
                        if changed:
                            cell.value = text
                            stats[f"{sheet_name.lower()}_{header.lower()}_updates"] += 1
                        continue
                    if (
                        header in DIRECT_ID_HEADERS
                        or header in JSON_ID_LIST_HEADERS
                        or header.startswith("TransferSubject")
                    ):
                        updated, changed = replace_cell_value(value, assignment.old_id, assignment.new_id)
                        if changed:
                            cell.value = updated
                            stats[f"{sheet_name.lower()}_{header.lower()}_updates"] += 1


def patch_workbook(
    full_workbook: Path,
    backup_root: Path,
    target_contexts: dict[str, set[str]],
) -> dict[str, int]:
    backup_file(full_workbook, backup_root / "full_workbook")
    wb = load_workbook(full_workbook)
    stats: defaultdict[str, int] = defaultdict(int)
    group_refs = collect_referenced_group_ids(wb, target_contexts)
    new_group_ids = duplicate_sourcegroups_for_splits(wb, group_refs, stats)
    patch_sheet_rows(wb, target_contexts, new_group_ids, stats)

    temp_path = full_workbook.with_name(f"{full_workbook.stem}.tmp{full_workbook.suffix}")
    wb.save(temp_path)
    wb.close()
    validate_workbook(temp_path)
    shutil.move(temp_path, full_workbook)
    return dict(stats)


def patch_json_object(node: Any, assignment: SplitAssignment) -> int:
    updates = 0
    if isinstance(node, dict):
        candidate_name = normalize_space(node.get("candidateName"))
        if not candidate_name and ("Firstname" in node or "Surname" in node):
            candidate_name = normalize_space(f"{node.get('Firstname', '')} {node.get('Surname', '')}")
        for key, value in list(node.items()):
            if key in {"Candidate_Id", "SourcePersonID", "PersonID"} and candidate_name == assignment.split_name:
                if str(value) == str(assignment.old_id):
                    node[key] = str(assignment.new_id) if isinstance(value, str) else assignment.new_id
                    updates += 1
            else:
                updates += patch_json_object(value, assignment)
    elif isinstance(node, list):
        for item in node:
            updates += patch_json_object(item, assignment)
    return updates


def patch_json_files(root: Path, backup_root: Path) -> dict[str, int]:
    stats: defaultdict[str, int] = defaultdict(int)
    backup_dir = backup_root / "website_json"
    for path in root.rglob("*.json"):
        original_text = path.read_text(encoding="utf-8")
        data = json.loads(original_text)
        file_updates = 0
        for assignment in ASSIGNMENTS:
            file_updates += patch_json_object(data, assignment)
        if file_updates:
            backup_target = backup_dir / path.relative_to(root)
            backup_target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(path, backup_target)
            path.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
            stats["json_files_touched"] += 1
            stats["json_id_updates"] += file_updates
    return dict(stats)


def scan_local_workbook(local_workbook: Path) -> dict[str, int]:
    wb = load_workbook(local_workbook, read_only=True, data_only=True)
    collision_ids = {assignment.old_id for assignment in ASSIGNMENTS}
    stats: defaultdict[str, int] = defaultdict(int)
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(isinstance(value, int) and value in collision_ids for value in row):
                stats[f"{ws_name}_collision_id_rows"] += 1
    wb.close()
    return dict(stats)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--full-workbook", default="Full election tables.xlsx")
    parser.add_argument("--website-json-root", default=r"election-viewer-package\data\elections")
    parser.add_argument("--local-workbook", default=r"_tmp_xls2rar_extract\out\wiki_lgov_modern\lgov-modern-wikipedia.stvfix.xlsx")
    args = parser.parse_args()

    backup_root = Path("backups") / f"personid-collision-fix-{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}"
    full_workbook = Path(args.full_workbook)
    target_contexts = load_target_contexts(full_workbook)
    workbook_stats = patch_workbook(full_workbook, backup_root, target_contexts)
    json_stats = patch_json_files(Path(args.website_json_root), backup_root)
    local_stats = scan_local_workbook(Path(args.local_workbook))
    print(
        json.dumps(
            {
                "backup_root": str(backup_root),
                "assignments": [assignment.__dict__ for assignment in ASSIGNMENTS],
                "full_workbook": workbook_stats,
                "website_json": json_stats,
                "local_workbook_scan": local_stats,
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
