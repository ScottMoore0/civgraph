#!/usr/bin/env python
"""Build pre-1922 Westminster election JSON files (1885-1918) from
parlconst's 7 ROI ERT sheets, plus update elections_index.json and
_pre1970_index.json with the new dates.

Outputs:
  data/external/parlconst/pre1922_westminster_results.csv  (unified table)
  election-viewer-package/data/elections/house-of-commons-of-the-united-kingdom/<date>/<slug>.json
  election-viewer-package/data/elections/house-of-commons-of-the-united-kingdom/_pre1970_index.json (updated)
  election-viewer-package/data/elections_index.json (updated)
"""
import csv, json, re, sys, io, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
EVP = REPO / "election-viewer-package"
ELECTIONS_DIR = EVP / "data" / "elections" / "house-of-commons-of-the-united-kingdom"
INDEX = ELECTIONS_DIR / "_pre1970_index.json"
ELECTIONS_INDEX = EVP / "data" / "elections_index.json"
OUT_CSV = REPO / "data" / "external" / "parlconst" / "pre1922_westminster_results.csv"

# 9 GEs with canonical polling dates
GE_DATES = {
    '1885':     '1885-11-24',
    '1886':     '1886-07-01',
    '1892':     '1892-07-04',
    '1895':     '1895-07-13',
    '1900':     '1900-09-26',
    '1906':     '1906-01-12',
    '1910 (J)': '1910-01-15',
    '1910 (D)': '1910-12-03',
    '1918':     '1918-12-14',
}

# Party code → name, colour
PARTY_DEFS = {
    'N':    {'name': 'Irish Parliamentary Party', 'colour': '#0E7C42'},
    'SF':   {'name': 'Sinn Féin',                 'colour': '#326760'},
    'PN':   {'name': 'Parnellite Nationalist',    'colour': '#1D9656'},
    'IndN': {'name': 'Independent Nationalist',   'colour': '#5DBC8E'},
    'U':    {'name': 'Irish Unionist Alliance',   'colour': '#1F4E8C'},
    'LU':   {'name': 'Liberal Unionist',          'colour': '#E0A60C'},
    'C':    {'name': 'Conservative',              'colour': '#0087DC'},
    'L':    {'name': 'Liberal',                   'colour': '#FAA61A'},
    'Lab':  {'name': 'Labour',                    'colour': '#E4003B'},
    'Oth':  {'name': 'Other',                     'colour': '#888888'},
}

ROI_SECTIONS = ['105','106','107','108','109','110','111']


def slugify(name):
    name = unicodedata.normalize('NFKD', name).encode('ascii','ignore').decode('ascii')
    s = re.sub(r"[^a-zA-Z0-9]+", '-', name).strip('-').lower()
    s = re.sub(r'-+', '-', s)
    return s or 'unknown'


def normalize_code(c):
    if c is None: return None
    c = str(c).strip()
    if not c: return None
    # Whitespace variants
    c = c.replace(' ', '')
    return c


def collect_results():
    """Walk all 7 ERT sheets. For each (constituency, year-column), read
    the party code. Returns list of (date, constituency, party_code)."""
    out = []
    constituencies_per_section = {}
    for section in ROI_SECTIONS:
        xlsx = next(Path('_tmp_parlconst/files').glob(f'{section}_other_dl_*.xlsx'))
        wb = load_workbook(xlsx, data_only=True)
        ws = wb['ERT']
        rows = list(ws.iter_rows(values_only=True))
        # Header row 2 (index 1) — columns D..L for 9 years
        header = rows[1]
        # Map year-label → column index
        year_cols = {}
        for i, cell in enumerate(header):
            if cell is None: continue
            s = str(cell).strip()
            if s in GE_DATES:
                year_cols[s] = i
        # Process data rows from row 3 (index 2) until we hit Total/empty
        constits = set()
        for r in rows[2:]:
            if r[1] is None: break
            constit = str(r[1]).strip()
            if not constit or constit.startswith('Total'): break
            constits.add(constit)
            for year, col in year_cols.items():
                code = normalize_code(r[col])
                if code:
                    out.append((GE_DATES[year], constit, code))
        constituencies_per_section[section] = constits
    return out, constituencies_per_section


def make_constituency_json(constit, code, date, body):
    party = PARTY_DEFS.get(code, {'name': code, 'colour': '#888888'})
    return {
        "Constituency": {
            "countInfo": {
                "Constituency_Name": constit,
                "Constituency_Number": "",
                "Number_Of_Seats": "1",
                "Spoiled": "",
                "Total_Electorate": "",
                "Total_Poll": "",
                "Valid_Poll": "",
            },
            "countGroup": [{
                "Candidate_First_Pref_Votes": "Unknown",
                "Candidate_Id": "",
                "Constituency_Number": "",
                "Count_Number": "1",
                "Firstname": "",
                "Occurred_On_Count": "",
                "Party_Colour": party['colour'],
                "Party_Name": party['name'],
                "Status": "Elected",
                "Surname": "",
                "Total_Votes": "",
                "Transfers": "0.00",
                "candidateName": "(candidate name not in source)",
                "id": 0,
            }],
        }
    }


def main():
    print("Collecting results from 7 ROI ERTs ...")
    results, by_section = collect_results()
    print(f"  {len(results)} (date, constituency, code) entries across {len(set(r[1] for r in results))} constituencies")

    # Code distribution
    from collections import Counter
    code_counts = Counter(r[2] for r in results)
    print(f"  party codes: {dict(code_counts.most_common())}")

    # Write unified CSV
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open('w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        w.writerow(['date','constituency','party_code','party_name'])
        for date, c, code in results:
            party = PARTY_DEFS.get(code, {'name': code})
            w.writerow([date, c, code, party['name']])
    print(f"\nwrote {OUT_CSV}  ({len(results)} rows)")

    # Group by date for JSON emission
    by_date = {}
    for date, constit, code in results:
        by_date.setdefault(date, {})[constit] = code

    # Emit per-constituency JSON files per date
    print("\nEmitting per-constituency JSON files ...")
    file_count = 0
    pre1970_index_entries = []
    for date in sorted(by_date.keys()):
        d_dir = ELECTIONS_DIR / date
        d_dir.mkdir(parents=True, exist_ok=True)
        constits_for_date = sorted(by_date[date].keys())
        for constit in constits_for_date:
            code = by_date[date][constit]
            slug = slugify(constit)
            obj = make_constituency_json(constit, code, date, 'House of Commons of the United Kingdom')
            (d_dir / f'{slug}.json').write_text(
                json.dumps(obj, indent=2, ensure_ascii=False) + '\n', encoding='utf-8')
            file_count += 1
        pre1970_index_entries.append({
            'date': date,
            'constituencies': constits_for_date,
        })
    print(f"  wrote {file_count} JSON files across {len(by_date)} dates")

    # Update _pre1970_index.json — merge with existing entries, preferring new
    existing = []
    if INDEX.exists():
        try:
            existing = json.loads(INDEX.read_text(encoding='utf-8'))
        except Exception:
            existing = []
    # Merge by date — for the new dates we replace; for old dates we keep existing
    new_dates = {e['date'] for e in pre1970_index_entries}
    merged = [e for e in existing if e.get('date') not in new_dates] + pre1970_index_entries
    merged.sort(key=lambda e: e.get('date', ''))
    INDEX.write_text(json.dumps(merged, indent=1, ensure_ascii=False) + '\n', encoding='utf-8')
    print(f"  updated {INDEX} ({len(merged)} entries)")


if __name__ == "__main__":
    main()
