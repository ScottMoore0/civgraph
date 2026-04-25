#!/usr/bin/env python
"""Convert pre-1970 NI Westminster results from the BK (Sydney Elliott)
spreadsheet into per-constituency JSON files matching the existing
'House of Commons of the United Kingdom' schema on the site.

Source: bkniresults.xlsx in repo root.
Output: election-viewer-package/data/elections/house-of-commons-of-the-united-kingdom/<date>/<slug>.json

Each sheet covers one election year. Within a sheet, contests are stacked
vertically. Each block starts with a title row of the form
"<Year> (<Day Month>) - <Constituency>: Electorate: X Total Poll: Y Valid Poll: Z"
or, after the first block in a sheet, just
"<Constituency>: Electorate: X Total Poll: Y Valid Poll: Z"
followed by a blank row, a "Candidate / Party / Votes" header row, the
candidate rows, and a totals + majority row.

1922-1929: NI used PR-STV with multi-member constituencies; this dataset
records first-preference counts only. Candidates are sorted by FP votes
and the top-N are flagged as elected (N = seats for that constituency).
1929-1969: single-member FPTP; top candidate elected.
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
SRC_XLSX = REPO / "bkniresults.xlsx"
OUT_BASE = REPO / "election-viewer-package" / "data" / "elections" / "house-of-commons-of-the-united-kingdom"

# Sheet -> ISO date for the election (BK headers in title row are authoritative;
# this map lets us locate the output folder before reading the sheet).
SHEETS_TO_DATE = {
    "1922":   "1922-11-15",
    "1923":   "1923-12-06",
    "1924":   "1924-10-29",
    "1929":   "1929-05-30",
    "1931":   "1931-10-27",
    "1935":   "1935-11-14",
    "1945":   "1945-07-05",
    "1950":   "1950-02-23",
    "1951":   "1951-10-25",
    "1955":   "1955-05-26",
    "1959":   "1959-10-08",
    "1964":   "1964-10-15",
    "1966":   "1966-03-31",
}

# 1922-1929 used PR-STV multi-member constituencies. Seat counts per
# constituency (from Wikipedia / Northern Ireland (Boundaries) Act 1922).
STV_SEATS_1922_1929 = {
    "Antrim": 2, "Armagh": 1, "Belfast East": 1, "Belfast North": 1,
    "Belfast South": 1, "Belfast West": 1, "Down": 2, "Fermanagh and Tyrone": 4,
    "Londonderry": 1,
}
# Default seats for unrecognised constituencies in those years
DEFAULT_STV_SEATS = 1

# Party label canonicalisation. BK uses short codes; we expand to full names so
# the existing party-colour table picks up the right shade.
PARTY_LABEL = {
    "U":     "Ulster Unionist Party",
    "UUP":   "Ulster Unionist Party",
    "OUP":   "Ulster Unionist Party",
    "DUP":   "Democratic Unionist Party",
    "UPNI":  "Unionist Party of Northern Ireland",
    "Ind U.":"Independent Unionist",
    "Ind U": "Independent Unionist",
    "IndU":  "Independent Unionist",
    "Ind":   "Independent",
    "Ind L": "Independent Labour",
    "Nat":   "Nationalist Party",
    "SF":    "Sinn Féin",
    "SDLP":  "SDLP",
    "All":   "Alliance",
    "Alliance":"Alliance",
    "PD":    "People's Democracy",
    "Lab":   "Northern Ireland Labour Party",
    "NILP":  "Northern Ireland Labour Party",
    "Soc R": "Republican Labour Party",
    "RL":    "Republican Labour Party",
    "Rep":   "Republican Clubs",
    "Co":    "Conservative",
    "Con":   "Conservative",
    "U (O'N)":"Ulster Unionist Party (Pro-O'Neill)",
    "U (Anti-O'N)":"Ulster Unionist Party (Anti-O'Neill)",
    "VUP":   "Vanguard Unionist Progressive Party",
    "WP":    "Workers' Party",
    "Loyalist":"Loyalist",
    "Unity": "Unity",
    "IIP":   "Irish Independence Party",
    "GP":    "Green Party",
    "Oth":   "Other",
}

# Curated colour palette for party labels (subset of scrape_old_lgov_wikipedia.py
# PARTY_COLOURS, expanded with historical-only labels).
PARTY_COLOURS = {
    "Ulster Unionist Party": "#48A5EE",
    "Ulster Unionist Party (Pro-O'Neill)": "#9CC6E8",
    "Ulster Unionist Party (Anti-O'Neill)": "#1E5C9E",
    "Democratic Unionist Party": "#D46A4C",
    "Unionist Party of Northern Ireland": "#FFA07A",
    "Independent Unionist": "#AADFFF",
    "Independent": "#DCDCDC",
    "Independent Labour": "#FF9999",
    "Independent Nationalist": "#CDFFAB",
    "Nationalist Party": "#32CD32",
    "Sinn Féin": "#326760",
    "SDLP": "#2AA82C",
    "Alliance": "#F6CB2F",
    "People's Democracy": "#FF0000",
    "Northern Ireland Labour Party": "#DC241F",
    "Republican Labour Party": "#85DE59",
    "Republican Clubs": "#930C1A",
    "Conservative": "#0087DC",
    "Vanguard Unionist Progressive Party": "#FF8C00",
    "Workers' Party": "#930C1A",
    "Loyalist": "#FFD700",
    "Unity": "#90EE90",
    "Irish Independence Party": "#228B22",
    "Green Party": "#8DC63F",
    "Other": "#999999",
}


def slugify(s: str) -> str:
    s = s.lower().replace("&", "and")
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"\s+", "-", s.strip())
    return s


# Constituency display-name canonicalisation: BK forms → site-canonical names
# matching the FGB Name attribute and the existing 1970+ election folders.
# BK uses inconsistent spellings ("West Belfast" / "Belfast West"; "Derry" /
# "Londonderry"; "&" vs "and"). Normalise to the post-1970 canonical labels.
CONST_CANONICAL = {
    "fermanagh and tyrone":          "Fermanagh and Tyrone",
    "fermanagh and south tyrone":    "Fermanagh and South Tyrone",
    "fermanagh & tyrone":            "Fermanagh and Tyrone",
    "fermanagh & south tyrone":      "Fermanagh and South Tyrone",
    "derry":                         "Londonderry",
    "londonderry":                   "Londonderry",
    "mid-ulster":                    "Mid Ulster",
    "mid ulster":                    "Mid Ulster",
    "queen's university belfast":    "Queen's University of Belfast",
    "queen's university":            "Queen's University of Belfast",
    "queens university":             "Queen's University of Belfast",
    "queens university belfast":     "Queen's University of Belfast",
    "queens university of belfast":  "Queen's University of Belfast",
    "antrim":                        "Antrim",
    "armagh":                        "Armagh",
    "down":                          "Down",
    "north down":                    "North Down",
    "south down":                    "South Down",
    "north antrim":                  "North Antrim",
    "south antrim":                  "South Antrim",
    "mid antrim":                    "Mid Antrim",
    # BK reverses the "Belfast X" convention for some sheets.
    "east belfast":                  "Belfast East",
    "north belfast":                 "Belfast North",
    "south belfast":                 "Belfast South",
    "west belfast":                  "Belfast West",
    "belfast east":                  "Belfast East",
    "belfast north":                 "Belfast North",
    "belfast south":                 "Belfast South",
    "belfast west":                  "Belfast West",
}


def canonicalise_constituency(raw: str) -> str:
    key = raw.strip().lower().rstrip(":").strip()
    # Normalise stray whitespace and ampersands for lookup
    key = re.sub(r"\s+", " ", key).replace("&", "and")
    return CONST_CANONICAL.get(key, raw.strip().rstrip(":").strip())


def parse_title_row(s: str, year: str):
    """Extract constituency name + electorate/poll/valid_poll from the title row.

    Examples:
      '2010 (6 May) - Fermanagh & South Tyrone: Electorate: 67,908 ...'
      'West Tyrone: Electorate: 61,148 ...'
      '1929 (30 May) - Fermanagh & Tyrone (4 seats)'
      'Antrim (2 seats):'
      'Derry:'                                       (uncontested, no metadata)
      'East Belfast (12 February 1943):'             (by-election within sheet)
    """
    s = s.strip()
    # Strip leading "<year> (<date>) - " if present
    m = re.match(r"^\d{4}\s*\([^)]+\)\s*-\s*(.+)$", s)
    if m:
        s = m.group(1).strip()
    # Split off ": Electorate / Total Poll / Valid Poll" trailing block
    if ":" in s:
        const, rest = s.split(":", 1)
    else:
        const, rest = s, ""
    const = const.strip()

    # Detect by-election dates in parens: "Constituency (12 February 1943)" or
    # "Constituency (24 January 1986)". These appear inside parent-year sheets
    # for interim contests. Capture as by-election date.
    by_date = None
    m = re.search(r"\((\d{1,2}\s+[A-Za-z]+\s+(?:19|20)\d{2})\)$", const)
    if m:
        by_date = m.group(1)
        const = const[:m.start()].strip()

    # Detect multi-member seat annotation: "Constituency (N Seats)"
    seats = 1
    m = re.search(r"\((\d+)\s*[Ss]eats?\)$", const)
    if m:
        seats = int(m.group(1))
        const = const[:m.start()].strip()

    # Numeric metadata from the post-colon segment
    def grab(label):
        m = re.search(label + r":\s*([\d,]+)", rest)
        return m.group(1).replace(",", "") if m else ""
    return {
        "constituency": const,
        "seats":        seats,
        "by_date":      by_date,
        "electorate":   grab("Electorate"),
        "total_poll":   grab("Total Poll"),
        "valid_poll":   grab("Valid Poll"),
    }


def parse_block(rows, header_idx, year):
    """Parse one constituency block beginning at the 'Candidate' header row."""
    # Find the title row: typically 1-2 rows above, may be the sheet's row 0.
    title_idx = None
    for k in (header_idx - 2, header_idx - 1):
        if k < 0: continue
        r = rows[k]
        if r and r[0] is not None and isinstance(r[0], str) and r[0].strip():
            s = r[0].strip()
            if s.lower() != "candidate":
                title_idx = k
                # Prefer the longer one (the actual title) if both rows have text.
    if title_idx is None:
        return None
    title = parse_title_row(rows[title_idx][0], year)
    # Read candidates: name in col0, party in col1, votes in col2
    candidates = []
    i = header_idx + 1
    while i < len(rows):
        r = rows[i]
        if not r or all(c is None for c in r):
            break
        name = r[0]
        if name is None:
            break
        s = str(name).strip()
        # Stop on totals row (col0 empty, col2 has total) — but we already broke on all-None.
        if s == "" or s.lower() in ("totals", "total"): break
        # Stop on "Majority" line (col0 is None usually but col4=='Majority').
        if len(r) > 4 and isinstance(r[4], str) and r[4].strip().lower().startswith("majority"):
            break
        # Stop on next title (a row that contains ":" and Electorate keyword)
        if "Electorate:" in s or re.search(r"^\d{4}\s*\(", s):
            break
        if s.lower() == "candidate":
            break
        party_raw = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        votes = r[2] if len(r) > 2 else None
        candidates.append({
            "name_raw":  s,
            "party_raw": party_raw,
            "votes":     votes,
        })
        i += 1

    return {"title": title, "candidates": candidates, "block_end": i}


def parse_name(name_raw):
    """Split 'Firstname Surname*' into (first, last, elected)."""
    s = name_raw.strip()
    elected = s.endswith("*")
    if elected: s = s[:-1].strip()
    # Some names contain commas, e.g. "Brady, Anthony Kevin"; flip to "First Last".
    if "," in s:
        last, first = s.split(",", 1)
        return first.strip(), last.strip(), elected
    parts = s.split(maxsplit=1)
    if len(parts) == 2:
        return parts[0], parts[1], elected
    return s, "", elected


def normalise_party(raw):
    raw = raw.strip()
    return PARTY_LABEL.get(raw, raw)


def party_colour(name):
    return PARTY_COLOURS.get(name, "#A0A0A0")


def candidate_id(year, slug, last, first):
    """Stable-ish ID per candidate. Same scheme as ARK converter."""
    base = f"{year}|{slug}|{last}|{first}".lower()
    h = 0
    for ch in base:
        h = (h * 131 + ord(ch)) & 0xFFFFFF
    return str(h)


def emit_json(year, sheet_name, block, out_dir, multi_member_seats):
    title = block["title"]
    const_canonical = canonicalise_constituency(title["constituency"])
    slug = slugify(const_canonical)
    # Trust the parsed "(N seats)" annotation when present; fall back to the
    # 1922-style multi-member map.
    seats = title.get("seats") or 1
    if seats == 1 and year in ("1922","1923","1929"):
        seats = multi_member_seats.get(const_canonical, 1)
    # Sort candidates by votes (numeric desc; "Unopposed" → infinity equivalent).
    def vkey(c):
        v = c["votes"]
        if isinstance(v, (int, float)): return -float(v)
        if isinstance(v, str) and v.lower().startswith("unopposed"): return -1e15
        try: return -float(v)
        except: return 0
    cands = sorted(block["candidates"], key=vkey)
    elected_n = seats
    countGroup = []
    valid_total = 0
    for idx, c in enumerate(cands):
        first, last, _ = parse_name(c["name_raw"])
        party = normalise_party(c["party_raw"])
        v = c["votes"]
        if isinstance(v, (int, float)):
            fp = f"{float(v):.2f}"
            valid_total += float(v)
        elif isinstance(v, str) and v.lower().startswith("unopposed"):
            fp = "Unopposed"
        else:
            try:
                fp_val = float(v)
                fp = f"{fp_val:.2f}"
                valid_total += fp_val
            except (TypeError, ValueError):
                fp = ""
        countGroup.append({
            "Candidate_First_Pref_Votes": fp,
            "Candidate_Id": candidate_id(year, slug, last, first),
            "Constituency_Number": "",
            "Count_Number": "1",
            "Firstname": first,
            "Occurred_On_Count": "",
            "Party_Colour": party_colour(party),
            "Party_Name": party,
            "Status": "Elected" if idx < elected_n else "",
            "Surname": last,
            "Total_Votes": fp if fp != "Unopposed" else "",
            "Transfers": "0.00",
            "candidateName": (first + " " + last).strip(),
            "id": idx,
        })

    # Compute Spoiled if total_poll and valid_poll available.
    tp = title["total_poll"]; vp = title["valid_poll"]
    spoiled = ""
    if tp and vp:
        try: spoiled = str(int(tp) - int(vp))
        except: spoiled = ""

    return {
        "Constituency": {
            "countInfo": {
                "Constituency_Name": const_canonical,
                "Constituency_Number": "",
                "Number_Of_Seats": str(seats),
                "Spoiled": spoiled,
                "Total_Electorate": title["electorate"],
                "Total_Poll": tp,
                "Valid_Poll": vp,
            },
            "countGroup": countGroup,
        }
    }, slug


_MONTHS = {m: i for i, m in enumerate(
    ["January","February","March","April","May","June",
     "July","August","September","October","November","December"], start=1)}

def parse_by_date(s):
    """Parse '12 February 1943' -> '1943-02-12'. Returns None on failure."""
    if not s: return None
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})$", s.strip())
    if not m: return None
    day = int(m.group(1))
    mon = _MONTHS.get(m.group(2).capitalize())
    year = int(m.group(3))
    if mon is None: return None
    return f"{year:04d}-{mon:02d}-{day:02d}"


def convert_sheet(wb, sheet_name, year_label):
    """Return ([(slug, json, const_name)], [(by_iso, slug, json, const_name)]).
    First list is GE results; second is by-elections from this sheet.
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    ge = []
    by = []
    header_idxs = [i for i, r in enumerate(rows)
                   if r and r[0] is not None and str(r[0]).strip() == "Candidate"]
    for hi in header_idxs:
        block = parse_block(rows, hi, year_label)
        if not block or not block["candidates"]:
            continue
        by_date_human = block["title"].get("by_date")
        seats_map = STV_SEATS_1922_1929 if year_label in ("1922","1923","1929") else {}
        if by_date_human:
            iso = parse_by_date(by_date_human)
            if iso is None:
                print(f"    ! couldn't parse by-election date '{by_date_human}'")
                continue
            # By-elections are always single-member here.
            json_obj, slug = emit_json(year_label, sheet_name, block, None, seats_map)
            json_obj["Constituency"]["countInfo"]["Number_Of_Seats"] = "1"
            const_name = json_obj["Constituency"]["countInfo"]["Constituency_Name"]
            by.append((iso, slug, json_obj, const_name))
        else:
            json_obj, slug = emit_json(year_label, sheet_name, block, None, seats_map)
            const_name = json_obj["Constituency"]["countInfo"]["Constituency_Name"]
            ge.append((slug, json_obj, const_name))
    if by:
        print(f"    ({len(by)} by-election(s) emitted)")
    return ge, by


def main():
    if not SRC_XLSX.exists():
        print(f"  ! source missing: {SRC_XLSX}")
        sys.exit(1)
    wb = load_workbook(str(SRC_XLSX), data_only=True, read_only=True)
    summary = []
    by_summary = {}  # iso_date -> list of constituency names
    for sheet, date in SHEETS_TO_DATE.items():
        if sheet not in wb.sheetnames:
            print(f"  ! sheet {sheet} not in workbook")
            continue
        out_dir = OUT_BASE / date
        out_dir.mkdir(parents=True, exist_ok=True)
        ge_results, by_results = convert_sheet(wb, sheet, sheet)
        const_names = []
        for slug, obj, const in ge_results:
            (out_dir / f"{slug}.json").write_text(
                json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")
            const_names.append(const)
        unique = sorted(set(const_names))
        summary.append({"date": date, "constituencies": unique})
        # By-elections — write each to its own date folder
        for by_iso, slug, obj, const in by_results:
            by_dir = OUT_BASE / by_iso
            by_dir.mkdir(parents=True, exist_ok=True)
            (by_dir / f"{slug}.json").write_text(
                json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")
            by_summary.setdefault(by_iso, []).append(const)
        print(f"  {sheet} -> {date}: {len(ge_results)} GE contests ({len(unique)} unique), "
              f"{len(by_results)} by-election(s)")
    # Write summaries
    (OUT_BASE / "_pre1970_index.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    by_payload = [{"date": iso, "constituencies": sorted(set(names))}
                  for iso, names in sorted(by_summary.items())]
    (OUT_BASE / "_pre1970_byelections_index.json").write_text(
        json.dumps(by_payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nWrote GE summary + {len(by_payload)} by-election dates")

if __name__ == "__main__":
    main()
