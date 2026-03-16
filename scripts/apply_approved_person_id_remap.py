"""Apply approved person-ID mappings from the match workbook to the modern local workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def normalize_approved(value) -> str:
    return str(value or "").strip().upper()


def load_approved_mapping(match_workbook: Path) -> dict[int, int]:
    wb = load_workbook(match_workbook, read_only=True, data_only=True)
    ws = wb["AttemptedMatches"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: pos for pos, header in enumerate(headers)}
    mapping: dict[int, int] = {}
    reverse: dict[int, int] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if normalize_approved(row[idx["approved"]]) != "Y":
            continue
        full_id = row[idx["full_person_id"]]
        modern_id = row[idx["modern_person_id"]]
        if not full_id or not modern_id:
            continue
        full_id = int(full_id)
        modern_id = int(modern_id)
        existing_full = mapping.get(modern_id)
        existing_modern = reverse.get(full_id)
        if existing_full is not None and existing_full != full_id:
            raise ValueError(f"Conflicting full ID for modern ID {modern_id}: {existing_full} vs {full_id}")
        if existing_modern is not None and existing_modern != modern_id:
            raise ValueError(f"Conflicting modern ID for full ID {full_id}: {existing_modern} vs {modern_id}")
        mapping[modern_id] = full_id
        reverse[full_id] = modern_id
    return mapping


def remap_scalar(value, mapping: dict[int, int]):
    if value is None or value == "":
        return value, False
    try:
        key = int(value)
    except (TypeError, ValueError):
        return value, False
    new_value = mapping.get(key)
    if new_value is None or new_value == key:
        return value, False
    return new_value, True


def remap_id_list(value, mapping: dict[int, int]):
    if value is None or value == "":
        return value, False
    parts = [part.strip() for part in str(value).split(",")]
    changed = False
    remapped: list[str] = []
    for part in parts:
        if not part:
            remapped.append(part)
            continue
        try:
            key = int(part)
        except ValueError:
            remapped.append(part)
            continue
        new_value = mapping.get(key, key)
        if new_value != key:
            changed = True
        remapped.append(str(new_value))
    return ", ".join(remapped), changed


def apply_mapping(input_workbook: Path, output_workbook: Path, mapping: dict[int, int]) -> dict[str, int]:
    wb = load_workbook(input_workbook)
    stats = {
        "electionresults_scalar_id_updates": 0,
        "electionresults_transfersubject_updates": 0,
        "transfers_scalar_id_updates": 0,
        "transfers_remainingcandidateids_updates": 0,
    }

    ws_results = wb["ElectionResults"]
    result_headers = [cell.value for cell in next(ws_results.iter_rows(min_row=1, max_row=1))]
    result_idx = {header: pos + 1 for pos, header in enumerate(result_headers)}
    result_scalar_headers = [header for header in result_headers if header == "PersonID"]
    result_transfer_subject_headers = [
        header
        for header in result_headers
        if isinstance(header, str) and header.startswith("TransferSubject")
    ]
    for row_num in range(2, ws_results.max_row + 1):
        for header in result_scalar_headers:
            cell = ws_results.cell(row=row_num, column=result_idx[header])
            new_value, changed = remap_scalar(cell.value, mapping)
            if changed:
                cell.value = new_value
                stats["electionresults_scalar_id_updates"] += 1
        for header in result_transfer_subject_headers:
            cell = ws_results.cell(row=row_num, column=result_idx[header])
            new_value, changed = remap_scalar(cell.value, mapping)
            if changed:
                cell.value = new_value
                stats["electionresults_transfersubject_updates"] += 1

    ws_transfers = wb["Transfers"]
    transfer_headers = [cell.value for cell in next(ws_transfers.iter_rows(min_row=1, max_row=1))]
    transfer_idx = {header: pos + 1 for pos, header in enumerate(transfer_headers)}
    transfer_scalar_headers = [
        header for header in transfer_headers if header in {"PersonID", "SourcePersonID"}
    ]
    for row_num in range(2, ws_transfers.max_row + 1):
        for header in transfer_scalar_headers:
            cell = ws_transfers.cell(row=row_num, column=transfer_idx[header])
            new_value, changed = remap_scalar(cell.value, mapping)
            if changed:
                cell.value = new_value
                stats["transfers_scalar_id_updates"] += 1

        remaining_cell = ws_transfers.cell(row=row_num, column=transfer_idx["RemainingCandidateIDsDesc"])
        new_value, changed = remap_id_list(remaining_cell.value, mapping)
        if changed:
            remaining_cell.value = new_value
            stats["transfers_remainingcandidateids_updates"] += 1

    wb.save(output_workbook)
    return stats


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--match-workbook", default=r"_tmp_xls2rar_extract\out\wiki_lgov_modern\person-id-match.xlsx")
    parser.add_argument("--modern-workbook", default=r"_tmp_xls2rar_extract\out\wiki_lgov_modern\lgov-modern-wikipedia.stvfix.xlsx")
    parser.add_argument("--output-workbook", default=None)
    args = parser.parse_args()

    mapping = load_approved_mapping(Path(args.match_workbook))
    input_path = Path(args.modern_workbook)
    output_path = Path(args.output_workbook) if args.output_workbook else input_path
    stats = apply_mapping(input_path, output_path, mapping)
    print({"approved_mapping_count": len(mapping), **stats})


if __name__ == "__main__":
    main()
