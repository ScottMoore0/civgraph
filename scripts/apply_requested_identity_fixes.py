"""Apply requested split/merge person-ID corrections across workbook, JSON, and modern local workbook."""

from __future__ import annotations

import json
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class SplitContextAssignment:
    label: str
    old_id: int
    new_id: int
    name: str
    contexts: tuple[tuple[str, str, str], ...]  # date, constituency, party
    website_json_files: tuple[str, ...]


SPLITS: tuple[SplitContextAssignment, ...] = (
    SplitContextAssignment(
        label="Austin Kelly Workers Party / Republican Clubs",
        old_id=63215,
        new_id=100008,
        name="Austin Kelly",
        contexts=(
            ("1982-10-20", "South Antrim", "Workers Party / Republican Clubs"),
            ("1983-06-09", "East Antrim", "Workers Party / Republican Clubs"),
            ("1987-06-11", "East Antrim", "Workers Party / Republican Clubs"),
            ("1996-05-30", "East Antrim", "Workers Party / Republican Clubs"),
        ),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-assembly/1982-10-20/south-antrim.json",
            r"election-viewer-package/data/elections/house-of-commons-of-the-united-kingdom/1983-06-09/east-antrim.json",
            r"election-viewer-package/data/elections/house-of-commons-of-the-united-kingdom/1987-06-11/east-antrim.json",
            r"election-viewer-package/data/elections/northern-ireland-forum-for-political-dialogue/1996-05-30/east-antrim.json",
        ),
    ),
    SplitContextAssignment(
        label="John Stewart NI Labour",
        old_id=8210,
        new_id=100009,
        name="John Stewart",
        contexts=(
            ("1973-06-28", "Belfast North", "NI Labour"),
            ("1975-05-01", "Belfast North", "NI Labour"),
        ),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-assembly/1973-06-28/belfast-north.json",
            r"election-viewer-package/data/elections/northern-ireland-constitutional-convention/1975-05-01/belfast-north.json",
        ),
    ),
    SplitContextAssignment(
        label="Peter Lavery Natural Law",
        old_id=58217,
        new_id=100010,
        name="Peter Lavery",
        contexts=(
            ("1996-05-30", "Fermanagh and South Tyrone", "Natural Law"),
        ),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-forum-for-political-dialogue/1996-05-30/fermanagh-and-south-tyrone.json",
        ),
    ),
    SplitContextAssignment(
        label="Richard Stewart Independent (Alan Chambers)",
        old_id=17230,
        new_id=100011,
        name="Richard Stewart",
        contexts=(
            ("1996-05-30", "Belfast West", "Independent (Alan Chambers)"),
        ),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-forum-for-political-dialogue/1996-05-30/belfast-west.json",
        ),
    ),
)


LOCAL_NAME_TO_ID = {
    "Austin Kelly": 63215,
    "Catherine Nelson": 70028,
    "Charlotte Carson": 100006,
    "Donal O'Cofaigh": 100001,
    "Gavin Malone": 100003,
    "John Stewart": 8210,
    "Paddy Meehan": 33653,
    "Peter Lavery": 58217,
    "Richard Stewart": 17230,
    "Stephen Dunne": 100002,
}


FULL_WORKBOOK = Path("Full election tables.xlsx")
LOCAL_WORKBOOK = Path(r"_tmp_xls2rar_extract/out/wiki_lgov_modern/lgov-modern-wikipedia.stvfix.xlsx")


def backup_file(src: Path, backup_root: Path) -> None:
    backup_root.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, backup_root / src.name)


def normalize_space(value: Any) -> str:
    return " ".join(str(value or "").split())


def format_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def replace_exact_id_token(text: str, old_id: int, new_id: int) -> tuple[str, bool]:
    old = str(old_id)
    new = str(new_id)
    updated: list[str] = []
    changed = False
    token = ""
    for ch in text:
        if ch.isdigit():
            token += ch
            continue
        if token:
            if token == old:
                updated.append(new)
                changed = True
            else:
                updated.append(token)
            token = ""
        updated.append(ch)
    if token:
        if token == old:
            updated.append(new)
            changed = True
        else:
            updated.append(token)
    return "".join(updated), changed


def replace_cell_value(value: Any, old_id: int, new_id: int) -> tuple[Any, bool]:
    if value is None or value == "":
        return value, False
    if isinstance(value, bool):
        return value, False
    if isinstance(value, int):
        return (new_id, True) if value == old_id else (value, False)
    if isinstance(value, float):
        if value.is_integer() and int(value) == old_id:
            return float(new_id), True
        return value, False
    if isinstance(value, str):
        return replace_exact_id_token(value, old_id, new_id)
    return value, False


def context_key(date_text: str, constituency: str, party: str) -> str:
    return f"{date_text}|{normalize_space(constituency)}|{normalize_space(party)}"


def derive_row_context(row_values: dict[str, Any]) -> str | None:
    date_text = format_date(row_values.get("Date"))
    constituency = row_values.get("Constituency")
    party = row_values.get("Party Name") or row_values.get("Source Party Name")
    if date_text and constituency and party:
        return context_key(date_text, str(constituency), str(party))
    return None


def build_target_context_sets() -> dict[str, set[str]]:
    result: dict[str, set[str]] = {}
    for assignment in SPLITS:
        result[assignment.label] = {
            context_key(date_text, constituency, party)
            for date_text, constituency, party in assignment.contexts
        }
    return result


def patch_full_workbook(path: Path, backup_root: Path) -> dict[str, int]:
    backup_file(path, backup_root)
    wb = load_workbook(path)
    context_sets = build_target_context_sets()
    stats: dict[str, int] = defaultdict(int)

    for sheet_name in wb.sheetnames:
        if sheet_name == "Names":
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        idx = {header: pos + 1 for pos, header in enumerate(headers)}
        for row_num in range(2, ws.max_row + 1):
            row_values = {header: ws.cell(row=row_num, column=col).value for header, col in idx.items()}
            row_context = derive_row_context(row_values)
            if not row_context:
                continue
            for assignment in SPLITS:
                if row_context not in context_sets[assignment.label]:
                    continue
                for header, col in idx.items():
                    cell = ws.cell(row=row_num, column=col)
                    new_value, changed = replace_cell_value(cell.value, assignment.old_id, assignment.new_id)
                    if changed:
                        cell.value = new_value
                        stats[f"{sheet_name}_cell_updates"] += 1

    # Update / add Names rows for split IDs and recalculate counts from patched ElectionResults.
    ws_names = wb["Names"]
    name_headers = [cell.value for cell in ws_names[1]]
    name_idx = {header: pos + 1 for pos, header in enumerate(name_headers)}

    existing_name_rows: dict[tuple[str, int], int] = {}
    for row_num in range(2, ws_names.max_row + 1):
        full_name = ws_names.cell(row=row_num, column=name_idx["Full Name usually known by"]).value
        person_id = ws_names.cell(row=row_num, column=name_idx["PersonID"]).value
        if full_name and person_id not in (None, ""):
            existing_name_rows[(str(full_name), int(person_id))] = row_num

    # Duplicate names rows for new split IDs if needed.
    for assignment in SPLITS:
        old_key = (assignment.name, assignment.old_id)
        new_key = (assignment.name, assignment.new_id)
        if new_key in existing_name_rows:
            continue
        old_row_num = existing_name_rows.get(old_key)
        if not old_row_num:
            raise ValueError(f"Missing Names row for {assignment.name} / {assignment.old_id}")
        row_values = [ws_names.cell(row=old_row_num, column=col).value for col in range(1, ws_names.max_column + 1)]
        row_values[name_idx["PersonID"] - 1] = assignment.new_id
        ws_names.append(row_values)
        existing_name_rows[new_key] = ws_names.max_row
        stats["Names_rows_added"] += 1

    # Recalculate counts for affected IDs based on patched ElectionResults.
    ws_results = wb["ElectionResults"]
    result_headers = [cell.value for cell in ws_results[1]]
    result_idx = {header: pos + 1 for pos, header in enumerate(result_headers)}
    counts: dict[int, dict[str, int]] = defaultdict(lambda: {"devolved": 0, "westminster": 0, "european": 0, "total": 0})
    for row_num in range(2, ws_results.max_row + 1):
        result_type = str(ws_results.cell(row=row_num, column=result_idx["ResultType"]).value or "")
        if not (
            result_type == "Candidate"
            or result_type.startswith("ListCandidate")
            or result_type.startswith("RegionalListCandidate")
        ):
            continue
        person_id = ws_results.cell(row=row_num, column=result_idx["PersonID"]).value
        if person_id in (None, ""):
            continue
        person_id = int(person_id)
        body = str(ws_results.cell(row=row_num, column=result_idx["ElectedBody"]).value or "")
        if "House of Commons" in body:
            counts[person_id]["westminster"] += 1
        elif "European" in body:
            counts[person_id]["european"] += 1
        else:
            counts[person_id]["devolved"] += 1
        counts[person_id]["total"] += 1

    affected_ids = {assignment.old_id for assignment in SPLITS} | {assignment.new_id for assignment in SPLITS}
    for assignment in SPLITS:
        for pid in (assignment.old_id, assignment.new_id):
            row_num = existing_name_rows[(assignment.name, pid)]
            ws_names.cell(row=row_num, column=name_idx["Times stood for devolved bodies"]).value = counts[pid]["devolved"]
            ws_names.cell(row=row_num, column=name_idx["Times stood for Westminster"]).value = counts[pid]["westminster"]
            ws_names.cell(row=row_num, column=name_idx["Times stood for European Parliament"]).value = counts[pid]["european"]
            ws_names.cell(row=row_num, column=name_idx["Total times stood"]).value = counts[pid]["total"]
            stats["Names_rows_recounted"] += 1

    temp_path = path.with_suffix(".tmp.xlsx")
    wb.save(temp_path)
    load_workbook(temp_path, read_only=True).close()
    temp_path.replace(path)
    return dict(stats)


def replace_json_values(obj: Any, old_id: int, new_id: int) -> tuple[Any, int]:
    changes = 0
    if isinstance(obj, dict):
        out = {}
        for key, value in obj.items():
            new_value, delta = replace_json_values(value, old_id, new_id)
            out[key] = new_value
            changes += delta
        return out, changes
    if isinstance(obj, list):
        out = []
        for item in obj:
            new_item, delta = replace_json_values(item, old_id, new_id)
            out.append(new_item)
            changes += delta
        return out, changes
    new_value, changed = replace_cell_value(obj, old_id, new_id)
    return new_value, int(changed)


def patch_website_json(backup_root: Path) -> dict[str, int]:
    stats: dict[str, int] = defaultdict(int)
    for assignment in SPLITS:
        for rel_path in assignment.website_json_files:
            path = Path(rel_path)
            backup_file(path, backup_root / "website-json")
            payload = json.loads(path.read_text(encoding="utf-8"))
            updated, changes = replace_json_values(payload, assignment.old_id, assignment.new_id)
            path.write_text(json.dumps(updated, ensure_ascii=False), encoding="utf-8")
            stats[str(path)] += changes
    return dict(stats)


def remap_scalar(value: Any, mapping: dict[int, int]):
    if value in (None, ""):
        return value, False
    try:
        current = int(value)
    except (TypeError, ValueError):
        return value, False
    new_value = mapping.get(current)
    if new_value is None or new_value == current:
        return value, False
    return new_value, True


def remap_id_list(value: Any, mapping: dict[int, int]):
    if value in (None, ""):
        return value, False
    parts = [part.strip() for part in str(value).split(",")]
    changed = False
    remapped = []
    for part in parts:
        if not part:
            remapped.append(part)
            continue
        try:
            current = int(part)
        except ValueError:
            remapped.append(part)
            continue
        new_value = mapping.get(current, current)
        if new_value != current:
            changed = True
        remapped.append(str(new_value))
    return ", ".join(remapped), changed


def patch_local_workbook(path: Path, backup_root: Path) -> dict[str, int]:
    backup_file(path, backup_root)
    wb = load_workbook(path)
    ws_results = wb["ElectionResults"]
    result_headers = [cell.value for cell in ws_results[1]]
    result_idx = {header: pos + 1 for pos, header in enumerate(result_headers)}

    current_ids_by_name: dict[str, set[int]] = defaultdict(set)
    for row_num in range(2, ws_results.max_row + 1):
        if ws_results.cell(row=row_num, column=result_idx["ResultType"]).value != "Candidate":
            continue
        name = ws_results.cell(row=row_num, column=result_idx["Name usually known by"]).value
        pid = ws_results.cell(row=row_num, column=result_idx["PersonID"]).value
        if name in LOCAL_NAME_TO_ID and pid not in (None, ""):
            current_ids_by_name[str(name)].add(int(pid))

    id_mapping: dict[int, int] = {}
    for name, target_id in LOCAL_NAME_TO_ID.items():
        ids = current_ids_by_name.get(name, set())
        if not ids:
            continue
        if len(ids) > 1:
            raise ValueError(f"Local workbook has multiple IDs for {name}: {sorted(ids)}")
        current_id = next(iter(ids))
        id_mapping[current_id] = target_id

    stats: dict[str, int] = defaultdict(int)
    result_transfer_subject_headers = [
        header for header in result_headers if isinstance(header, str) and header.startswith("TransferSubject")
    ]
    for row_num in range(2, ws_results.max_row + 1):
        person_cell = ws_results.cell(row=row_num, column=result_idx["PersonID"])
        new_value, changed = remap_scalar(person_cell.value, id_mapping)
        if changed:
            person_cell.value = new_value
            stats["ElectionResults_PersonID_updates"] += 1
        for header in result_transfer_subject_headers:
            cell = ws_results.cell(row=row_num, column=result_idx[header])
            new_value, changed = remap_scalar(cell.value, id_mapping)
            if changed:
                cell.value = new_value
                stats["ElectionResults_TransferSubject_updates"] += 1

    ws_transfers = wb["Transfers"]
    transfer_headers = [cell.value for cell in ws_transfers[1]]
    transfer_idx = {header: pos + 1 for pos, header in enumerate(transfer_headers)}
    for row_num in range(2, ws_transfers.max_row + 1):
        for header in ("PersonID", "TransferSubject", "SourcePersonID"):
            cell = ws_transfers.cell(row=row_num, column=transfer_idx[header])
            new_value, changed = remap_scalar(cell.value, id_mapping)
            if changed:
                cell.value = new_value
                stats[f"Transfers_{header}_updates"] += 1
        remaining = ws_transfers.cell(row=row_num, column=transfer_idx["RemainingCandidateIDsDesc"])
        new_value, changed = remap_id_list(remaining.value, id_mapping)
        if changed:
            remaining.value = new_value
            stats["Transfers_RemainingCandidateIDsDesc_updates"] += 1

    temp_path = path.with_suffix(".tmp.xlsx")
    wb.save(temp_path)
    load_workbook(temp_path, read_only=True).close()
    temp_path.replace(path)
    stats["local_name_mapping_count"] = len(id_mapping)
    return dict(stats)


def main() -> None:
    timestamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    backup_root = Path("backups") / f"requested-identity-fixes-{timestamp}"
    stats = {
        "full_workbook": patch_full_workbook(FULL_WORKBOOK, backup_root / "full-workbook"),
        "website_json": patch_website_json(backup_root),
        "local_workbook": patch_local_workbook(LOCAL_WORKBOOK, backup_root / "local-workbook"),
    }
    print(json.dumps(stats, indent=2))


if __name__ == "__main__":
    main()
