#!/usr/bin/env python
"""Extend the pre-1922 Westminster JSON corpus with the 3 NI sections from
parlconst (101 Antrim+Belfast(W), 102 Down+Armagh+Belfast(E), 103 Fermanagh+
Tyrone+Londonderry).

Same 9 GEs as the ROI build (1885, 1886, 1892, 1895, 1900, 1906, 1910 J,
1910 D, 1918) emit per-constituency JSON files into
election-viewer-package/data/elections/house-of-commons-of-the-united-kingdom/<date>/
and merge into _pre1970_index.json. Also append rows to the unified CSV
data/external/parlconst/pre1922_westminster_results.csv.
"""
import csv, json, re, sys, io, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict

REPO = Path(__file__).resolve().parent.parent
EVP = REPO / "election-viewer-package"
ELECTIONS_DIR = EVP / "data" / "elections" / "house-of-commons-of-the-united-kingdom"
INDEX = ELECTIONS_DIR / "_pre1970_index.json"
RESULTS_CSV = REPO / "data" / "external" / "parlconst" / "pre1922_westminster_results.csv"

GE_DATES = {
    '1885':     '1885-11-24',
    '1886':     '1886-07-01',
    '1892':     '1892-07-04',
    '1895':     '1895-07-13',
    '1900':     '1900-09-26',
    '1906':     '1906-01-12',
    '1910 (J)': '1910-01-15',
    '1910 (D)': '1910-12-03',
    '1918':     '1918-12-14',
}

PARTY_DEFS = {
    'N':    {'name': 'Irish Parliamentary Party', 'colour': '#0E7C42'},
    'SF':   {'name': 'Sinn Féin',                 'colour': '#326760'},
    'PN':   {'name': 'Parnellite Nationalist',    'colour': '#1D9656'},
    'IndN': {'name': 'Independent Nationalist',   'colour': '#5DBC8E'},
    'U':    {'name': 'Irish Unionist Alliance',   'colour': '#1F4E8C'},
    'IndU': {'name': 'Independent Unionist',      'colour': '#3F7CC0'},
    'LU':   {'name': 'Liberal Unionist',          'colour': '#E0A60C'},
    'C':    {'name': 'Conservative',              'colour': '#0087DC'},
    'L':    {'name': 'Liberal',                   'colour': '#FAA61A'},
    'Lab':  {'name': 'Labour',                    'colour': '#E4003B'},
    'La U': {'name': 'Labour Unionist',           'colour': '#7A2A82'},  # historical NI variant
    'Oth':  {'name': 'Other',                     'colour': '#888888'},
}

NI_SECTIONS = ['NI_101', 'NI_102', 'NI_103']


def slugify(name):
    name = unicodedata.normalize('NFKD', name).encode('ascii','ignore').decode('ascii')
    s = re.sub(r"[^a-zA-Z0-9]+", '-', name).strip('-').lower()
    return re.sub(r'-+', '-', s) or 'unknown'


def normalize_code(c):
    if c is None: return None
    c = str(c).strip()
    if not c: return None
    return c


def collect_results():
    """Walk NI_101/102/103 ERT sheets. Skip blank-name rows. Stop when
    encountering the next-era header (e.g. '1950 - 1983')."""
    out = []
    for section in NI_SECTIONS:
        xlsx = next(Path('_tmp_parlconst/files').glob(f'{section}_other_dl_*.xlsx'))
        wb = load_workbook(xlsx, data_only=True)
        ws = wb['ERT']
        rows = list(ws.iter_rows(values_only=True))
        header = rows[1]
        year_cols = {}
        for i, cell in enumerate(header):
            if cell is None: continue
            s = str(cell).strip()
            if s in GE_DATES:
                year_cols[s] = i
        consec_blank = 0
        for r in rows[2:]:
            constit = r[1]
            if constit is None:
                consec_blank += 1
                if consec_blank > 4: break  # gap → end of section
                continue
            consec_blank = 0
            constit = str(constit).strip()
            if not constit: continue
            if constit.startswith('Total'): continue
            # Detect next-era header (e.g. '1950 - 1983') as termination
            if re.match(r'\d{4}\s*-\s*\d{4}$', constit): break
            for year, col in year_cols.items():
                code = normalize_code(r[col])
                if code:
                    out.append((GE_DATES[year], constit, code))
    return out


def make_constituency_json(constit, code):
    party = PARTY_DEFS.get(code, {'name': code, 'colour': '#888888'})
    return {
        "Constituency": {
            "countInfo": {
                "Constituency_Name": constit,
                "Constituency_Number": "",
                "Number_Of_Seats": "1",
                "Spoiled": "",
                "Total_Electorate": "",
                "Total_Poll": "",
                "Valid_Poll": "",
            },
            "countGroup": [{
                "Candidate_First_Pref_Votes": "Unknown",
                "Candidate_Id": "",
                "Constituency_Number": "",
                "Count_Number": "1",
                "Firstname": "",
                "Occurred_On_Count": "",
                "Party_Colour": party['colour'],
                "Party_Name": party['name'],
                "Status": "Elected",
                "Surname": "",
                "Total_Votes": "",
                "Transfers": "0.00",
                "candidateName": "(candidate name not in source)",
                "id": 0,
            }],
        }
    }


def main():
    print("Collecting NI ERT results ...")
    results = collect_results()
    print(f"  {len(results)} (date, constituency, code) entries across {len(set(r[1] for r in results))} NI constituencies")
    from collections import Counter
    code_counts = Counter(r[2] for r in results)
    print(f"  party codes: {dict(code_counts.most_common())}")

    # Append to unified CSV
    existing_rows = []
    if RESULTS_CSV.exists():
        with RESULTS_CSV.open('r', encoding='utf-8', newline='') as f:
            existing_rows = list(csv.reader(f))
    existing_keys = {tuple(r[:3]) for r in existing_rows[1:]}
    new_csv_rows = []
    for date, c, code in results:
        if (date, c, code) in existing_keys: continue
        party = PARTY_DEFS.get(code, {'name': code})
        new_csv_rows.append([date, c, code, party['name']])
    print(f"  appending {len(new_csv_rows)} rows to CSV")
    with RESULTS_CSV.open('a', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        for row in new_csv_rows: w.writerow(row)

    # Group by date
    by_date = defaultdict(dict)
    for date, constit, code in results:
        by_date[date][constit] = code

    # Emit per-constituency JSON files (don't overwrite if existing has more data)
    file_count = 0
    pre1970_index_entries = []
    for date in sorted(by_date.keys()):
        d_dir = ELECTIONS_DIR / date
        d_dir.mkdir(parents=True, exist_ok=True)
        constits_for_date = sorted(by_date[date].keys())
        for constit in constits_for_date:
            code = by_date[date][constit]
            slug = slugify(constit)
            fpath = d_dir / f'{slug}.json'
            if fpath.exists():
                # Existing file (probably from ROI builder OR Wiki backfill) — skip overwrite
                continue
            obj = make_constituency_json(constit, code)
            fpath.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + '\n', encoding='utf-8')
            file_count += 1
        pre1970_index_entries.append({'date': date, 'constituencies': constits_for_date})
    print(f"\nwrote {file_count} new JSON files")

    # Update _pre1970_index.json — merge constituency lists for existing dates
    if INDEX.exists():
        existing = json.loads(INDEX.read_text(encoding='utf-8'))
    else:
        existing = []
    by_existing_date = {e.get('date'): e for e in existing if isinstance(e, dict)}
    for entry in pre1970_index_entries:
        date = entry['date']
        if date in by_existing_date:
            cur = set(by_existing_date[date].get('constituencies') or [])
            cur.update(entry['constituencies'])
            by_existing_date[date]['constituencies'] = sorted(cur)
        else:
            existing.append(entry)
            by_existing_date[date] = entry
    existing.sort(key=lambda e: e.get('date',''))
    INDEX.write_text(json.dumps(existing, indent=1, ensure_ascii=False) + '\n', encoding='utf-8')
    print(f"  updated {INDEX} ({len(existing)} entries)")


if __name__ == "__main__":
    main()
