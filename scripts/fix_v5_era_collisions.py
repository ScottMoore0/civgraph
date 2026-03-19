#!/usr/bin/env python
"""Fix era collisions in v5: split PIDs where pre-1970 and post-1970
candidates sharing a name were incorrectly merged.

Strategy: For any PID with appearances spanning >40 years, check if there's
a gap of >25 years between consecutive appearances.  If so, split at the gap.
This handles the common case of Stormont-era politicians (1920s-60s) being
merged with modern politicians (1970s+) who happen to share a name.

Exceptions: known long careers (verified same person) are preserved.
"""

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl


def norm(name):
    if not name: return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


# Known genuinely long careers (verified same person — do NOT split)
KNOWN_LONG_CAREERS = {
    "99683",  # David Bleakley — NI Labour politician 1949-1998
    "60268",  # Erskine Holmes — NI Labour/NILRC 1965-2016
    "83992",  # Eamonn McCann — NI Labour/PBP 1969-2019
    "96686",  # William Beattie — DUP/Protestant Unionist 1929-2001 (born 1930s — could be father+son, but common enough to keep)
}


def main():
    input_path = Path("Full election tables - comprehensive - personid-v5.xlsx")
    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb["ElectionResults"]

    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))

    pid_col = col["PersonID"]
    rt_col = col["ResultType"]
    name_col = col["Name usually known by"]
    date_col = col["Date"]
    party_col = col["Party Name"]

    # Build per-PID date/row index
    pid_rows: dict[str, list[tuple[int, int]]] = defaultdict(list)  # pid -> [(year, row_idx)]
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col])
        date = str(row[date_col] or "")
        if date[:4].isdigit():
            pid_rows[pid].append((int(date[:4]), i))

    # Find PIDs to split
    max_pid = max(
        (int(str(row[pid_col])) for row in rows if row[pid_col] is not None and str(row[pid_col]).isdigit()),
        default=0
    )
    next_id = max_pid + 1

    splits = 0
    split_log = []

    for pid, year_indices in pid_rows.items():
        if pid in KNOWN_LONG_CAREERS:
            continue

        years_sorted = sorted(year_indices, key=lambda x: x[0])
        if not years_sorted:
            continue

        span = years_sorted[-1][0] - years_sorted[0][0]
        if span <= 40:
            continue

        # Find the largest gap
        max_gap = 0
        max_gap_pos = -1
        for k in range(1, len(years_sorted)):
            gap = years_sorted[k][0] - years_sorted[k - 1][0]
            if gap > max_gap:
                max_gap = gap
                max_gap_pos = k

        if max_gap < 25:
            continue  # No large enough gap — might genuinely be the same person

        # Split: everything before the gap stays, everything after gets new PID
        before = years_sorted[:max_gap_pos]
        after = years_sorted[max_gap_pos:]

        new_pid = str(next_id)
        next_id += 1

        for _, idx in after:
            rows[idx][pid_col] = new_pid

        splits += 1
        before_name = rows[before[0][1]][name_col] if before else "?"
        after_name = rows[after[0][1]][name_col] if after else "?"
        before_party = rows[before[0][1]][party_col] if before else "?"
        after_party = rows[after[0][1]][party_col] if after else "?"
        split_log.append({
            "pid": pid,
            "before": f"{before_name} ({before_party}, {before[0][0]}-{before[-1][0]})",
            "after": f"{after_name} ({after_party}, {after[0][0]}-{after[-1][0]})",
            "gap": max_gap,
            "new_pid": new_pid,
        })

    print(f"\nSplit {splits} era collisions")
    for s in split_log:
        print(f"  PID {s['pid']}: {s['before']} | gap={s['gap']}yr | {s['after']} -> PID {s['new_pid']}")

    # Update Transfers
    print("\nUpdating Transfers...")
    name_to_pid: dict[str, str] = {}
    for row in rows:
        if row[rt_col] != "Candidate": continue
        name = row[name_col]
        pid = str(row[pid_col] or "")
        if name and pid:
            n = norm(name)
            if n: name_to_pid[n] = pid

    ws_t = wb["Transfers"]
    t_h = [cell.value for cell in ws_t[1]]
    t_pid_idx = t_h.index("PersonID")
    t_name_idx = t_h.index("Name")
    t_upd = 0
    for t_row in ws_t.iter_rows(min_row=2):
        nv = t_row[t_name_idx].value
        if nv:
            n = norm(nv)
            if n in name_to_pid:
                new = name_to_pid[n]
                if str(t_row[t_pid_idx].value or "") != new:
                    t_row[t_pid_idx].value = new
                    t_upd += 1
    print(f"  Updated {t_upd} Transfers rows")

    # Write
    print(f"\nSaving {input_path}...")
    for i, rd in enumerate(rows):
        for j, val in enumerate(rd):
            ws.cell(row=i + 2, column=j + 1, value=val)
    wb.save(input_path)

    # Final stats
    pid_names = defaultdict(set)
    pid_bodies = defaultdict(set)
    pid_dates_final = defaultdict(set)
    total = 0
    for row in rows:
        if row[rt_col] != "Candidate": continue
        total += 1
        pid = str(row[pid_col])
        pid_names[pid].add(row[name_col] or "")
        pid_bodies[pid].add(row[col["ElectedBody"]] or "")
        pid_dates_final[pid].add(str(row[date_col] or ""))

    cross = sum(1 for bs in pid_bodies.values() if len(bs) > 1)
    multi = sum(1 for ns in pid_names.values() if len(ns) > 1)
    long_spans = sum(1 for pid in pid_dates_final
                     if max((int(d[:4]) for d in pid_dates_final[pid] if d[:4].isdigit()), default=0) -
                        min((int(d[:4]) for d in pid_dates_final[pid] if d[:4].isdigit()), default=0) > 40)

    print(f"\n{'='*60}")
    print(f"  V5 POST-FIX STATE")
    print(f"{'='*60}")
    print(f"  Candidate rows: {total}")
    print(f"  Unique PersonIDs: {len(pid_names)}")
    print(f"  Cross-body people: {cross}")
    print(f"  PIDs with name variants: {multi}")
    print(f"  Remaining >40yr spans: {long_spans}")

    body_counts = defaultdict(int)
    for row in rows:
        if row[rt_col] != "Candidate": continue
        body_counts[row[col["ElectedBody"]]] += 1
    for body in sorted(body_counts):
        bp = len({p for p, bs in pid_bodies.items() if body in bs})
        print(f"    {body}: {body_counts[body]} candidacies, {bp} people")

    Path("personid_v5_erasplit_log.json").write_text(
        json.dumps(split_log, indent=2, default=str, ensure_ascii=False), encoding="utf-8")


if __name__ == "__main__":
    main()
