#!/usr/bin/env python
"""Fix remaining PersonID issues in the v3 workbook.

1. Clean trailing dash artefacts from names (e.g. "John Carson –" → "John Carson")
2. Split genuine collisions where different people share a PID
3. Add missing first-name variant mappings so legitimate variants aren't flagged

Identified collisions (different people, same PID):
- Neil Kelly (Alliance, Antrim) ≠ Niall Kelly (SDLP, Balmoral) — same year, different everything
- Joseph McBride (DUP, Armagh) ≠ Joseph McBride (SDLP, Magherafelt) — same name, different person
- Paddy McGrath (SDLP, Fermanagh) ≠ Paddy McGrath (Labour, Mid Ulster) — different person
- Sean Montgomery (SF, Castlereagh) ≠ John Montgomery (Alliance/SDLP) — different person
- James Murray (Liberal, EU) ≠ Seamus Murray (WP, Newry) — different person
- John Lynch (SF/Ind, Armagh 1959-64) ≠ Seán Lynch (SF, Fermanagh 2005+) — 46yr gap, different
- Sean McCloskey (Ind, Limavady 1981) ≠ John McCloskey (SF, E Londonderry 1996) — different
- Sean Quinn (Liberal, Down 1960s) ≠ Sean Quinn (SDLP, Down 1977+) — likely different
- Liam Kennedy (Ind, Belfast 1997+) ≠ William Kennedy (CLP/UUP, Belfast 1945-69) — different
- Patrick Doherty (SDLP/Alliance, various) + Paddy Doherty (Nat, Omagh) — at least 2 people
- Thomas Black (DUP, Armagh) ≠ Thomas Black (Labour, Belfast) — different
- Albert Johnston (DUP, Castlereagh 1985) ≠ Bert Johnston (UUP→DUP, Fermanagh) — different
- Liam Norris (Alliance, Belfast 2019) ≠ William Norris (DUP, Limavady 1977-85) — different
- Ciaran Mulholland (Labour, Lagan Valley 1996) ≠ Kieran Mulholland (SF, Glens 2011+) — different
- Catherine Kelly (SF, Omagh 2014+) ≠ Kathleen Kelly (SF, E Londonderry 1996) — possibly different
- Paddy Agnew (NI Labour, S Armagh 1945) ≠ Patrick Agnew (Unity/IIP, Armagh 1973-81) — likely different
- Liam Gallagher (WP, Derry 1973-77) ≠ William Gallagher (H-Block, Strabane 1981) — different
- Thomas Caldwell (Alliance/UUP, Larne 1989-93) ≠ Tom Caldwell (Ind, Belfast 1969-70) — different
- J.J. Magee (WP, various 1977-93) ≠ JJ Magee (SF, Belfast 2011+) — different
- Sean Laverty (SF, S Antrim 1983) ≠ John Laverty (DUP, Castlereagh 2019+) — different
- Bobby Irvine (Ind, Holywood 1997) ≠ Robert Irvine (DUP/UUP, Enniskillen 1997-2001) — different
- John Davey (SF, E Londonderry 1983-87) ≠ Sean Davey (SF, Ballymena 2011) — different
- Sean O'Hare (WP, Belfast 1973-77) ≠ Sean O'Hare (SDLP, Crotlieve 2011) + John O'Hare (SDLP, 2019) — multiple people
- Liam Logan (not present) vs William Logan (UUP, Ballymoney) — William Logan is one person, no split needed
- Drew Thompson (Andrew Thompson UUP Antrim) ≠ Andrew Thompson (Cons, Newtownards + UIV, Belfast) — possible 2-3 people
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl


def normalize_name(name: str) -> str:
    if not name:
        return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


# Collisions to split: (pid, split_key_function)
# Each entry: the PID, and a function that returns a group key for each row
# Rows with different keys get different PIDs

SPLITS = [
    # PID 21661: Neil Kelly (Alliance) ≠ Niall Kelly (SDLP)
    (21661, lambda r: "niall" if "Niall" in (r["name"] or "") else "neil"),
    # PID 24174: Joseph McBride (DUP, Armagh) ≠ Joseph McBride (SDLP, Magherafelt)
    (24174, lambda r: "dup" if r["party"] == "DUP" else "sdlp"),
    # PID 2799: John Montgomery (Alliance→SDLP) ≠ Sean Montgomery (SF)
    (2799, lambda r: "sean" if "Sean" in (r["name"] or "") else "john"),
    # PID 59264: James Murray (Liberal) ≠ Seamus Murray (WP)
    (59264, lambda r: "seamus" if "Seamus" in (r["name"] or "") else "james"),
    # PID 73893: John Lynch (1959-64) ≠ Seán Lynch (2005+)
    (73893, lambda r: "sean" if "Seán" in (r["name"] or "") or "Sean" in (r["name"] or "") else "john"),
    # PID 67448: Sean McCloskey (IIP) ≠ John McCloskey (SF)
    (67448, lambda r: "john" if "John" in (r["name"] or "") else "sean"),
    # PID 60498: John Quinn (Liberal 1966-70) ≠ Sean Quinn (SDLP 1977+)
    (60498, lambda r: "sean" if "Sean" in (r["name"] or "") else "john"),
    # PID 51502: Liam Kennedy (Ind 1997+) ≠ William Kennedy (CLP/UUP 1945-69)
    (51502, lambda r: "liam" if "Liam" in (r["name"] or "") else "william"),
    # PID 26409: Multiple Patrick/Paddy Dohertys — split by party+area
    (26409, lambda r: "nat_omagh" if r["party"] == "Nationalist Party" else
                      "nilab" if r["party"] == "NI Labour" else
                      "alliance" if r["party"] == "Alliance" else "sdlp"),
    # PID 52501: Thomas Black (DUP, Armagh) ≠ Thomas Black (Labour, Belfast)
    (52501, lambda r: "labour" if "Labour" in (r["party"] or "") else "dup"),
    # PID 42028: Albert Johnston (DUP, Castlereagh) ≠ Bert Johnston (UUP→DUP, Fermanagh)
    (42028, lambda r: "albert" if "Albert" in (r["name"] or "") or "Castlereagh" in (r["const"] or "") else "bert"),
    # PID 42569: Liam Norris (Alliance 2019) ≠ William Norris (DUP, Limavady 1977-85)
    (42569, lambda r: "liam" if "Liam" in (r["name"] or "") else "william"),
    # PID 79116: Ciaran Mulholland (Labour 1996) ≠ Kieran Mulholland (SF, Glens 2011+)
    (79116, lambda r: "ciaran" if "Ciaran" in (r["name"] or "") else "kieran"),
    # PID 36316: Catherine Kelly (SF, Omagh 2014+) ≠ Kathleen Kelly (SF, E Londonderry 1996)
    (36316, lambda r: "kathleen" if "Kathleen" in (r["name"] or "") else "catherine"),
    # PID 84707: Paddy Agnew (Labour 1945) ≠ Patrick Agnew (Unity/IIP 1973-81)
    (84707, lambda r: "paddy45" if "1945" in (r["date"] or "") else "patrick73"),
    # PID 88879: Liam Gallagher (WP, Derry) ≠ William Gallagher (H-Block, Strabane)
    (88879, lambda r: "william" if "William" in (r["name"] or "") else "liam"),
    # PID 89719: Tom Caldwell (Ind, Belfast 1969-70) ≠ Thomas Caldwell (Alliance/UUP, Larne 1989+)
    (89719, lambda r: "tom69" if r["date"] and r["date"][:4] < "1980" else "thomas89"),
    # PID 91406: J.J. Magee (WP, 1977-93) ≠ JJ Magee (SF, 2011+)
    (91406, lambda r: "wp" if "Workers" in (r["party"] or "") else "sf"),
    # PID 96041: Sean Laverty (SF, 1983) ≠ John Laverty (DUP, 2019+)
    (96041, lambda r: "john" if "John" in (r["name"] or "") else "sean"),
    # PID 11796: Bobby Irvine (Ind, Holywood) ≠ Robert Irvine (DUP/UUP, Enniskillen)
    (11796, lambda r: "bobby_holywood" if "Holywood" in (r["const"] or "") or "Bobby" in (r["name"] or "") else "robert_enniskillen"),
    # PID 32087: John Davey (SF, E Londonderry 1983-87) ≠ Sean Davey (SF, Ballymena 2011)
    (32087, lambda r: "sean" if "Sean" in (r["name"] or "") else "john"),
    # PID 33467: Multiple O'Hares across decades
    (33467, lambda r: "wp_70s" if "Workers" in (r["party"] or "") and r["date"] and r["date"][:4] < "1980" else
                      "sdlp_crot" if "Crotlieve" in (r["const"] or "") else
                      "sdlp_lagan" if "Lagan" in (r["const"] or "") else "other"),
    # PID 701: Paddy McGrath (SDLP, Fermanagh) ≠ Paddy McGrath (Labour, Mid Ulster)
    (701, lambda r: "labour" if "Labour" in (r["party"] or "") else "sdlp"),
    # PID 74365: John McKee (UPNI, Down 1977) ≠ Sean McKee (SDLP, Antrim 1997+)
    (74365, lambda r: "john" if "John" in (r["name"] or "") else "sean"),
    # PID 39588: Liam Johnston ≠ William Johnston (multiple Williams too)
    (39588, lambda r: "liam" if "Liam" in (r["name"] or "") else "william"),
    # PID 59944: Seamus Lynch (WP) ≠ James Lynch (H-Block)
    (59944, lambda r: "james_hblock" if "H-Block" in (r["party"] or "") or "Anti" in (r["party"] or "") else "seamus"),
    # PID 74806: Multiple Patrick Kellys
    (74806, lambda r: "pearse" if "Pearse" in (r["name"] or "") or "Democratic" in (r["party"] or "") else
                      "paddy23" if r["date"] and r["date"][:4] >= "2020" else "patrick"),
]


def main() -> None:
    input_path = Path("Full election tables - comprehensive - personid-v3.xlsx")
    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb["ElectionResults"]

    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))

    pid_col = col["PersonID"]
    rt_col = col["ResultType"]
    name_col = col["Name usually known by"]
    party_col = col["Party Name"]
    date_col = col["Date"]
    const_col = col["Constituency"]

    # Step 1: Clean trailing dash artefacts from names
    print("\nStep 1: Cleaning name artefacts...")
    cleaned = 0
    for i, row in enumerate(rows):
        name = row[name_col]
        if name and isinstance(name, str):
            # Remove trailing " –" or " -" (Wikipedia en-dash artefact)
            new_name = re.sub(r"\s*[–—-]\s*$", "", name).strip()
            # Remove trailing special chars like ♭
            new_name = re.sub(r"\s*[♭†‡*]+\s*$", "", new_name).strip()
            if new_name != name:
                rows[i][name_col] = new_name
                # Also fix Source Name if it matches
                src_col = col.get("Source Name")
                if src_col and row[src_col] == name:
                    rows[i][src_col] = new_name
                cleaned += 1
    print(f"  Cleaned {cleaned} name artefacts")

    # Step 2: Split collisions
    print("\nStep 2: Splitting collisions...")

    # Find max PID for fresh ID assignment
    max_pid = 0
    for row in rows:
        pid = row[pid_col]
        if pid is not None:
            try:
                v = int(str(pid))
                if v > max_pid:
                    max_pid = v
            except ValueError:
                pass
    next_id = max(max_pid + 1, 900001)

    total_splits = 0
    split_log = []
    for target_pid, key_fn in SPLITS:
        target_str = str(target_pid)
        # Find all rows with this PID
        groups = defaultdict(list)
        for i, row in enumerate(rows):
            if row[rt_col] != "Candidate":
                continue
            if str(row[pid_col]) != target_str:
                continue
            r = {
                "name": row[name_col],
                "party": row[party_col],
                "date": str(row[date_col] or ""),
                "const": row[const_col],
            }
            key = key_fn(r)
            groups[key].append(i)

        if len(groups) <= 1:
            continue

        # Keep first group on original PID, assign new PIDs to others
        group_list = sorted(groups.items())
        for group_key, indices in group_list[1:]:
            new_pid = str(next_id)
            next_id += 1
            for idx in indices:
                rows[idx][pid_col] = new_pid
            total_splits += 1
            split_log.append({
                "original_pid": target_str,
                "group": group_key,
                "new_pid": new_pid,
                "rows": len(indices),
                "sample_name": rows[indices[0]][name_col],
            })

    print(f"  Split {total_splits} collision groups")
    for s in split_log:
        print(f"    PID {s['original_pid']} → {s['new_pid']}: {s['sample_name']} ({s['rows']} rows)")

    # Step 3: Update Transfers sheet
    print("\nStep 3: Updating Transfers sheet...")
    name_to_pid = {}
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        name = row[name_col]
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if name and pid:
            n = normalize_name(name)
            if n:
                name_to_pid[n] = pid

    ws_t = wb["Transfers"]
    t_headers = [cell.value for cell in ws_t[1]]
    t_pid_idx = t_headers.index("PersonID") if "PersonID" in t_headers else None
    t_name_idx = t_headers.index("Name") if "Name" in t_headers else None
    t_updates = 0
    if t_pid_idx is not None and t_name_idx is not None:
        for t_row in ws_t.iter_rows(min_row=2):
            name_val = t_row[t_name_idx].value
            if name_val:
                n = normalize_name(name_val)
                if n in name_to_pid:
                    old = str(t_row[t_pid_idx].value) if t_row[t_pid_idx].value else ""
                    new = name_to_pid[n]
                    if old != new:
                        t_row[t_pid_idx].value = new
                        t_updates += 1
    print(f"  Updated {t_updates} Transfers rows")

    # Write
    print(f"\nWriting {input_path}...")
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=i + 2, column=j + 1, value=val)
    wb.save(input_path)

    # Final verification
    final_pid_names = defaultdict(set)
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col])
        name = row[name_col] or ""
        if name:
            final_pid_names[pid].add(name)

    multi = {p: names for p, names in final_pid_names.items() if len(names) > 1}
    print(f"\n  Unique PIDs: {len(final_pid_names)}")
    print(f"  PIDs with multiple name variants: {len(multi)}")
    print(f"  Splits applied: {total_splits}")
    print(f"  Names cleaned: {cleaned}")

    # Save log
    Path("personid_v3_fixes_log.json").write_text(
        json.dumps({"splits": split_log, "names_cleaned": cleaned}, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
