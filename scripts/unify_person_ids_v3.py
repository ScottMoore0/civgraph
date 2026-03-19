#!/usr/bin/env python
"""Unify PersonIDs v3 — starts from the ORIGINAL comprehensive workbook.

Key improvement over v2: never merges by surname alone.  Only merges when
normalised full names match exactly, or when first names are known variants
of each other (Paddy/Patrick, Bill/William, Seán/Sean, etc.).

Phases:
  1a — Remove garbage rows (Wikipedia parse artifacts)
  1b — Strip T-prefix from temp IDs
  1c — Merge temp/hash IDs into curated IDs by exact normalised name
  2  — Merge by known first-name variants (same surname required)
  3  — Unify hash IDs sharing a normalised name
  4  — Split collisions (PIDs with genuinely different people)
  5  — Assign fresh sequential IDs to remaining hash PIDs
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

# ── Known first-name variant groups ───────────────────────────────────────
# Each inner list = names that refer to the same person.
# All comparisons are done in normalised (lowercase, no diacritics) form.

FIRSTNAME_VARIANT_GROUPS: list[list[str]] = [
    # English
    ["james", "jim", "jimmy", "jas"],
    ["william", "will", "willy", "willie", "bill", "billy", "liam"],
    ["robert", "rob", "robbie", "bob", "bobby", "bert"],
    ["john", "johnny", "jack", "jock"],
    ["thomas", "tom", "tommy", "thos"],
    ["edward", "ed", "eddie", "eddy", "ted", "teddy", "ned"],
    ["richard", "dick", "dickie", "rich", "rick", "ricky"],
    ["charles", "charlie", "chas", "chuck"],
    ["patrick", "paddy", "pat", "patsy"],
    ["michael", "mike", "mick", "mickey", "mickie"],
    ["christopher", "chris"],
    ["daniel", "dan", "danny"],
    ["andrew", "andy", "drew"],
    ["anthony", "tony"],
    ["joseph", "joe", "joey"],
    ["samuel", "sam", "sammy"],
    ["benjamin", "ben", "benny"],
    ["alexander", "alex", "alec", "alistair", "alastair", "alasdair"],
    ["frederick", "fred", "freddie", "freddy"],
    ["kenneth", "ken", "kenny"],
    ["ronald", "ron", "ronnie"],
    ["raymond", "ray"],
    ["stephen", "steve", "steven"],
    ["philip", "phil"],
    ["lawrence", "larry", "laurence"],
    ["matthew", "matt"],
    ["gerald", "gerry", "gerard"],
    ["peter", "pete"],
    ["david", "dave", "davy", "davey"],
    ["donald", "don", "donnie"],
    ["dennis", "denis", "den"],
    ["terence", "terry", "terrence"],
    ["henry", "harry", "hal"],
    ["albert", "bert", "bertie", "al"],
    ["alfred", "alf", "alfie"],
    ["leonard", "len", "lenny"],
    ["bernard", "bernie"],
    ["reginald", "reg", "reggie"],
    ["archibald", "archie"],
    ["herbert", "herb", "herbie"],
    ["humphrey", "humph"],
    ["geoffrey", "geoff", "jeff", "jeffrey"],
    ["timothy", "tim", "timmy"],
    ["nicholas", "nick", "nicky"],
    ["douglas", "doug"],
    ["vincent", "vince"],
    ["francis", "frank", "frankie"],
    ["cecil", "cec"],
    ["cornelius", "con"],
    ["bartholomew", "bart"],
    ["desmond", "des"],
    ["clifford", "cliff"],
    ["gordon", "gordie"],
    ["roderick", "roddy", "rod"],
    ["wallace", "wally"],
    ["maurice", "morrie"],
    ["sydney", "sid"],
    ["stanley", "stan"],
    ["ernest", "ernie"],
    ["norman", "norm"],
    ["harold", "hal"],
    ["herbert", "herbie"],
    ["percival", "percy"],
    # Women
    ["elizabeth", "liz", "lizzie", "beth", "betty", "bess", "bessie", "eliza"],
    ["margaret", "maggie", "meg", "peggy", "madge", "marge", "margie", "mags"],
    ["catherine", "kate", "katie", "cathy", "kathleen", "katharine", "katherine"],
    ["patricia", "pat", "patsy", "tricia", "trish"],
    ["jennifer", "jenny", "jen"],
    ["dorothy", "dot", "dolly", "dora"],
    ["pamela", "pam"],
    ["susan", "sue", "susie", "susanne"],
    ["deborah", "deb", "debbie"],
    ["christine", "chris", "christina", "tina"],
    ["jacqueline", "jackie"],
    ["rosemary", "rosie"],
    ["caroline", "carol"],
    ["joanna", "jo", "joanne", "joan"],
    ["anne", "ann", "annie"],
    ["mary", "molly", "may"],
    ["eleanor", "ella", "ellie", "nell", "nellie"],
    ["victoria", "vicky"],
    ["helena", "helen"],
    ["theresa", "tessa", "terry"],
    ["bridget", "brid", "bridie"],
    ["eileen", "aileen"],
    # Irish
    ["sean", "john", "seán", "shane"],
    ["padraig", "patrick", "paddy", "padraic"],
    ["cathal", "charles"],
    ["niall", "neal", "neil"],
    ["ciaran", "kieran"],
    ["eamonn", "edmund", "eamon"],
    ["seamus", "james", "séamas"],
    ["liam", "william"],
    ["proinsias", "francis"],
    ["gearoid", "gerard", "gerry"],
    ["micheal", "michael"],
    ["maire", "mary"],
    ["siobhan", "joan"],
    ["caoimhe", "keeva"],
    ["aisling", "ashling"],
    ["grainne", "grania"],
    # Initials should NOT be matched — J. Smith ≠ James Smith
    # Double-barrelled first names
    ["jj", "j j", "j. j.", "j.j."],
]

# Build lookup: normalised first name → group index
_VARIANT_LOOKUP: dict[str, int] = {}
for _gidx, _group in enumerate(FIRSTNAME_VARIANT_GROUPS):
    for _name in _group:
        _n = _name.lower().strip()
        _VARIANT_LOOKUP[_n] = _gidx


def are_firstname_variants(fn1: str, fn2: str) -> bool:
    """Check if two first names are known variants of each other."""
    n1 = fn1.lower().strip()
    n2 = fn2.lower().strip()
    if n1 == n2:
        return True
    g1 = _VARIANT_LOOKUP.get(n1)
    g2 = _VARIANT_LOOKUP.get(n2)
    if g1 is not None and g2 is not None and g1 == g2:
        return True
    # One is a prefix of the other with ≥3 chars (e.g. "Rob" / "Robert")
    if len(n1) >= 3 and len(n2) >= 3:
        if n1.startswith(n2) or n2.startswith(n1):
            return True
    return False


# ── Name normalisation ────────────────────────────────────────────────────

def normalize_name(name: str) -> str:
    if not name:
        return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def split_name(name: str) -> tuple[str, str]:
    """Split display name into (firstname, surname).  Returns normalised."""
    parts = name.strip().split()
    if not parts:
        return ("", "")
    if len(parts) == 1:
        return ("", normalize_name(parts[0]))
    return (normalize_name(" ".join(parts[:-1])), normalize_name(parts[-1]))


def is_garbage_name(name: str) -> bool:
    if not name or not name.strip():
        return True
    cleaned = name.strip()
    garbage = {
        "party", "(politician)", "ireland", "voice", "fein",
        "fáin", "ecology", "clubs", "labour", "conservative",
    }
    return cleaned.lower().strip("() ") in garbage or len(cleaned) <= 2


def pid_type(pid: str) -> str:
    if not pid:
        return "none"
    if pid.startswith("T"):
        return "temp"
    if pid.isdigit():
        return "curated" if len(pid) <= 6 else "hash"
    return "other"


def names_are_same_person(name1: str, name2: str) -> bool:
    """Determine if two names could plausibly refer to the same person."""
    n1 = normalize_name(name1)
    n2 = normalize_name(name2)
    if n1 == n2:
        return True

    fn1, sn1 = split_name(name1)
    fn2, sn2 = split_name(name2)

    # Different surname → definitely different people
    if sn1 != sn2:
        return False

    # Same surname, check first names
    return are_firstname_variants(fn1, fn2)


# ── Main processing ───────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Unify PersonIDs v3")
    parser.add_argument("input", help="Input XLSX (original comprehensive workbook)")
    parser.add_argument("output", help="Output XLSX path")
    parser.add_argument("--log", default="personid_v3_log.json", help="Change log")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    log_path = Path(args.log)

    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb["ElectionResults"]

    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))

    pid_col = col["PersonID"]
    rt_col = col["ResultType"]
    name_col = col["Name usually known by"]
    party_col = col["Party Name"]
    date_col = col["Date"]
    body_col = col["ElectedBody"]
    const_col = col["Constituency"]

    log: dict[str, Any] = {"phases": {}}

    # ── Phase 1a: Remove garbage rows ─────────────────────────────────────
    print("\nPhase 1a: Removing garbage rows...")
    garbage_count = 0
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        if is_garbage_name(row[name_col]):
            rows[i][rt_col] = "GarbageRemoved"
            garbage_count += 1
    log["phases"]["1a"] = {"garbage_removed": garbage_count}
    print(f"  Removed {garbage_count} garbage rows")

    # ── Phase 1b: Strip T-prefix ──────────────────────────────────────────
    print("\nPhase 1b: Stripping T-prefix...")
    t_count = 0
    for i, row in enumerate(rows):
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid.startswith("T"):
            rows[i][pid_col] = pid[1:]
            t_count += 1
    log["phases"]["1b"] = {"t_stripped": t_count}
    print(f"  Stripped {t_count} T-prefixes")

    # ── Phase 1c: Exact normalised name merge into curated IDs ────────────
    print("\nPhase 1c: Merging by exact normalised name into curated IDs...")

    # Build norm_name → curated PID (only from curated-ID rows)
    norm_to_curated: dict[str, str] = {}
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) != "curated":
            continue
        name = row[name_col]
        if not name:
            continue
        n = normalize_name(name)
        if n and n not in norm_to_curated:
            norm_to_curated[n] = pid

    exact_merged = 0
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) == "curated":
            continue
        name = row[name_col]
        if not name:
            continue
        n = normalize_name(name)
        if n in norm_to_curated:
            rows[i][pid_col] = norm_to_curated[n]
            exact_merged += 1
    log["phases"]["1c"] = {"exact_merged": exact_merged, "curated_names": len(norm_to_curated)}
    print(f"  Merged {exact_merged} rows ({len(norm_to_curated)} curated names available)")

    # ── Phase 2: Merge by first-name variants into curated IDs ────────────
    print("\nPhase 2: Merging by first-name variants into curated IDs...")

    # Build surname → [(curated_pid, firstname_norm, display_name)] index
    surname_index: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) != "curated":
            continue
        name = row[name_col] or ""
        fn, sn = split_name(name)
        if sn and fn:
            surname_index[sn].append((pid, fn, name))

    variant_merged = 0
    variant_log: list[dict] = []
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) == "curated":
            continue
        name = row[name_col] or ""
        fn, sn = split_name(name)
        if not sn or not fn:
            continue

        candidates = surname_index.get(sn, [])
        matches = [(cpid, cfn, cname) for cpid, cfn, cname in candidates
                    if are_firstname_variants(fn, cfn)]

        # Only merge if exactly one curated PID matches
        matched_pids = set(cpid for cpid, _, _ in matches)
        if len(matched_pids) == 1:
            target_pid = matched_pids.pop()
            rows[i][pid_col] = target_pid
            variant_merged += 1
            variant_log.append({
                "name": name, "matched_to": matches[0][2],
                "curated_pid": target_pid,
            })

    log["phases"]["2"] = {"variant_merged": variant_merged, "details": variant_log[:50]}
    print(f"  Merged {variant_merged} rows via first-name variants")

    # ── Phase 3: Unify hash IDs sharing a normalised name ─────────────────
    print("\nPhase 3: Unifying hash IDs sharing a normalised name...")

    norm_to_hash: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) != "hash":
            continue
        name = row[name_col] or ""
        n = normalize_name(name)
        if n:
            norm_to_hash[n].add(pid)

    unify_map: dict[str, str] = {}
    for n, pids in norm_to_hash.items():
        if len(pids) <= 1:
            continue
        canonical = sorted(pids)[0]
        for pid in pids:
            if pid != canonical:
                unify_map[pid] = canonical

    unified = 0
    for i, row in enumerate(rows):
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid in unify_map:
            rows[i][pid_col] = unify_map[pid]
            unified += 1

    log["phases"]["3"] = {"hash_unified": len(unify_map), "rows": unified}
    print(f"  Unified {len(unify_map)} duplicate hash PIDs ({unified} rows)")

    # ── Phase 3b: Merge hash IDs into curated IDs via first-name variants ─
    print("\nPhase 3b: Second pass — merging remaining hash IDs via variants...")

    # Rebuild surname index with any newly-curated entries
    surname_index2: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) != "curated":
            continue
        name = row[name_col] or ""
        fn, sn = split_name(name)
        if sn and fn:
            # Deduplicate
            key = (pid, sn, fn)
            existing = [(p, f, n) for p, f, n in surname_index2[sn] if p == pid and f == fn]
            if not existing:
                surname_index2[sn].append((pid, fn, name))

    pass2_merged = 0
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) == "curated":
            continue
        name = row[name_col] or ""
        fn, sn = split_name(name)
        if not sn or not fn:
            continue

        candidates = surname_index2.get(sn, [])
        matches = [(cpid, cfn, cname) for cpid, cfn, cname in candidates
                    if are_firstname_variants(fn, cfn)]
        matched_pids = set(cpid for cpid, _, _ in matches)
        if len(matched_pids) == 1:
            rows[i][pid_col] = matched_pids.pop()
            pass2_merged += 1

    log["phases"]["3b"] = {"pass2_merged": pass2_merged}
    print(f"  Merged {pass2_merged} more rows")

    # ── Phase 4: Split collisions ─────────────────────────────────────────
    print("\nPhase 4: Splitting collisions (PIDs with different people)...")

    # Find PIDs with multiple genuinely-different names
    pid_name_groups: dict[str, list[tuple[str, list[int]]]] = defaultdict(list)
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        name = row[name_col] or ""
        if not pid or not name:
            continue

        # Try to add to existing group for this PID
        found_group = False
        for group_name, group_indices in pid_name_groups[pid]:
            if names_are_same_person(name, group_name):
                group_indices.append(i)
                found_group = True
                break
        if not found_group:
            pid_name_groups[pid].append((name, [i]))

    # Find the max curated ID for fresh ID assignment
    max_curated = 0
    for row in rows:
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid.isdigit() and len(pid) <= 6:
            v = int(pid)
            if v > max_curated:
                max_curated = v

    next_id = max(max_curated + 1, 200001)  # Start fresh IDs at 200001

    splits = 0
    split_log: list[dict] = []
    for pid, groups in pid_name_groups.items():
        if len(groups) <= 1:
            continue
        # Multiple distinct-name groups share this PID — collision!
        # Keep the first group on the original PID, split the rest
        for group_name, group_indices in groups[1:]:
            new_pid = str(next_id)
            next_id += 1
            for idx in group_indices:
                rows[idx][pid_col] = new_pid
            splits += 1
            split_log.append({
                "original_pid": pid,
                "kept_name": groups[0][0],
                "split_name": group_name,
                "new_pid": new_pid,
                "rows": len(group_indices),
            })

    log["phases"]["4"] = {"collisions_split": splits, "splits": split_log}
    print(f"  Split {splits} collisions into new PIDs")

    # ── Phase 5: Assign fresh IDs to remaining hash PIDs ──────────────────
    print("\nPhase 5: Assigning fresh sequential IDs...")

    hash_to_fresh: dict[str, str] = {}
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) != "hash":
            continue
        if pid not in hash_to_fresh:
            hash_to_fresh[pid] = str(next_id)
            next_id += 1
        rows[i][pid_col] = hash_to_fresh[pid]

    log["phases"]["5"] = {"fresh_ids": len(hash_to_fresh), "id_range": f"200001-{next_id - 1}"}
    print(f"  Assigned {len(hash_to_fresh)} fresh IDs (200001-{next_id - 1})")

    # ── Update Transfers sheet ────────────────────────────────────────────
    print("\nUpdating Transfers sheet...")

    # Build name → final PID mapping
    name_to_final: dict[str, str] = {}
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        name = row[name_col]
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if name and pid:
            n = normalize_name(name)
            if n:
                name_to_final[n] = pid

    ws_t = wb["Transfers"]
    t_headers = [cell.value for cell in ws_t[1]]
    t_pid_idx = t_headers.index("PersonID") if "PersonID" in t_headers else None
    t_name_idx = t_headers.index("Name") if "Name" in t_headers else None

    t_updates = 0
    if t_pid_idx is not None and t_name_idx is not None:
        for t_row in ws_t.iter_rows(min_row=2):
            name_val = t_row[t_name_idx].value
            if name_val:
                n = normalize_name(name_val)
                if n in name_to_final:
                    old_val = str(t_row[t_pid_idx].value) if t_row[t_pid_idx].value else ""
                    new_val = name_to_final[n]
                    if old_val != new_val:
                        t_row[t_pid_idx].value = new_val
                        t_updates += 1

    print(f"  Updated {t_updates} Transfers PersonIDs")

    # ── Write results ─────────────────────────────────────────────────────
    print(f"\nWriting {output_path}...")
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=i + 2, column=j + 1, value=val)
    wb.save(output_path)

    # ── Final stats ───────────────────────────────────────────────────────
    final_pids = set()
    final_names = set()
    pid_type_counts = defaultdict(int)
    pid_multi_names: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        name = row[name_col] or ""
        pt = pid_type(pid)
        pid_type_counts[pt] += 1
        final_pids.add(pid)
        final_names.add(name)
        if name:
            pid_multi_names[pid].add(name)

    multi_name_pids = {p: names for p, names in pid_multi_names.items() if len(names) > 1}
    # Check for remaining collisions
    remaining_collisions = 0
    for pid, names in multi_name_pids.items():
        name_list = sorted(names)
        for i in range(len(name_list)):
            for j in range(i + 1, len(name_list)):
                if not names_are_same_person(name_list[i], name_list[j]):
                    remaining_collisions += 1
                    break

    log["summary"] = {
        "total_candidates": sum(pid_type_counts.values()),
        "pid_types": dict(pid_type_counts),
        "unique_pids": len(final_pids),
        "unique_names": len(final_names),
        "pids_with_multiple_names": len(multi_name_pids),
        "remaining_collisions": remaining_collisions,
    }

    print(f"\n{'='*60}")
    print(f"  FINAL SUMMARY")
    print(f"{'='*60}")
    print(f"  Total candidate rows: {sum(pid_type_counts.values())}")
    print(f"  PID types: {dict(pid_type_counts)}")
    print(f"  Unique PersonIDs: {len(final_pids)}")
    print(f"  Unique names: {len(final_names)}")
    print(f"  PIDs with multiple name variants: {len(multi_name_pids)}")
    print(f"  Remaining collisions: {remaining_collisions}")
    print(f"  Output: {output_path}")

    log_path.write_text(json.dumps(log, indent=2, default=str, ensure_ascii=False), encoding="utf-8")
    print(f"  Change log: {log_path}")


if __name__ == "__main__":
    main()
