"""Apply the second requested same-name identity split/merge corrections."""

from __future__ import annotations

import json
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class SplitContextAssignment:
    label: str
    old_id: int
    new_id: int
    name: str
    contexts: tuple[tuple[str, str, str], ...]  # date, constituency, party
    website_json_files: tuple[str, ...]


@dataclass(frozen=True)
class LocalContextRemap:
    label: str
    name: str
    target_id: int
    contexts: tuple[tuple[str, str, str], ...]  # date, constituency, party


SPLITS: tuple[SplitContextAssignment, ...] = (
    SplitContextAssignment(
        label="David Taylor Green / Ecology",
        old_id=18241,
        new_id=100012,
        name="David Taylor",
        contexts=(("1996-05-30", "Belfast West", "Green / Ecology"),),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-forum-for-political-dialogue/1996-05-30/belfast-west.json",
        ),
    ),
    SplitContextAssignment(
        label="David Taylor UKUP",
        old_id=18241,
        new_id=100013,
        name="David Taylor",
        contexts=(("1996-05-30", "Foyle", "UKUP"),),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-forum-for-political-dialogue/1996-05-30/foyle.json",
        ),
    ),
    SplitContextAssignment(
        label="Glenn Barr Vanguard",
        old_id=7192,
        new_id=100014,
        name="Glenn Barr",
        contexts=(
            ("1973-06-28", "Londonderry", "Vanguard Unionist Progressive Party"),
            ("1975-05-01", "Londonderry", "Vanguard Unionist Progressive Party"),
        ),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-assembly/1973-06-28/londonderry.json",
            r"election-viewer-package/data/elections/northern-ireland-constitutional-convention/1975-05-01/londonderry.json",
        ),
    ),
    SplitContextAssignment(
        label="John Doherty Workers Party / Republican Clubs",
        old_id=82766,
        new_id=100015,
        name="John Doherty",
        contexts=(("1996-05-30", "West Tyrone", "Workers Party / Republican Clubs"),),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-forum-for-political-dialogue/1996-05-30/west-tyrone.json",
        ),
    ),
    SplitContextAssignment(
        label="Martin Kelly CISTA",
        old_id=33545,
        new_id=100016,
        name="Martin Kelly",
        contexts=(
            ("2015-05-07", "Upper Bann", "CISTA"),
            ("2016-05-05", "Upper Bann", "CISTA"),
        ),
        website_json_files=(
            r"election-viewer-package/data/elections/house-of-commons-of-the-united-kingdom/2015-05-07/upper-bann.json",
            r"election-viewer-package/data/elections/northern-ireland-assembly/2016-05-05/upper-bann.json",
        ),
    ),
    SplitContextAssignment(
        label="Stephen Nicholl UKUP",
        old_id=41828,
        new_id=100017,
        name="Stephen Nicholl",
        contexts=(
            ("1996-05-30", "Northern Ireland", "UKUP"),
            ("1996-05-30", "South Antrim", "UKUP"),
        ),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-forum-for-political-dialogue/1996-05-30/northern-ireland.json",
            r"election-viewer-package/data/elections/northern-ireland-forum-for-political-dialogue/1996-05-30/south-antrim.json",
        ),
    ),
    SplitContextAssignment(
        label="Thomas Burns DUP",
        old_id=63034,
        new_id=100018,
        name="Thomas Burns",
        contexts=(
            ("1973-06-28", "Belfast South", "DUP"),
            ("1975-05-01", "Belfast South", "DUP"),
        ),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-assembly/1973-06-28/belfast-south.json",
            r"election-viewer-package/data/elections/northern-ireland-constitutional-convention/1975-05-01/belfast-south.json",
        ),
    ),
    SplitContextAssignment(
        label="Robert Stewart UUP",
        old_id=42476,
        new_id=100019,
        name="Robert Stewart",
        contexts=(("1973-06-28", "Belfast South", "UUP"),),
        website_json_files=(
            r"election-viewer-package/data/elections/northern-ireland-assembly/1973-06-28/belfast-south.json",
        ),
    ),
)


LOCAL_CONTEXT_REMAPS: tuple[LocalContextRemap, ...] = (
    LocalContextRemap(
        label="David Taylor UUP local",
        name="David Taylor",
        target_id=18241,
        contexts=(
            ("2014-05-22", "Slieve Gullion", "UUP"),
            ("2019-05-02", "Slieve Gullion", "UUP"),
            ("2023-05-18", "Slieve Gullion", "UUP"),
        ),
    ),
    LocalContextRemap(
        label="Glenn Barr UUP local",
        name="Glenn Barr",
        target_id=7192,
        contexts=(
            ("2014-05-22", "Banbridge", "UUP"),
            ("2019-05-02", "Banbridge", "UUP"),
            ("2023-05-18", "Banbridge", "UUP"),
        ),
    ),
    LocalContextRemap(
        label="John Boyle Aontu local",
        name="John Boyle",
        target_id=49548,
        contexts=(("2023-05-18", "Limavady", "Aontú"),),
    ),
    LocalContextRemap(
        label="John Doherty Alliance local",
        name="John Doherty",
        target_id=82766,
        contexts=(("2019-05-02", "Foyleside", "Alliance"),),
    ),
    LocalContextRemap(
        label="Martin Kelly Aontu local",
        name="Martin Kelly",
        target_id=33545,
        contexts=(("2019-05-02", "Armagh", "Aontú"),),
    ),
    LocalContextRemap(
        label="Stephen Nicholl UUP local",
        name="Stephen Nicholl",
        target_id=41828,
        contexts=(
            ("2014-05-22", "Ballymena", "UUP"),
            ("2019-05-02", "Ballymena", "UUP"),
        ),
    ),
    LocalContextRemap(
        label="Thomas Burns SDLP local",
        name="Thomas Burns",
        target_id=63034,
        contexts=(
            ("2014-05-22", "Airport", "SDLP"),
            ("2019-05-02", "Airport", "SDLP"),
            ("2023-05-18", "Airport", "SDLP"),
        ),
    ),
)


FULL_WORKBOOK = Path("Full election tables.xlsx")
LOCAL_WORKBOOK = Path(r"_tmp_xls2rar_extract/out/wiki_lgov_modern/lgov-modern-wikipedia.stvfix.xlsx")


def backup_file(src: Path, backup_root: Path) -> None:
    backup_root.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, backup_root / src.name)


def normalize_space(value: Any) -> str:
    return " ".join(str(value or "").split())


def format_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def replace_exact_id_token(text: str, old_id: int, new_id: int) -> tuple[str, bool]:
    old = str(old_id)
    new = str(new_id)
    updated: list[str] = []
    changed = False
    token = ""
    for ch in text:
        if ch.isdigit():
            token += ch
            continue
        if token:
            if token == old:
                updated.append(new)
                changed = True
            else:
                updated.append(token)
            token = ""
        updated.append(ch)
    if token:
        if token == old:
            updated.append(new)
            changed = True
        else:
            updated.append(token)
    return "".join(updated), changed


def replace_cell_value(value: Any, old_id: int, new_id: int) -> tuple[Any, bool]:
    if value is None or value == "":
        return value, False
    if isinstance(value, bool):
        return value, False
    if isinstance(value, int):
        return (new_id, True) if value == old_id else (value, False)
    if isinstance(value, float):
        if value.is_integer() and int(value) == old_id:
            return float(new_id), True
        return value, False
    if isinstance(value, str):
        return replace_exact_id_token(value, old_id, new_id)
    return value, False


def context_key(date_text: str, constituency: str, party: str) -> str:
    return f"{date_text}|{normalize_space(constituency)}|{normalize_space(party)}"


def derive_row_context(row_values: dict[str, Any]) -> str | None:
    date_text = format_date(row_values.get("Date"))
    constituency = row_values.get("Constituency")
    party = row_values.get("Party Name") or row_values.get("Source Party Name")
    if date_text and constituency and party:
        return context_key(date_text, str(constituency), str(party))
    return None


def build_target_context_sets() -> dict[str, set[str]]:
    result: dict[str, set[str]] = {}
    for assignment in SPLITS:
        result[assignment.label] = {
            context_key(date_text, constituency, party)
            for date_text, constituency, party in assignment.contexts
        }
    return result


def build_local_context_sets() -> dict[str, set[str]]:
    result: dict[str, set[str]] = {}
    for assignment in LOCAL_CONTEXT_REMAPS:
        result[assignment.label] = {
            context_key(date_text, constituency, party)
            for date_text, constituency, party in assignment.contexts
        }
    return result


def patch_full_workbook(path: Path, backup_root: Path) -> dict[str, int]:
    backup_file(path, backup_root)
    wb = load_workbook(path)
    context_sets = build_target_context_sets()
    stats: dict[str, int] = defaultdict(int)

    for sheet_name in wb.sheetnames:
        if sheet_name == "Names":
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        idx = {header: pos + 1 for pos, header in enumerate(headers)}
        for row_num in range(2, ws.max_row + 1):
            row_values = {header: ws.cell(row=row_num, column=col).value for header, col in idx.items()}
            row_context = derive_row_context(row_values)
            if not row_context:
                continue
            for assignment in SPLITS:
                if normalize_space(row_values.get("Name usually known by")) != assignment.name:
                    continue
                if row_context not in context_sets[assignment.label]:
                    continue
                for header, col in idx.items():
                    cell = ws.cell(row=row_num, column=col)
                    new_value, changed = replace_cell_value(cell.value, assignment.old_id, assignment.new_id)
                    if changed:
                        cell.value = new_value
                        stats[f"{sheet_name}_cell_updates"] += 1

    ws_names = wb["Names"]
    name_headers = [cell.value for cell in ws_names[1]]
    name_idx = {header: pos + 1 for pos, header in enumerate(name_headers)}

    existing_name_rows: dict[tuple[str, int], int] = {}
    for row_num in range(2, ws_names.max_row + 1):
        full_name = ws_names.cell(row=row_num, column=name_idx["Full Name usually known by"]).value
        person_id = ws_names.cell(row=row_num, column=name_idx["PersonID"]).value
        if full_name and person_id not in (None, ""):
            existing_name_rows[(str(full_name), int(person_id))] = row_num

    for assignment in SPLITS:
        old_key = (assignment.name, assignment.old_id)
        new_key = (assignment.name, assignment.new_id)
        if new_key in existing_name_rows:
            continue
        old_row_num = existing_name_rows.get(old_key)
        if not old_row_num:
            raise ValueError(f"Missing Names row for {assignment.name} / {assignment.old_id}")
        row_values = [ws_names.cell(row=old_row_num, column=col).value for col in range(1, ws_names.max_column + 1)]
        row_values[name_idx["PersonID"] - 1] = assignment.new_id
        ws_names.append(row_values)
        existing_name_rows[new_key] = ws_names.max_row
        stats["Names_rows_added"] += 1

    ws_results = wb["ElectionResults"]
    result_headers = [cell.value for cell in ws_results[1]]
    result_idx = {header: pos + 1 for pos, header in enumerate(result_headers)}
    counts: dict[int, dict[str, int]] = defaultdict(lambda: {"devolved": 0, "westminster": 0, "european": 0, "total": 0})
    for row_num in range(2, ws_results.max_row + 1):
        result_type = str(ws_results.cell(row=row_num, column=result_idx["ResultType"]).value or "")
        if not (
            result_type == "Candidate"
            or result_type.startswith("ListCandidate")
            or result_type.startswith("RegionalListCandidate")
        ):
            continue
        person_id = ws_results.cell(row=row_num, column=result_idx["PersonID"]).value
        if person_id in (None, ""):
            continue
        person_id = int(person_id)
        body = str(ws_results.cell(row=row_num, column=result_idx["ElectedBody"]).value or "")
        if "House of Commons" in body:
            counts[person_id]["westminster"] += 1
        elif "European" in body:
            counts[person_id]["european"] += 1
        else:
            counts[person_id]["devolved"] += 1
        counts[person_id]["total"] += 1

    for assignment in SPLITS:
        for pid in (assignment.old_id, assignment.new_id):
            row_num = existing_name_rows[(assignment.name, pid)]
            ws_names.cell(row=row_num, column=name_idx["Times stood for devolved bodies"]).value = counts[pid]["devolved"]
            ws_names.cell(row=row_num, column=name_idx["Times stood for Westminster"]).value = counts[pid]["westminster"]
            ws_names.cell(row=row_num, column=name_idx["Times stood for European Parliament"]).value = counts[pid]["european"]
            ws_names.cell(row=row_num, column=name_idx["Total times stood"]).value = counts[pid]["total"]
            stats["Names_rows_recounted"] += 1

    temp_path = path.with_suffix(".tmp.xlsx")
    wb.save(temp_path)
    load_workbook(temp_path, read_only=True).close()
    temp_path.replace(path)
    return dict(stats)


def replace_json_values(obj: Any, old_id: int, new_id: int) -> tuple[Any, int]:
    changes = 0
    if isinstance(obj, dict):
        out = {}
        for key, value in obj.items():
            new_value, delta = replace_json_values(value, old_id, new_id)
            out[key] = new_value
            changes += delta
        return out, changes
    if isinstance(obj, list):
        out = []
        for item in obj:
            new_item, delta = replace_json_values(item, old_id, new_id)
            out.append(new_item)
            changes += delta
        return out, changes
    new_value, changed = replace_cell_value(obj, old_id, new_id)
    return new_value, int(changed)


def patch_website_json(backup_root: Path) -> dict[str, int]:
    stats: dict[str, int] = defaultdict(int)
    for assignment in SPLITS:
        for rel_path in assignment.website_json_files:
            path = Path(rel_path)
            backup_file(path, backup_root / "website-json")
            payload = json.loads(path.read_text(encoding="utf-8"))
            updated, changes = replace_json_values(payload, assignment.old_id, assignment.new_id)
            path.write_text(json.dumps(updated, ensure_ascii=False), encoding="utf-8")
            stats[str(path)] += changes
    return dict(stats)


def remap_scalar(value: Any, mapping: dict[int, int]) -> tuple[Any, bool]:
    if value in (None, ""):
        return value, False
    try:
        current = int(value)
    except (TypeError, ValueError):
        return value, False
    new_value = mapping.get(current)
    if new_value is None or new_value == current:
        return value, False
    return new_value, True


def remap_id_list(value: Any, mapping: dict[int, int]) -> tuple[Any, bool]:
    if value in (None, ""):
        return value, False
    parts = [part.strip() for part in str(value).split(",")]
    changed = False
    remapped = []
    for part in parts:
        if not part:
            remapped.append(part)
            continue
        try:
            current = int(part)
        except ValueError:
            remapped.append(part)
            continue
        new_value = mapping.get(current, current)
        if new_value != current:
            changed = True
        remapped.append(str(new_value))
    return ", ".join(remapped), changed


def patch_local_workbook(path: Path, backup_root: Path) -> dict[str, int]:
    backup_file(path, backup_root)
    wb = load_workbook(path)
    ws_results = wb["ElectionResults"]
    result_headers = [cell.value for cell in ws_results[1]]
    result_idx = {header: pos + 1 for pos, header in enumerate(result_headers)}
    context_sets = build_local_context_sets()

    rows_for_remap: dict[int, int] = {}
    id_mapping: dict[int, int] = {}
    candidate_rows: dict[str, list[tuple[int, str, int]]] = defaultdict(list)

    for row_num in range(2, ws_results.max_row + 1):
        if ws_results.cell(row=row_num, column=result_idx["ResultType"]).value != "Candidate":
            continue
        name = normalize_space(ws_results.cell(row=row_num, column=result_idx["Name usually known by"]).value)
        party = normalize_space(ws_results.cell(row=row_num, column=result_idx["Party Name"]).value)
        constituency = normalize_space(ws_results.cell(row=row_num, column=result_idx["Constituency"]).value)
        date_text = format_date(ws_results.cell(row=row_num, column=result_idx["Date"]).value)
        pid = ws_results.cell(row=row_num, column=result_idx["PersonID"]).value
        if pid in (None, ""):
            continue
        pid = int(pid)
        candidate_rows[name].append((row_num, context_key(date_text, constituency, party), pid))

    for assignment in LOCAL_CONTEXT_REMAPS:
        for row_num, row_context, current_id in candidate_rows.get(assignment.name, []):
            if row_context in context_sets[assignment.label]:
                rows_for_remap[row_num] = assignment.target_id
                id_mapping[current_id] = assignment.target_id

    result_transfer_subject_headers = [
        header for header in result_headers if isinstance(header, str) and header.startswith("TransferSubject")
    ]
    stats: dict[str, int] = defaultdict(int)

    for row_num, target_id in rows_for_remap.items():
        person_cell = ws_results.cell(row=row_num, column=result_idx["PersonID"])
        old_id = int(person_cell.value)
        if old_id == target_id:
            continue
        person_cell.value = target_id
        stats["ElectionResults_PersonID_updates"] += 1
        for header in result_transfer_subject_headers:
            cell = ws_results.cell(row=row_num, column=result_idx[header])
            new_value, changed = remap_scalar(cell.value, {old_id: target_id})
            if changed:
                cell.value = new_value
                stats["ElectionResults_TransferSubject_updates"] += 1

    ws_transfers = wb["Transfers"]
    transfer_headers = [cell.value for cell in ws_transfers[1]]
    transfer_idx = {header: pos + 1 for pos, header in enumerate(transfer_headers)}

    for row_num in range(2, ws_transfers.max_row + 1):
        for header in ("PersonID", "TransferSubject", "SourcePersonID"):
            cell = ws_transfers.cell(row=row_num, column=transfer_idx[header])
            new_value, changed = remap_scalar(cell.value, id_mapping)
            if changed:
                cell.value = new_value
                stats[f"Transfers_{header}_updates"] += 1
        remaining = ws_transfers.cell(row=row_num, column=transfer_idx["RemainingCandidateIDsDesc"])
        new_value, changed = remap_id_list(remaining.value, id_mapping)
        if changed:
            remaining.value = new_value
            stats["Transfers_RemainingCandidateIDsDesc_updates"] += 1

    temp_path = path.with_suffix(".tmp.xlsx")
    wb.save(temp_path)
    load_workbook(temp_path, read_only=True).close()
    temp_path.replace(path)
    stats["local_context_mapping_count"] = len(rows_for_remap)
    return dict(stats)


def main() -> None:
    timestamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    backup_root = Path("backups") / f"requested-identity-fixes-round2-{timestamp}"
    stats = {
        "full_workbook": patch_full_workbook(FULL_WORKBOOK, backup_root / "full-workbook"),
        "website_json": patch_website_json(backup_root),
        "local_workbook": patch_local_workbook(LOCAL_WORKBOOK, backup_root / "local-workbook"),
    }
    print(json.dumps(stats, indent=2))


if __name__ == "__main__":
    main()
