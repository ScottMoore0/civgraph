#!/usr/bin/env python
"""Deep analysis for v8 — exhaustive search for remaining issues."""

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

def norm(name):
    if not name: return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()

def split_fn_sn(name):
    parts = name.strip().split()
    if not parts: return ("", "")
    if len(parts) == 1: return ("", norm(parts[0]))
    return (norm(" ".join(parts[:-1])), norm(parts[-1]))

def normalize_party(p):
    if not p: return ""
    l = p.lower()
    for needle, code in [("sdlp","sdlp"),("democratic unionist","dup"),("dup","dup"),
        ("ulster unionist","uup"),("uup","uup"),("sinn","sf"),("alliance","all"),
        ("independent","ind"),("green","grn"),("tuv","tuv"),("workers","wp"),
        ("pup","pup"),("labour","lab"),("conservative","con"),("ukip","ukip"),
        ("ukup","ukup"),("vanguard","vupp"),("republican clubs","wp"),
        ("socialist","soc"),("people before","pbp"),("national front","nf")]:
        if needle in l: return code
    return l[:15]

VGROUPS = [
    ["james","jim","jimmy","seamus"],["william","will","willie","bill","billy","liam"],
    ["robert","rob","robbie","bob","bobby"],["john","johnny","jack","sean","shane"],
    ["thomas","tom","tommy"],["edward","ed","eddie","ted","eamonn","eamon","edmund"],
    ["richard","dick","dickie","ricky"],["charles","charlie","cathal"],
    ["patrick","paddy","pat","patsy","padraig"],["michael","mike","mick","mickey"],
    ["christopher","chris"],["daniel","dan","danny"],["andrew","andy","drew"],
    ["anthony","tony"],["joseph","joe","joey"],["samuel","sam","sammy"],
    ["alexander","alex","alec","alistair","alastair"],["frederick","fred","freddie"],
    ["kenneth","ken","kenny"],["ronald","ron","ronnie"],["raymond","ray"],
    ["stephen","steve","steven"],["philip","phil"],["lawrence","larry","laurence"],
    ["gerald","gerry","gerard","gearoid"],["peter","pete"],["david","dave","davy"],
    ["henry","harry","hal"],["albert","bert","bertie"],["alfred","alf","alfie"],
    ["francis","frank","frankie","proinsias"],["desmond","des"],["stanley","stan"],
    ["ernest","ernie"],["norman","norm"],["donald","don","donnie"],["dennis","denis"],
    ["terence","terry","terrence"],["vincent","vince"],["douglas","doug"],
    ["timothy","tim"],["nicholas","nick"],["geoffrey","geoff","jeff","jeffrey"],
    ["elizabeth","liz","lizzie","beth","betty","bess"],
    ["margaret","maggie","meg","peggy","madge"],
    ["catherine","kate","katie","cathy","kathleen"],
    ["patricia","pat","tricia","trish"],["dorothy","dot","dolly"],
    ["jacqueline","jackie"],["anne","ann","annie"],
    ["mary","molly","may","maire"],["bridget","brid","bridie"],
    ["niall","neal","neil"],["ciaran","kieran"],
    ["archibald","archie"],["reginald","reg","reggie"],
    ["leonard","len","lenny"],["bernard","bernie"],
    ["pamela","pam"],["susan","sue","susie"],["deborah","deb","debbie"],
    ["christine","chris","christina","tina"],["rosemary","rosie"],["caroline","carol"],
    ["joanna","jo","joanne","joan"],["jennifer","jenny","jen"],
    ["helena","helen"],["victoria","vicky"],
    ["cornelius","con"],["bartholomew","bart"],["roderick","roddy","rod"],
    ["wallace","wally"],["maurice","morrie"],["cecil","cec"],
    ["clifford","cliff"],["sydney","sid"],["percy","percival"],
    ["owen","eugene"],["colm","columba"],["declan","dec"],
    ["cathal","charles"],["malachy","malachi"],
]
_VM = {}
for gi, grp in enumerate(VGROUPS):
    for n in grp: _VM[n.lower()] = gi

def are_fn_var(f1, f2):
    n1, n2 = f1.lower().strip(), f2.lower().strip()
    if n1 == n2: return True
    g1, g2 = _VM.get(n1), _VM.get(n2)
    if g1 is not None and g2 is not None and g1 == g2: return True
    if len(n1) >= 3 and len(n2) >= 3 and (n1.startswith(n2) or n2.startswith(n1)): return True
    return False

# Load ARK
ark_lookup = json.load(open("_tmp_ark_name_lookup.json", encoding="utf-8"))
ark_candidates = json.load(open("_tmp_ark_candidates.json", encoding="utf-8"))
ark_ctx = {}
for c in ark_candidates:
    full = c["full_name"]
    parts = full.split()
    if len(parts) < 2: continue
    short = f"{parts[0]} {parts[-1]}"
    ark_ctx[(norm(short), norm(c["constituency"]), c["year"])] = norm(full)


print("Loading workbook...")
wb = openpyxl.load_workbook("Full election tables - comprehensive - personid-v8.xlsx", read_only=True)
ws = wb["ElectionResults"]
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col = {name: i for i, name in enumerate(headers)}

records = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    if row[col["ResultType"]] != "Candidate": continue
    date = str(row[col["Date"]] or "")
    year = date[:4] if date[:4].isdigit() else ""
    records.append({
        "idx": i, "pid": str(row[col["PersonID"]]),
        "name": row[col["Name usually known by"]] or "",
        "party": row[col["Party Name"]] or "",
        "date": date, "year": year,
        "body": row[col["ElectedBody"]] or "",
        "const": row[col["Constituency"]] or "",
    })

profiles = defaultdict(lambda: {
    "names": set(), "parties": set(), "years": set(),
    "bodies": set(), "consts": set(), "rows": 0,
})
for r in records:
    p = profiles[r["pid"]]
    p["names"].add(r["name"])
    p["parties"].add(r["party"])
    if r["year"]: p["years"].add(int(r["year"]))
    p["bodies"].add(r["body"])
    p["consts"].add(r["const"])
    p["rows"] += 1

# ══════════════════════════════════════════════════════════════════════════
# 1. EXHAUSTIVE variant-merge search: find ALL pairs of PIDs where
#    names are variants and party+temporal signals are strong
# ══════════════════════════════════════════════════════════════════════════
print("\n=== 1. Exhaustive variant-merge search ===")

sn_index = defaultdict(list)
for pid, p in profiles.items():
    for name in p["names"]:
        fn, sn = split_fn_sn(name)
        if sn and fn:
            sn_index[sn].append((pid, fn, name))

merge_candidates = []
seen = set()
for sn, entries in sn_index.items():
    pids_set = set(e[0] for e in entries)
    if len(pids_set) <= 1: continue
    for i in range(len(entries)):
        for j in range(i + 1, len(entries)):
            p1, fn1, n1 = entries[i]
            p2, fn2, n2 = entries[j]
            if p1 == p2: continue
            pair = tuple(sorted([p1, p2]))
            if pair in seen: continue
            if not are_fn_var(fn1, fn2): continue
            seen.add(pair)
            pr1, pr2 = profiles[p1], profiles[p2]
            np1 = {normalize_party(x) for x in pr1["parties"]}
            np2 = {normalize_party(x) for x in pr2["parties"]}
            party_match = bool(np1 & np2 - {""})
            y1, y2 = sorted(pr1["years"]), sorted(pr2["years"])
            gap = min(abs(a - b) for a in y1 for b in y2) if y1 and y2 else 999
            geo = bool(pr1["consts"] & pr2["consts"])

            # Check ARK
            ark1 = set()
            ark2 = set()
            for name in pr1["names"]:
                for f in ark_lookup.get(norm(name), []):
                    ark1.add(norm(f))
            for name in pr2["names"]:
                for f in ark_lookup.get(norm(name), []):
                    ark2.add(norm(f))
            ark_confirms = bool(ark1 & ark2)
            ark_denies = bool(ark1 and ark2 and not ark1 & ark2)
            # Refine: ARK only denies if both have 3+ part names with different middles
            if ark_denies:
                real_deny = False
                for a in ark1:
                    for b in ark2:
                        ap, bp = a.split(), b.split()
                        if len(ap) >= 3 and len(bp) >= 3 and ap[-1] == bp[-1] and ap[0] == bp[0]:
                            m1 = " ".join(ap[1:-1])
                            m2 = " ".join(bp[1:-1])
                            if m1 and m2 and m1 != m2 and not m1.startswith(m2[:3]) and not m2.startswith(m1[:3]):
                                real_deny = True
                ark_denies = real_deny

            if ark_denies:
                continue  # ARK says different people

            score = 0
            if party_match: score += 40
            if gap <= 5: score += 20
            elif gap <= 10: score += 15
            elif gap <= 20: score += 10
            if geo: score += 30
            if ark_confirms: score += 10
            score += 10  # name match base

            if score >= 70:
                merge_candidates.append({
                    "p1": p1, "p2": p2, "n1": n1, "n2": n2,
                    "score": score, "gap": gap, "geo": geo,
                    "party_match": party_match, "ark_confirms": ark_confirms,
                    "parties1": sorted(pr1["parties"]), "parties2": sorted(pr2["parties"]),
                    "years1": y1[:3], "years2": y2[:3],
                })

# Exclude already-known intentional splits
INTENTIONAL = {
    "100001","100002","100003","100004","100005","100006","100007",
    "100008","100009","100010","100011","100012","100013","100014",
    "100015","100016","100017","100018","100019",
}
merge_candidates = [m for m in merge_candidates
                    if m["p1"] not in INTENTIONAL and m["p2"] not in INTENTIONAL]

print(f"Merge candidates (score >= 70, ARK not denied): {len(merge_candidates)}")
for m in sorted(merge_candidates, key=lambda x: -x["score"])[:40]:
    ark_tag = " [ARK+]" if m["ark_confirms"] else ""
    print(f"  score={m['score']}: {m['n1']} (PID {m['p1']}) <-> {m['n2']} (PID {m['p2']}){ark_tag}")
    print(f"    parties: {m['parties1'][:3]} vs {m['parties2'][:3]}, gap={m['gap']}yr, geo={m['geo']}")


# ══════════════════════════════════════════════════════════════════════════
# 2. Check for PIDs that are ONLY in one data source but should link
# ══════════════════════════════════════════════════════════════════════════
print(f"\n=== 2. Single-body PIDs that might link to multi-body PIDs ===")
single_body = {pid for pid, p in profiles.items() if len(p["bodies"]) == 1 and p["rows"] == 1}
print(f"Single-appearance PIDs: {len(single_body)}")

# Save
json.dump({
    "merge_candidates": merge_candidates,
    "single_body_count": len(single_body),
}, open("_tmp_v8_analysis.json", "w", encoding="utf-8"),
    indent=2, default=str, ensure_ascii=False)
print(f"\nSaved to _tmp_v8_analysis.json")
