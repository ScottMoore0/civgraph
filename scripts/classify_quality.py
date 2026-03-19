#!/usr/bin/env python
"""Classify every row in the v8 workbook by data quality."""

import json
import re
import unicodedata
from collections import defaultdict, Counter

import openpyxl


def norm(name):
    if not name: return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


# Build ARK name set
ark_candidates = json.load(open("_tmp_ark_candidates.json", encoding="utf-8"))
ark_names = set()
for c in ark_candidates:
    ark_names.add(norm(c["full_name"]))
    parts = c["full_name"].split()
    if len(parts) >= 2:
        ark_names.add(norm(f"{parts[0]} {parts[-1]}"))

wb = openpyxl.load_workbook("Full election tables - comprehensive - personid-v8.xlsx", read_only=True)
ws = wb["ElectionResults"]
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col = {name: i for i, name in enumerate(headers)}

# Build PID profiles
pid_profiles = defaultdict(lambda: {"rows": 0, "bodies": set(), "years": set()})
all_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    all_rows.append(row)
    rt = row[col["ResultType"]]
    if rt != "Candidate":
        continue
    pid = str(row[col["PersonID"]])
    date = str(row[col["Date"]] or "")
    year = date[:4]
    p = pid_profiles[pid]
    p["rows"] += 1
    p["bodies"].add(row[col["ElectedBody"]] or "")
    if year.isdigit():
        p["years"].add(int(year))

# Classify each row
total = 0
garbage = 0
summary = 0
groups = Counter()

for row in all_rows:
    rt = row[col["ResultType"]]
    if rt == "GarbageRemoved":
        garbage += 1
        continue
    if rt != "Candidate":
        summary += 1
        continue
    total += 1

    pid = str(row[col["PersonID"]])
    name = row[col["Name usually known by"]] or ""
    body = row[col["ElectedBody"]] or ""
    date = str(row[col["Date"]] or "")
    year = date[:4] if date[:4].isdigit() else ""
    year_int = int(year) if year else 0

    pid_val = int(pid) if pid.isdigit() else 0
    prof = pid_profiles[pid]

    # PID provenance
    if pid_val < 100000:
        pid_prov = "curated"
    elif pid_val <= 100019:
        pid_prov = "curated"  # intentional splits are still curated
    else:
        pid_prov = "script"

    # Data source
    if body == "Parliament of Northern Ireland":
        data_src = "wiki_stormont"
    elif body == "House of Commons of the United Kingdom" and year_int < 1970:
        data_src = "wiki_westminster"
    elif body == "Local Government" and year_int < 2014:
        data_src = "wiki_lgov"
    else:
        data_src = "election_viewer"

    # ARK coverage
    has_ark = norm(name) in ark_names

    # PID corroboration
    multi_election = len(prof["years"]) > 1
    multi_body = len(prof["bodies"]) > 1

    # Has vote data
    has_votes = row[col["Votes1"]] is not None and row[col["Votes1"]] != 0

    # Has party normalisation
    has_dedup = row[col["Deduplicated Party Name"]] is not None and str(row[col["Deduplicated Party Name"]]).strip() != ""

    pass


# Actually, let's just count the raw dimensions
dims = Counter()
for row in all_rows:
    rt = row[col["ResultType"]]
    if rt != "Candidate":
        continue
    pid = str(row[col["PersonID"]])
    name = row[col["Name usually known by"]] or ""
    body = row[col["ElectedBody"]] or ""
    date = str(row[col["Date"]] or "")
    year = date[:4] if date[:4].isdigit() else ""
    year_int = int(year) if year else 0
    pid_val = int(pid) if pid.isdigit() else 0
    prof = pid_profiles[pid]

    pid_prov = "curated" if pid_val <= 100019 else "script"
    if body == "Parliament of Northern Ireland":
        data_src = "wiki_stormont"
    elif body == "House of Commons of the United Kingdom" and year_int < 1970:
        data_src = "wiki_westminster"
    elif body == "Local Government" and year_int < 2014:
        data_src = "wiki_lgov"
    else:
        data_src = "election_viewer"
    has_ark = norm(name) in ark_names
    multi_elec = len(prof["years"]) > 1
    has_votes = row[col["Votes1"]] is not None and row[col["Votes1"]] != 0
    has_dedup = row[col["Deduplicated Party Name"]] is not None and str(row[col["Deduplicated Party Name"]]).strip() != ""

    dims[(pid_prov, data_src, "ark" if has_ark else "no_ark",
          "multi" if multi_elec else "single",
          "votes" if has_votes else "no_votes",
          "dedup" if has_dedup else "no_dedup")] += 1

print(f"Candidate rows: {total}")
print(f"Summary rows: {summary}")
print(f"Garbage rows: {garbage}")
print()
print("=== RAW DIMENSIONS ===")
for key in sorted(dims.keys(), key=lambda k: -dims[k]):
    print(f"  {' | '.join(key)}: {dims[key]}")

# Simplified grouping
print()
print("=== SIMPLIFIED GROUPING ===")
simple = Counter()
for key, count in dims.items():
    pid_prov, data_src, ark, multi, votes, dedup = key
    if pid_prov == "curated" and data_src == "election_viewer" and multi == "multi":
        simple["A_gold"] += count
    elif pid_prov == "curated" and data_src == "election_viewer" and multi == "single":
        simple["B_curated_single"] += count
    elif pid_prov == "curated" and data_src != "election_viewer" and multi == "multi":
        simple["C_curated_cross_source"] += count
    elif pid_prov == "curated" and data_src != "election_viewer":
        simple["D_curated_scraped"] += count
    elif pid_prov == "script" and ark == "ark" and multi == "multi":
        simple["E_script_ark_multi"] += count
    elif pid_prov == "script" and ark == "ark" and multi == "single":
        simple["F_script_ark_single"] += count
    elif pid_prov == "script" and ark == "no_ark" and multi == "multi":
        simple["G_script_noark_multi"] += count
    elif pid_prov == "script" and ark == "no_ark" and multi == "single":
        simple["H_script_noark_single"] += count
    else:
        simple["Z_other"] += count

for group in sorted(simple.keys()):
    print(f"  {group}: {simple[group]}")
