#!/usr/bin/env python
"""Build a three-sheet workbook comparing person IDs between source workbooks."""

from __future__ import annotations

import argparse
import difflib
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def normalize_space(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_name_for_match(value: str) -> str:
    text = normalize_space(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().replace("’", "'")
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def load_full_people(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws_names = wb["Names"]
    headers = [cell.value for cell in next(ws_names.iter_rows(min_row=1, max_row=1))]
    idx = {header: pos for pos, header in enumerate(headers)}
    years_by_person: dict[int, set[str]] = {}
    constituencies_by_person: dict[int, set[str]] = {}
    parties_by_person: dict[int, set[str]] = {}
    ws_results = wb["ElectionResults"]
    result_headers = [cell.value for cell in next(ws_results.iter_rows(min_row=1, max_row=1))]
    result_idx = {header: pos for pos, header in enumerate(result_headers)}
    for row in ws_results.iter_rows(min_row=2, values_only=True):
        result_type = normalize_space(row[result_idx["ResultType"]])
        if (
            result_type != "Candidate"
            and not result_type.startswith("ListCandidate")
            and not result_type.startswith("RegionalListCandidate")
        ):
            continue
        person_id = row[result_idx["PersonID"]]
        date_value = row[result_idx["Date"]]
        if not person_id or not date_value:
            continue
        years_by_person.setdefault(person_id, set()).add(str(date_value)[:4])
        constituency = normalize_space(row[result_idx["Constituency"]])
        if constituency:
            constituencies_by_person.setdefault(person_id, set()).add(constituency)
        party_name = normalize_space(row[result_idx["Party Name"]])
        if party_name:
            parties_by_person.setdefault(person_id, set()).add(party_name)

    rows: list[dict[str, Any]] = []
    seen_ids: set[int] = set()
    for row in ws_names.iter_rows(min_row=2, values_only=True):
        person_id = row[idx["PersonID"]]
        full_name = normalize_space(row[idx["Full Name usually known by"]])
        if not person_id or not full_name:
            continue
        if person_id in seen_ids:
            continue
        seen_ids.add(person_id)
        rows.append(
            {
                "full_name": full_name,
                "first_name": normalize_space(row[idx["First Name"]]),
                "last_name": normalize_space(row[idx["Last Name"]]),
                "person_id": person_id,
                "gender": normalize_space(row[idx["Gender"]]),
                "years_stood": ", ".join(sorted(years_by_person.get(person_id, set()))),
                "constituencies_stood": ", ".join(sorted(constituencies_by_person.get(person_id, set()))),
                "parties_stood_in": ", ".join(sorted(parties_by_person.get(person_id, set()))),
                "match_key": normalize_name_for_match(full_name),
            }
        )
    return rows


def load_modern_people(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["ElectionResults"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {header: pos for pos, header in enumerate(headers)}
    people: dict[int, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["ResultType"]] != "Candidate":
            continue
        person_id = row[idx["PersonID"]]
        full_name = normalize_space(row[idx["Name usually known by"]])
        if not person_id or not full_name:
            continue
        entry = people.setdefault(
            person_id,
            {
                "full_name": full_name,
                "first_name": normalize_space(row[idx["First Name"]]),
                "last_name": normalize_space(row[idx["Last Name"]]),
                "person_id": person_id,
                "years": set(),
                "councils": set(),
                "parties": set(),
                "match_key": normalize_name_for_match(full_name),
            },
        )
        if row[idx["Date"]]:
            entry["years"].add(str(row[idx["Date"]])[:4])
        if row[idx["Council"]]:
            entry["councils"].add(normalize_space(row[idx["Council"]]))
        if row[idx["Party Name"]]:
            entry["parties"].add(normalize_space(row[idx["Party Name"]]))
    return [
        {
            **entry,
            "years": ", ".join(sorted(entry["years"])),
            "councils": ", ".join(sorted(entry["councils"])),
            "parties": ", ".join(sorted(entry["parties"])),
        }
        for entry in sorted(people.values(), key=lambda item: item["full_name"])
    ]


def attempt_matches(full_people: list[dict[str, Any]], modern_people: list[dict[str, Any]]) -> list[dict[str, Any]]:
    modern_by_key: dict[str, list[dict[str, Any]]] = {}
    for entry in modern_people:
        modern_by_key.setdefault(entry["match_key"], []).append(entry)

    matches: list[dict[str, Any]] = []
    used_modern_ids: set[int] = set()
    for full_entry in full_people:
        exact = modern_by_key.get(full_entry["match_key"], [])
        if len(exact) == 1:
            modern_entry = exact[0]
            used_modern_ids.add(modern_entry["person_id"])
            matches.append(
                {
                    "full_name": full_entry["full_name"],
                    "full_person_id": full_entry["person_id"],
                    "gender": full_entry["gender"],
                    "full_years_stood": full_entry["years_stood"],
                    "full_constituencies_stood": full_entry["constituencies_stood"],
                    "full_parties_stood_in": full_entry["parties_stood_in"],
                    "modern_name": modern_entry["full_name"],
                    "modern_person_id": modern_entry["person_id"],
                    "modern_years": modern_entry["years"],
                    "modern_councils": modern_entry["councils"],
                    "modern_parties_stood_in": modern_entry["parties"],
                    "match_method": "exact_normalized_name",
                    "match_score": 1.0,
                }
            )
            continue

        best_ratio = 0.0
        best_entry: dict[str, Any] | None = None
        full_last = normalize_name_for_match(full_entry["last_name"])
        for modern_entry in modern_people:
            if full_last and normalize_name_for_match(modern_entry["last_name"]) != full_last:
                continue
            ratio = difflib.SequenceMatcher(None, full_entry["match_key"], modern_entry["match_key"]).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_entry = modern_entry
        if best_entry and best_ratio >= 0.94:
            used_modern_ids.add(best_entry["person_id"])
            matches.append(
                {
                    "full_name": full_entry["full_name"],
                    "full_person_id": full_entry["person_id"],
                    "gender": full_entry["gender"],
                    "full_years_stood": full_entry["years_stood"],
                    "full_constituencies_stood": full_entry["constituencies_stood"],
                    "full_parties_stood_in": full_entry["parties_stood_in"],
                    "modern_name": best_entry["full_name"],
                    "modern_person_id": best_entry["person_id"],
                    "modern_years": best_entry["years"],
                    "modern_councils": best_entry["councils"],
                    "modern_parties_stood_in": best_entry["parties"],
                    "match_method": "fuzzy_same_last_name",
                    "match_score": round(best_ratio, 4),
                }
            )
        else:
            matches.append(
                {
                    "full_name": full_entry["full_name"],
                    "full_person_id": full_entry["person_id"],
                    "gender": full_entry["gender"],
                    "full_years_stood": full_entry["years_stood"],
                    "full_constituencies_stood": full_entry["constituencies_stood"],
                    "full_parties_stood_in": full_entry["parties_stood_in"],
                    "modern_name": None,
                    "modern_person_id": None,
                    "modern_years": None,
                    "modern_councils": None,
                    "modern_parties_stood_in": None,
                    "match_method": "unmatched",
                    "match_score": 0.0,
                }
            )

    for modern_entry in modern_people:
        if modern_entry["person_id"] in used_modern_ids:
            continue
        matches.append(
            {
                "full_name": None,
                "full_person_id": None,
                "gender": None,
                "full_years_stood": None,
                "full_constituencies_stood": None,
                "full_parties_stood_in": None,
                "modern_name": modern_entry["full_name"],
                "modern_person_id": modern_entry["person_id"],
                "modern_years": modern_entry["years"],
                "modern_councils": modern_entry["councils"],
                "modern_parties_stood_in": modern_entry["parties"],
                "match_method": "modern_only",
                "match_score": 0.0,
            }
        )
    return matches


def write_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--full-workbook", default="Full election tables.xlsx")
    parser.add_argument("--modern-workbook", default=r"_tmp_xls2rar_extract\out\wiki_lgov_modern\lgov-modern-wikipedia.stvfix.xlsx")
    parser.add_argument("--output", default=r"_tmp_xls2rar_extract\out\wiki_lgov_modern\person-id-match.xlsx")
    args = parser.parse_args()

    full_people = load_full_people(Path(args.full_workbook))
    modern_people = load_modern_people(Path(args.modern_workbook))
    matches = attempt_matches(full_people, modern_people)

    wb = Workbook()
    ws_full = wb.active
    ws_full.title = "FullWorkbookPeople"
    write_sheet(
        ws_full,
        ["full_name", "first_name", "last_name", "person_id", "gender", "years_stood", "constituencies_stood", "parties_stood_in"],
        full_people,
    )
    ws_modern = wb.create_sheet("ModernLocalPeople")
    write_sheet(
        ws_modern,
        ["full_name", "first_name", "last_name", "person_id", "years", "councils", "parties"],
        modern_people,
    )
    ws_matches = wb.create_sheet("AttemptedMatches")
    write_sheet(
        ws_matches,
        [
            "full_name",
            "full_person_id",
            "gender",
            "full_years_stood",
            "full_constituencies_stood",
            "full_parties_stood_in",
            "modern_name",
            "modern_person_id",
            "modern_years",
            "modern_councils",
            "modern_parties_stood_in",
            "match_method",
            "match_score",
        ],
        matches,
    )
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    main()
