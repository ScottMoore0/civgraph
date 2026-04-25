#!/usr/bin/env python
"""Build a 3-sheet Excel listing the unique party labels found in:
  Sheet 1: Local elections 2014 onwards (LGD-2014 era)
  Sheet 2: Local elections before 2014 (1973-2011)
  Sheet 3: Non-local elections (NI Assembly, Westminster, European, etc.)
Each sheet shows distinct party labels with how many times each appears
and which years it shows up in.
"""
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

REPO = Path(__file__).resolve().parent.parent
ELECT_BASE = REPO / "election-viewer-package" / "data" / "elections"
OUT_PATH = REPO / "election-viewer-package" / "data" / "party-names-by-election-type.xlsx"

POST2014_DATES = {"2014-05-22", "2018-10-18", "2019-05-02", "2023-05-18"}

def collect_parties(election_jsons):
    """Walk the iterable of (label, json_path) and build party_label -> {dates, counts}.
    Schemas differ across years:
      - Pre-2014 ARK conversions:   countGroup[].Party
      - Post-2014 LG / non-local:    countGroup[].Party_Name (canonical raw label)
    Try both."""
    out = defaultdict(lambda: {"count": 0, "dates": set()})
    for label, jpath in election_jsons:
        try:
            data = json.loads(jpath.read_text(encoding="utf-8"))
            cg = data.get("Constituency", {}).get("countGroup", [])
            for c in cg:
                p = (c.get("Party") or c.get("Party_Name") or "").strip()
                if not p: continue
                out[p]["count"] += 1
                out[p]["dates"].add(label)
        except Exception:
            continue
    return out

def write_sheet(ws, title, parties):
    ws.title = title
    ws.append(["Party label (as recorded in source data)", "Total candidate-occurrences",
               "Number of elections featured in", "Election dates featured"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEEEEE")
    for label in sorted(parties.keys(), key=lambda k: (-parties[k]["count"], k.lower())):
        info = parties[label]
        ws.append([label, info["count"], len(info["dates"]),
                   ", ".join(sorted(info["dates"]))])
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 60

def main():
    # Scan local-government dir
    lg_dir = ELECT_BASE / "local-government"
    pre, post = [], []
    for date_dir in sorted(lg_dir.iterdir()):
        if not date_dir.is_dir(): continue
        date = date_dir.name
        bucket = post if date in POST2014_DATES else pre
        for jp in sorted(date_dir.glob("*.json")):
            if jp.stem.startswith("_"): continue
            bucket.append((date, jp))

    # Scan non-local-government bodies
    nonlocal_files = []
    for body_dir in sorted(ELECT_BASE.iterdir()):
        if not body_dir.is_dir(): continue
        if body_dir.name == "local-government": continue
        for date_dir in sorted(body_dir.iterdir()):
            if not date_dir.is_dir(): continue
            for jp in sorted(date_dir.glob("*.json")):
                if jp.stem.startswith("_"): continue
                nonlocal_files.append((date_dir.name, jp))

    print(f"Scanning {len(post)} post-2014 LG files, {len(pre)} pre-2014 LG files, {len(nonlocal_files)} non-local files")

    parties_post  = collect_parties(post)
    parties_pre   = collect_parties(pre)
    parties_other = collect_parties(nonlocal_files)

    print(f"  post-2014 LG distinct labels: {len(parties_post)}")
    print(f"  pre-2014 LG  distinct labels: {len(parties_pre)}")
    print(f"  non-local    distinct labels: {len(parties_other)}")

    wb = Workbook()
    write_sheet(wb.active,           "LG 2014+",  parties_post)
    write_sheet(wb.create_sheet(),   "LG pre-2014", parties_pre)
    write_sheet(wb.create_sheet(),   "Non-LG",     parties_other)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")

if __name__ == "__main__":
    main()
