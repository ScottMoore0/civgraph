#!/usr/bin/env python
"""Deep analysis for v5 PersonID improvements."""

import openpyxl, re, unicodedata, json
from collections import defaultdict

def norm(name):
    if not name: return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()

def split_fn_sn(name):
    parts = name.strip().split()
    if not parts: return ("", "")
    if len(parts) == 1: return ("", norm(parts[0]))
    return (norm(" ".join(parts[:-1])), norm(parts[-1]))

VGROUPS = [
    ["james","jim","jimmy","seamus"],["william","will","willie","bill","billy","liam"],
    ["robert","rob","robbie","bob","bobby"],["john","johnny","jack","sean","shane"],
    ["thomas","tom","tommy"],["edward","ed","eddie","ted","eamonn","eamon","edmund"],
    ["richard","dick","dickie","ricky"],["charles","charlie","cathal"],
    ["patrick","paddy","pat","patsy","padraig"],["michael","mike","mick","mickey"],
    ["christopher","chris"],["daniel","dan","danny"],["andrew","andy","drew"],
    ["anthony","tony"],["joseph","joe","joey"],["samuel","sam","sammy"],
    ["alexander","alex","alec","alistair","alastair"],["frederick","fred","freddie"],
    ["kenneth","ken","kenny"],["ronald","ron","ronnie"],["raymond","ray"],
    ["stephen","steve","steven"],["philip","phil"],["lawrence","larry","laurence"],
    ["gerald","gerry","gerard","gearoid"],["peter","pete"],["david","dave","davy"],
    ["henry","harry","hal"],["albert","bert","bertie"],["alfred","alf","alfie"],
    ["francis","frank","frankie","proinsias"],["desmond","des"],["stanley","stan"],
    ["ernest","ernie"],["norman","norm"],["donald","don","donnie"],["dennis","denis"],
    ["terence","terry","terrence"],["vincent","vince"],["douglas","doug"],
    ["timothy","tim"],["nicholas","nick"],["geoffrey","geoff","jeff","jeffrey"],
    ["elizabeth","liz","lizzie","beth","betty","bess"],
    ["margaret","maggie","meg","peggy","madge"],
    ["catherine","kate","katie","cathy","kathleen"],
    ["patricia","pat","tricia","trish"],["dorothy","dot","dolly"],
    ["jacqueline","jackie"],["anne","ann","annie"],
    ["mary","molly","may","maire"],["bridget","brid","bridie"],
    ["niall","neal","neil"],["ciaran","kieran"],
    ["archibald","archie"],["reginald","reg","reggie"],
    ["leonard","len","lenny"],["bernard","bernie"],
    ["pamela","pam"],["susan","sue","susie"],["deborah","deb","debbie"],
    ["christine","chris","christina","tina"],["rosemary","rosie"],["caroline","carol"],
    ["joanna","jo","joanne","joan"],["jennifer","jenny","jen"],
]
_VM = {}
for gi, grp in enumerate(VGROUPS):
    for n in grp:
        _VM[n.lower()] = gi

def are_fn_var(f1, f2):
    n1, n2 = f1.lower().strip(), f2.lower().strip()
    if n1 == n2: return True
    g1, g2 = _VM.get(n1), _VM.get(n2)
    if g1 is not None and g2 is not None and g1 == g2: return True
    if len(n1) >= 3 and len(n2) >= 3 and (n1.startswith(n2) or n2.startswith(n1)): return True
    return False

def normalize_party(p):
    if not p: return ""
    l = p.lower()
    for needle, code in [("sdlp","sdlp"),("democratic unionist","dup"),("dup","dup"),
        ("ulster unionist","uup"),("uup","uup"),("sinn","sf"),("alliance","all"),
        ("independent","ind"),("green","grn"),("tuv","tuv"),("workers","wp"),
        ("pup","pup"),("labour","lab"),("conservative","con"),("ukip","ukip"),("ukup","ukup")]:
        if needle in l: return code
    return l[:15]

wb = openpyxl.load_workbook("Full election tables - comprehensive - personid-v5.xlsx", read_only=True)
ws = wb["ElectionResults"]
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col = {name: i for i, name in enumerate(headers)}

profiles = defaultdict(lambda: {"names": set(), "parties": set(), "dates": set(),
    "bodies": set(), "consts": set(), "rows": 0})
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[col["ResultType"]] != "Candidate": continue
    pid = str(row[col["PersonID"]])
    p = profiles[pid]
    p["names"].add(row[col["Name usually known by"]] or "")
    p["parties"].add(row[col["Party Name"]] or "")
    p["dates"].add(str(row[col["Date"]] or ""))
    p["bodies"].add(row[col["ElectedBody"]] or "")
    p["consts"].add(row[col["Constituency"]] or "")
    p["rows"] += 1

# === 1. Variant merges: same surname, variant first name, different PID ===
sn_index = defaultdict(list)
for pid, p in profiles.items():
    for name in p["names"]:
        fn, sn = split_fn_sn(name)
        if sn and fn:
            sn_index[sn].append((pid, fn, name))

variant_candidates = []
seen_pairs = set()
for sn, entries in sn_index.items():
    pids_set = set(e[0] for e in entries)
    if len(pids_set) <= 1: continue
    for i in range(len(entries)):
        for j in range(i + 1, len(entries)):
            p1, fn1, n1 = entries[i]
            p2, fn2, n2 = entries[j]
            if p1 == p2: continue
            pair = tuple(sorted([p1, p2]))
            if pair in seen_pairs: continue
            if are_fn_var(fn1, fn2):
                seen_pairs.add(pair)
                prof1, prof2 = profiles[p1], profiles[p2]
                # Score: party overlap + temporal proximity + geographic overlap
                np1 = {normalize_party(x) for x in prof1["parties"]}
                np2 = {normalize_party(x) for x in prof2["parties"]}
                party_match = bool(np1 & np2 - {""})
                years1 = [int(d[:4]) for d in prof1["dates"] if d[:4].isdigit()]
                years2 = [int(d[:4]) for d in prof2["dates"] if d[:4].isdigit()]
                gap = min(abs(y1 - y2) for y1 in years1 for y2 in years2) if years1 and years2 else 999
                geo_match = bool(prof1["consts"] & prof2["consts"])
                variant_candidates.append({
                    "p1": p1, "p2": p2, "n1": n1, "n2": n2,
                    "party_match": party_match, "gap": gap, "geo_match": geo_match,
                    "parties1": sorted(prof1["parties"]),
                    "parties2": sorted(prof2["parties"]),
                    "dates1": sorted(prof1["dates"])[:4],
                    "dates2": sorted(prof2["dates"])[:4],
                    "consts1": sorted(prof1["consts"])[:3],
                    "consts2": sorted(prof2["consts"])[:3],
                })

# Classify
safe = [v for v in variant_candidates if v["party_match"] and v["gap"] <= 15]
possible = [v for v in variant_candidates if v["party_match"] and v["gap"] > 15 and v["gap"] <= 40]
risky = [v for v in variant_candidates if not v["party_match"] or v["gap"] > 40]

print(f"=== VARIANT MERGE CANDIDATES ===")
print(f"Total: {len(variant_candidates)}")
print(f"Safe (party match + <=15yr gap): {len(safe)}")
print(f"Possible (party match + 15-40yr gap): {len(possible)}")
print(f"Risky (no party match or >40yr gap): {len(risky)}")
print()
print("SAFE merges:")
for v in safe:
    print(f"  {v['n1']} (PID {v['p1']}) <-> {v['n2']} (PID {v['p2']})")
    print(f"    parties: {v['parties1']} vs {v['parties2']}, gap={v['gap']}yr, geo={v['geo_match']}")
print()
print("POSSIBLE merges:")
for v in possible:
    print(f"  {v['n1']} (PID {v['p1']}) <-> {v['n2']} (PID {v['p2']})")
    print(f"    parties: {v['parties1']} vs {v['parties2']}, gap={v['gap']}yr")
print()
print("RISKY (do NOT merge):")
for v in risky[:20]:
    print(f"  {v['n1']} (PID {v['p1']}) <-> {v['n2']} (PID {v['p2']})")
    print(f"    parties: {v['parties1']} vs {v['parties2']}, gap={v['gap']}yr")

# Save analysis
json.dump({"safe": safe, "possible": possible, "risky": risky},
          open("_tmp_v5_analysis.json", "w", encoding="utf-8"),
          indent=2, default=str, ensure_ascii=False)
