#!/usr/bin/env python
"""Unify PersonIDs in the comprehensive election workbook.

Operates on a copy of the workbook.  Phases:

  Phase 1a — Remove garbage rows (Wikipedia parse artifacts)
  Phase 1b — Strip T-prefix from temp IDs to unify with hash IDs
  Phase 1c — Merge temp/hash IDs into curated IDs by exact normalised name
  Phase 2  — Cross-election fuzzy matching for remaining temp/hash IDs
  Phase 3  — Assign fresh sequential IDs to unmatched candidates

Produces a change log so the user can review before approving.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl


# ── Name normalisation ────────────────────────────────────────────────────

def normalize_name(name: str) -> str:
    """Normalise a name for matching: lowercase, strip diacritics, collapse whitespace."""
    if not name:
        return ""
    # NFKD decomposition then strip combining marks
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    # Lowercase, strip punctuation except hyphens and apostrophes
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    lowered = re.sub(r"\s+", " ", lowered).strip()
    return lowered


def normalize_party(party: str) -> str:
    if not party:
        return ""
    lowered = party.strip().lower()
    mappings = [
        ("social democratic and labour party", "sdlp"), ("sdlp", "sdlp"),
        ("democratic unionist", "dup"), ("dup", "dup"),
        ("ulster unionist", "uup"), ("uup", "uup"),
        ("sinn f", "sf"), ("alliance", "alliance"),
        ("independent", "ind"), ("green", "green"),
        ("traditional unionist voice", "tuv"), ("tuv", "tuv"),
        ("workers party", "wp"), ("republican clubs", "wp"),
        ("progressive unionist", "pup"), ("pup", "pup"),
        ("labour", "lab"), ("conservative", "con"),
        ("nationalist party", "nat"), ("ukip", "ukip"),
        ("ni labour", "nilab"), ("northern ireland labour", "nilab"),
    ]
    for needle, code in mappings:
        if needle in lowered:
            return code
    return lowered[:20]


def is_garbage_name(name: str) -> bool:
    """Detect Wikipedia parsing artifacts that leaked into the Name field."""
    if not name or not name.strip():
        return True
    cleaned = name.strip()
    # Known garbage patterns from wiki link residue
    garbage = {
        "party", "(politician)", "ireland", "voice", "fein",
        "fáin", "ecology", "clubs", "labour", "conservative",
    }
    return cleaned.lower().strip("() ") in garbage or len(cleaned) <= 2


def pid_type(pid: str) -> str:
    """Classify a PersonID string."""
    if not pid:
        return "none"
    if pid.startswith("T"):
        return "temp"
    if pid.isdigit():
        return "curated" if len(pid) <= 6 else "hash"
    return "other"


# ── Matching logic ────────────────────────────────────────────────────────

def score_match(
    source_parties: set[str],
    source_dates: list[str],
    source_constituencies: set[str],
    target_parties: set[str],
    target_dates: list[str],
    target_constituencies: set[str],
) -> int:
    """Score a candidate match (0-100)."""
    score = 10  # base name match score

    # Party (0-40)
    src_norm = {normalize_party(p) for p in source_parties if p}
    tgt_norm = {normalize_party(p) for p in target_parties if p}
    if src_norm & tgt_norm:
        score += 40
    elif "ind" in src_norm or "ind" in tgt_norm:
        score += 15
    # Known party switches
    elif any(
        (a in src_norm and b in tgt_norm) or (b in src_norm and a in tgt_norm)
        for a, b in [("uup", "dup"), ("uup", "ind"), ("dup", "tuv"), ("wp", "sf")]
    ):
        score += 20

    # Temporal proximity (0-20)
    all_years = []
    for d in source_dates + target_dates:
        try:
            all_years.append(int(str(d)[:4]))
        except (ValueError, TypeError):
            pass
    if len(all_years) >= 2:
        gap = max(all_years) - min(all_years)
        if gap <= 8:
            score += 20
        elif gap <= 15:
            score += 15
        elif gap <= 25:
            score += 10
        elif gap <= 40:
            score += 5

    # Geographic proximity (0-30)
    src_const = {c.lower() for c in source_constituencies if c}
    tgt_const = {c.lower() for c in target_constituencies if c}
    if src_const & tgt_const:
        score += 30
    elif any(s in t or t in s for s in src_const for t in tgt_const):
        score += 15

    return score


# ── Main processing ───────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Unify PersonIDs in election workbook")
    parser.add_argument("input", help="Input XLSX path")
    parser.add_argument("--output", help="Output XLSX path (default: overwrite input)")
    parser.add_argument("--log", default="personid_unification_log.json", help="Change log output")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else input_path
    log_path = Path(args.log)

    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb["ElectionResults"]

    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    # Read all data rows
    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))

    pid_col = col["PersonID"]
    rt_col = col["ResultType"]
    name_col = col["Name usually known by"]
    party_col = col["Party Name"]
    date_col = col["Date"]
    body_col = col["ElectedBody"]
    const_col = col["Constituency"]
    src_name_col = col["Source Name"]

    log: dict[str, Any] = {"phases": {}, "changes": [], "summary": {}}

    # ── Phase 1a: Remove garbage rows ─────────────────────────────────────
    print("\nPhase 1a: Identifying garbage rows...")
    garbage_indices: list[int] = []
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        name = row[name_col]
        if is_garbage_name(name):
            garbage_indices.append(i)

    # Mark garbage rows (set ResultType to "GarbageRemoved" rather than deleting,
    # so the user can review)
    for i in garbage_indices:
        old_name = rows[i][name_col]
        old_pid = rows[i][pid_col]
        rows[i][rt_col] = "GarbageRemoved"
        log["changes"].append({
            "phase": "1a",
            "row": i + 2,
            "action": "garbage_removed",
            "old_name": str(old_name),
            "old_pid": str(old_pid),
        })

    log["phases"]["1a"] = {"garbage_rows_removed": len(garbage_indices)}
    print(f"  Found {len(garbage_indices)} garbage rows (marked as GarbageRemoved)")

    # ── Phase 1b: Strip T-prefix ──────────────────────────────────────────
    print("\nPhase 1b: Stripping T-prefix from temp IDs...")
    t_stripped = 0
    for i, row in enumerate(rows):
        if row[rt_col] not in ("Candidate", "NonTransferable"):
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid.startswith("T"):
            new_pid = pid[1:]
            rows[i][pid_col] = new_pid
            t_stripped += 1

    log["phases"]["1b"] = {"t_prefixes_stripped": t_stripped}
    print(f"  Stripped T-prefix from {t_stripped} rows")

    # ── Phase 1c: Merge hash/temp IDs into curated IDs by exact name ─────
    print("\nPhase 1c: Merging temp/hash IDs into curated IDs by exact name...")

    # Build name→curated PID mapping
    name_to_curated: dict[str, str] = {}
    name_to_curated_details: dict[str, dict] = {}
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) != "curated":
            continue
        name = row[name_col]
        if not name:
            continue
        norm = normalize_name(name)
        if norm and norm not in name_to_curated:
            name_to_curated[norm] = pid
            name_to_curated_details[norm] = {
                "display_name": name,
                "curated_pid": pid,
            }

    # Apply curated PIDs to hash/temp rows with matching names
    merged_count = 0
    merged_names: dict[str, dict] = {}
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        pt = pid_type(pid)
        if pt == "curated":
            continue
        name = row[name_col]
        if not name:
            continue
        norm = normalize_name(name)
        if norm in name_to_curated:
            old_pid = pid
            new_pid = name_to_curated[norm]
            rows[i][pid_col] = new_pid
            merged_count += 1
            if norm not in merged_names:
                merged_names[norm] = {
                    "display_name": name,
                    "curated_pid": new_pid,
                    "old_pids": set(),
                    "count": 0,
                }
            merged_names[norm]["old_pids"].add(old_pid)
            merged_names[norm]["count"] += 1

    # Log merge details
    merge_log = []
    for norm, info in sorted(merged_names.items()):
        merge_log.append({
            "name": info["display_name"],
            "curated_pid": info["curated_pid"],
            "old_pids": sorted(info["old_pids"]),
            "rows_updated": info["count"],
        })

    log["phases"]["1c"] = {
        "names_merged": len(merged_names),
        "rows_updated": merged_count,
        "curated_names_available": len(name_to_curated),
        "merges": merge_log,
    }
    print(f"  Merged {merged_count} rows across {len(merged_names)} names into curated IDs")

    # ── Phase 2: Cross-election fuzzy matching ────────────────────────────
    print("\nPhase 2: Cross-election fuzzy matching for remaining hash IDs...")

    # Build profiles for curated-PID people
    curated_profiles: dict[str, dict] = {}
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) != "curated":
            continue
        if pid not in curated_profiles:
            curated_profiles[pid] = {
                "names": set(),
                "norm_names": set(),
                "parties": set(),
                "dates": [],
                "constituencies": set(),
            }
        name = row[name_col] or ""
        curated_profiles[pid]["names"].add(name)
        curated_profiles[pid]["norm_names"].add(normalize_name(name))
        curated_profiles[pid]["parties"].add(row[party_col] or "")
        curated_profiles[pid]["dates"].append(row[date_col])
        curated_profiles[pid]["constituencies"].add(row[const_col] or "")

    # Build profiles for remaining hash-PID candidates
    hash_profiles: dict[str, dict] = {}
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) != "hash":
            continue
        if pid not in hash_profiles:
            hash_profiles[pid] = {
                "names": set(),
                "norm_names": set(),
                "parties": set(),
                "dates": [],
                "constituencies": set(),
                "row_indices": [],
            }
        name = row[name_col] or ""
        hash_profiles[pid]["names"].add(name)
        hash_profiles[pid]["norm_names"].add(normalize_name(name))
        hash_profiles[pid]["parties"].add(row[party_col] or "")
        hash_profiles[pid]["dates"].append(row[date_col])
        hash_profiles[pid]["constituencies"].add(row[const_col] or "")
        hash_profiles[pid]["row_indices"].append(i)

    # Build surname→curated PID index for fuzzy matching
    surname_index: dict[str, list[str]] = defaultdict(list)
    for pid, prof in curated_profiles.items():
        for name in prof["names"]:
            parts = name.strip().split()
            if parts:
                surname = normalize_name(parts[-1])
                if surname:
                    surname_index[surname].append(pid)

    # Try to match each hash profile to a curated profile
    fuzzy_matched = 0
    fuzzy_ambiguous = 0
    fuzzy_log: list[dict] = []
    ambiguous_log: list[dict] = []

    for hash_pid, hprof in hash_profiles.items():
        # Get surname for index lookup
        surname_candidates: set[str] = set()
        for name in hprof["names"]:
            parts = name.strip().split()
            if parts:
                surname = normalize_name(parts[-1])
                if surname and surname in surname_index:
                    surname_candidates.update(surname_index[surname])

        if not surname_candidates:
            continue

        # Score each candidate
        scored: list[tuple[int, str]] = []
        for curated_pid in surname_candidates:
            cprof = curated_profiles[curated_pid]
            # Check normalised name similarity (must share at least surname)
            h_norms = hprof["norm_names"]
            c_norms = cprof["norm_names"]

            # Exact normalised name match
            if h_norms & c_norms:
                # Already handled in Phase 1c — this shouldn't happen
                continue

            # Same surname, check full match score
            s = score_match(
                hprof["parties"], hprof["dates"], hprof["constituencies"],
                cprof["parties"], cprof["dates"], cprof["constituencies"],
            )
            if s >= 50:
                scored.append((s, curated_pid))

        if not scored:
            continue

        scored.sort(key=lambda x: -x[0])
        best_score, best_pid = scored[0]
        second_score = scored[1][0] if len(scored) > 1 else 0
        margin = best_score - second_score

        entry = {
            "hash_pid": hash_pid,
            "name": sorted(hprof["names"])[0],
            "best_match_pid": best_pid,
            "best_match_name": sorted(curated_profiles[best_pid]["names"])[0],
            "best_score": best_score,
            "second_score": second_score,
            "margin": margin,
            "hash_parties": sorted(hprof["parties"]),
            "match_parties": sorted(curated_profiles[best_pid]["parties"]),
            "hash_dates": sorted(set(str(d) for d in hprof["dates"])),
            "match_dates": sorted(set(str(d) for d in curated_profiles[best_pid]["dates"])),
        }

        if best_score >= 70 and (len(scored) == 1 or margin >= 20):
            # Auto-match
            for idx in hprof["row_indices"]:
                rows[idx][pid_col] = best_pid
            fuzzy_matched += 1
            entry["action"] = "auto_matched"
            fuzzy_log.append(entry)
        else:
            fuzzy_ambiguous += 1
            entry["action"] = "ambiguous"
            ambiguous_log.append(entry)

    log["phases"]["2"] = {
        "auto_matched": fuzzy_matched,
        "ambiguous": fuzzy_ambiguous,
        "hash_profiles_checked": len(hash_profiles),
        "matches": fuzzy_log,
        "ambiguous_cases": ambiguous_log,
    }
    print(f"  Auto-matched {fuzzy_matched} hash PIDs to curated IDs")
    print(f"  Found {fuzzy_ambiguous} ambiguous cases (logged for review)")

    # ── Phase 2b: Unify hash IDs that share a normalised name ─────────────
    print("\nPhase 2b: Unifying hash IDs that share a normalised name...")

    # Group remaining hash PIDs by normalised name
    norm_to_hash_pids: dict[str, set[str]] = defaultdict(set)
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) != "hash":
            continue
        name = row[name_col] or ""
        norm = normalize_name(name)
        if norm:
            norm_to_hash_pids[norm].add(pid)

    # For groups with multiple hash PIDs, pick one canonical PID
    unified_count = 0
    unify_map: dict[str, str] = {}
    for norm, pids in norm_to_hash_pids.items():
        if len(pids) <= 1:
            continue
        canonical = sorted(pids)[0]
        for pid in pids:
            if pid != canonical:
                unify_map[pid] = canonical

    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid in unify_map:
            rows[i][pid_col] = unify_map[pid]
            unified_count += 1

    log["phases"]["2b"] = {
        "hash_pids_unified": len(unify_map),
        "rows_updated": unified_count,
    }
    print(f"  Unified {len(unify_map)} duplicate hash PIDs ({unified_count} rows)")

    # ── Phase 3: Assign fresh sequential IDs ──────────────────────────────
    print("\nPhase 3: Assigning fresh sequential IDs to remaining hash PIDs...")

    # Find max existing curated ID
    max_curated = 0
    for row in rows:
        if row[rt_col] not in ("Candidate", "NonTransferable"):
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid.isdigit() and len(pid) <= 6:
            max_curated = max(max_curated, int(pid))

    # Also check the known nextId from the existing system
    next_id = max(max_curated + 1, 100020)

    # Assign fresh IDs to remaining hash PIDs
    hash_to_fresh: dict[str, int] = {}
    fresh_log: list[dict] = []
    for i, row in enumerate(rows):
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if pid_type(pid) != "hash":
            continue
        if pid not in hash_to_fresh:
            hash_to_fresh[pid] = next_id
            name = row[name_col] or ""
            fresh_log.append({
                "old_hash_pid": pid,
                "new_pid": next_id,
                "name": name,
            })
            next_id += 1
        rows[i][pid_col] = str(hash_to_fresh[pid])

    log["phases"]["3"] = {
        "fresh_ids_assigned": len(hash_to_fresh),
        "id_range": f"{min(hash_to_fresh.values()) if hash_to_fresh else 0}-{next_id - 1}",
        "assignments": fresh_log,
    }
    print(f"  Assigned {len(hash_to_fresh)} fresh IDs (range {min(hash_to_fresh.values()) if hash_to_fresh else 'N/A'}-{next_id - 1})")

    # ── Update Transfers sheet ────────────────────────────────────────────
    print("\nUpdating Transfers sheet...")
    ws_t = wb["Transfers"]
    t_headers = [cell.value for cell in ws_t[1]]
    t_pid_col = t_headers.index("PersonID") if "PersonID" in t_headers else None
    t_src_col = t_headers.index("SourcePersonID") if "SourcePersonID" in t_headers else None

    # Build complete PID remap (all old→new mappings from all phases)
    pid_remap: dict[str, str] = {}
    # Phase 1b: T-prefix stripping
    # (already done inline, but we need to track for Transfers)
    # Phase 1c merges + Phase 2 fuzzy matches are in the rows already
    # Let's just rebuild the remap from original→current by re-reading
    # Actually, let's build it from all the changes we made

    # Simpler approach: build a mapping from the current state of ElectionResults
    # Map (name, date, body, constituency) → current PID
    # Then apply to Transfers by matching

    # For Transfers, the PersonID field uses the same IDs as ElectionResults
    # Build old PID → new PID mapping
    # We need to track all remaps across phases

    # Collect all PID changes: for Transfers, remap PersonID and SourcePersonID
    # The simplest approach: build name→currentPID from the final ElectionResults state
    name_to_final_pid: dict[str, str] = {}
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        name = row[name_col]
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        if name and pid:
            norm = normalize_name(name)
            if norm:
                name_to_final_pid[norm] = pid

    t_updates = 0
    if t_pid_col is not None:
        t_name_col_idx = t_headers.index("Name") if "Name" in t_headers else None
        for t_row in ws_t.iter_rows(min_row=2):
            if t_name_col_idx is not None:
                name_val = t_row[t_name_col_idx].value
                if name_val:
                    norm = normalize_name(name_val)
                    if norm in name_to_final_pid:
                        old_val = t_row[t_pid_col].value
                        new_val = name_to_final_pid[norm]
                        if str(old_val) != new_val:
                            t_row[t_pid_col].value = new_val
                            t_updates += 1

    print(f"  Updated {t_updates} PersonID values in Transfers sheet")

    # ── Write results back ────────────────────────────────────────────────
    print(f"\nWriting {output_path}...")

    # Clear and rewrite ElectionResults
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=i + 2, column=j + 1, value=val)

    wb.save(output_path)

    # ── Final summary ─────────────────────────────────────────────────────
    # Re-count PID types after all changes
    final_counts = defaultdict(int)
    final_unique_pids = set()
    final_unique_names = set()
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        pid = str(row[pid_col]) if row[pid_col] is not None else ""
        pt = pid_type(pid)
        final_counts[pt] += 1
        final_unique_pids.add(pid)
        name = row[name_col]
        if name:
            final_unique_names.add(name)

    log["summary"] = {
        "total_candidate_rows": sum(final_counts.values()),
        "pid_types": dict(final_counts),
        "unique_pids": len(final_unique_pids),
        "unique_names": len(final_unique_names),
        "output_file": str(output_path),
    }

    print(f"\n{'='*60}")
    print(f"  FINAL SUMMARY")
    print(f"{'='*60}")
    print(f"  Total candidate rows: {sum(final_counts.values())}")
    print(f"  PID types: {dict(final_counts)}")
    print(f"  Unique PersonIDs: {len(final_unique_pids)}")
    print(f"  Unique names: {len(final_unique_names)}")
    print(f"  Output: {output_path}")

    # Write log
    log_path.write_text(json.dumps(log, indent=2, default=str), encoding="utf-8")
    print(f"  Change log: {log_path}")


if __name__ == "__main__":
    main()
