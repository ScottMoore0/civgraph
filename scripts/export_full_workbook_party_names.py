#!/usr/bin/env python
"""Export unique party names from Full election tables.xlsx."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def normalize(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def collect_unique_party_names(workbook_path: Path) -> list[str]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["ElectionResults"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {str(value): index for index, value in enumerate(header_row) if value is not None}
    canonical_idx = header_map["Party Name"]

    names: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        value = normalize(row[canonical_idx] if canonical_idx < len(row) else "")
        if value:
            names.add(value)
    wb.close()
    return sorted(names, key=str.casefold)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default="Full election tables.xlsx")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    names = collect_unique_party_names(workbook_path)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["party_name"])
        for name in names:
            writer.writerow([name])

    print(f"Wrote {len(names)} rows to {output_path}")


if __name__ == "__main__":
    main()
