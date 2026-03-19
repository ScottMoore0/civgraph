#!/usr/bin/env python
"""Deep analysis for v7 — find every remaining PersonID issue."""

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl


def norm(name):
    if not name: return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def split_fn_sn(name):
    parts = name.strip().split()
    if not parts: return ("", "")
    if len(parts) == 1: return ("", norm(parts[0]))
    return (norm(" ".join(parts[:-1])), norm(parts[-1]))


def normalize_party(p):
    if not p: return ""
    l = p.lower()
    for needle, code in [("sdlp","sdlp"),("democratic unionist","dup"),("dup","dup"),
        ("ulster unionist","uup"),("uup","uup"),("sinn","sf"),("alliance","all"),
        ("independent","ind"),("green","grn"),("tuv","tuv"),("workers","wp"),
        ("pup","pup"),("labour","lab"),("conservative","con"),("ukip","ukip"),("ukup","ukup"),
        ("vanguard","vupp"),("republican clubs","wp")]:
        if needle in l: return code
    return l[:15]


# Load ARK lookup
ark_lookup = json.load(open("_tmp_ark_name_lookup.json", encoding="utf-8"))
ark_candidates = json.load(open("_tmp_ark_candidates.json", encoding="utf-8"))

# Build ARK context lookup: (norm_short_name, norm_constituency, year) -> norm_full_name
ark_ctx = {}
for c in ark_candidates:
    full = c["full_name"]
    parts = full.split()
    if len(parts) < 2: continue
    short = f"{parts[0]} {parts[-1]}"
    const = norm(c["constituency"])
    year = c["year"]
    key = (norm(short), const, year)
    ark_ctx[key] = norm(full)


def get_ark_full(name, constituency="", year=""):
    """Get ARK full name for a candidate in context."""
    n = norm(name)
    const_n = norm(constituency)
    # Try with context first
    full = ark_ctx.get((n, const_n, year))
    if full:
        return full
    # Fallback: any full name for this short name
    fulls = ark_lookup.get(n, [])
    if len(fulls) == 1:
        return norm(fulls[0])
    return None


print("Loading workbook...")
wb = openpyxl.load_workbook("Full election tables - comprehensive - personid-v7.xlsx", read_only=True)
ws = wb["ElectionResults"]
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col = {name: i for i, name in enumerate(headers)}

# Build complete record database
records = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    if row[col["ResultType"]] != "Candidate":
        continue
    date = str(row[col["Date"]] or "")
    year = date[:4] if date[:4].isdigit() else ""
    records.append({
        "idx": i,
        "pid": str(row[col["PersonID"]]),
        "name": row[col["Name usually known by"]] or "",
        "party": row[col["Party Name"]] or "",
        "date": date,
        "year": year,
        "body": row[col["ElectedBody"]] or "",
        "const": row[col["Constituency"]] or "",
        "council": row[col["Council"]] or "",
        "ark_full": get_ark_full(
            row[col["Name usually known by"]] or "",
            row[col["Constituency"]] or "",
            year,
        ),
    })

# Build PID profiles
profiles = defaultdict(lambda: {
    "names": set(), "parties": set(), "dates": set(), "years": set(),
    "bodies": set(), "consts": set(), "ark_fulls": set(), "rows": 0,
})
for r in records:
    p = profiles[r["pid"]]
    p["names"].add(r["name"])
    p["parties"].add(r["party"])
    p["dates"].add(r["date"])
    if r["year"]: p["years"].add(int(r["year"]))
    p["bodies"].add(r["body"])
    p["consts"].add(r["const"])
    if r["ark_full"]: p["ark_fulls"].add(r["ark_full"])
    p["rows"] += 1


# ══════════════════════════════════════════════════════════════════════════
# ANALYSIS 1: ARK-based splits — PIDs where ARK shows different full names
# that indicate genuinely different people
# ══════════════════════════════════════════════════════════════════════════
print("\n=== ANALYSIS 1: ARK-based same-name-different-person detection ===")

ark_splits = []
for pid, prof in profiles.items():
    if len(prof["ark_fulls"]) <= 1:
        continue
    # Multiple distinct ARK full names — check if they indicate different people
    fulls = sorted(prof["ark_fulls"])
    for i in range(len(fulls)):
        for j in range(i + 1, len(fulls)):
            f1_parts = fulls[i].split()
            f2_parts = fulls[j].split()
            if len(f1_parts) < 3 or len(f2_parts) < 3:
                continue
            sn1, sn2 = f1_parts[-1], f2_parts[-1]
            fn1, fn2 = f1_parts[0], f2_parts[0]
            mid1 = " ".join(f1_parts[1:-1])
            mid2 = " ".join(f2_parts[1:-1])
            if sn1 != sn2 or fn1 != fn2:
                continue
            if not mid1 or not mid2:
                continue
            # Same first, same surname, both have middle names
            # Check if middles are truly different (not typos)
            if mid1 == mid2:
                continue
            # Typo detection: if edit distance <= 2, probably same person
            if len(mid1) > 3 and len(mid2) > 3:
                # Simple check: share first 3 chars?
                if mid1[:3] == mid2[:3]:
                    continue
                # One is abbreviation of other?
                if mid1.replace(".", "").replace(" ", "") == mid2.replace(".", "").replace(" ", ""):
                    continue
                if mid1.startswith(mid2) or mid2.startswith(mid1):
                    continue
            # Genuinely different middle names
            ark_splits.append({
                "pid": pid,
                "f1": fulls[i], "f2": fulls[j],
                "names": sorted(prof["names"]),
                "parties": sorted(prof["parties"]),
                "years": sorted(prof["years"]),
            })

print(f"Genuinely different middle names: {len(ark_splits)}")
for s in ark_splits[:30]:
    print(f"  PID {s['pid']}: '{s['f1']}' vs '{s['f2']}'")
    print(f"    names={s['names']} parties={s['parties']} years={s['years'][:5]}")


# ══════════════════════════════════════════════════════════════════════════
# ANALYSIS 2: ARK-based merges — different PIDs that share an ARK full name
# ══════════════════════════════════════════════════════════════════════════
print("\n=== ANALYSIS 2: ARK-based merge opportunities ===")

ark_full_to_pids = defaultdict(set)
for pid, prof in profiles.items():
    for af in prof["ark_fulls"]:
        if af:
            ark_full_to_pids[af].add(pid)

ark_merges = []
for af, pids in ark_full_to_pids.items():
    if len(pids) <= 1:
        continue
    # Multiple PIDs share the same ARK full name
    pid_list = sorted(pids, key=lambda p: int(p) if p.isdigit() else 999999999)
    prof_list = [(p, profiles[p]) for p in pid_list]
    # Check party + temporal compatibility
    for j in range(1, len(prof_list)):
        p1, pr1 = prof_list[0]
        p2, pr2 = prof_list[j]
        np1 = {normalize_party(x) for x in pr1["parties"]}
        np2 = {normalize_party(x) for x in pr2["parties"]}
        party_ok = bool(np1 & np2 - {""})
        y1 = sorted(pr1["years"])
        y2 = sorted(pr2["years"])
        gap = min(abs(a - b) for a in y1 for b in y2) if y1 and y2 else 999
        if party_ok and gap <= 20:
            ark_merges.append({
                "ark_full": af,
                "p1": p1, "p2": p2,
                "names1": sorted(pr1["names"]), "names2": sorted(pr2["names"]),
                "parties1": sorted(pr1["parties"]), "parties2": sorted(pr2["parties"]),
                "gap": gap,
            })

print(f"ARK-full-name merge candidates (party match, <=20yr gap): {len(ark_merges)}")
for m in ark_merges[:30]:
    print(f"  '{m['ark_full']}': PID {m['p1']} {m['names1']} vs PID {m['p2']} {m['names2']}")
    print(f"    parties: {m['parties1']} vs {m['parties2']}, gap={m['gap']}yr")


# ══════════════════════════════════════════════════════════════════════════
# ANALYSIS 3: Same-constituency same-date duplicates
# (same person appearing twice in one election — data error)
# ══════════════════════════════════════════════════════════════════════════
print("\n=== ANALYSIS 3: Same person appearing twice in same constituency/date ===")

election_key_to_pids = defaultdict(list)
for r in records:
    key = (r["date"], r["body"], r["const"])
    election_key_to_pids[key].append((r["pid"], r["name"], r["party"]))

dups = []
for key, entries in election_key_to_pids.items():
    pid_set = set(e[0] for e in entries)
    if len(entries) != len(pid_set):
        # Same PID appears multiple times in same constituency/date
        from collections import Counter
        pid_counts = Counter(e[0] for e in entries)
        for pid, count in pid_counts.items():
            if count > 1:
                dups.append({"key": key, "pid": pid, "count": count,
                             "entries": [e for e in entries if e[0] == pid]})

print(f"Same-PID duplicates in same election: {len(dups)}")
for d in dups[:15]:
    print(f"  {d['key']}: PID {d['pid']} appears {d['count']}x: {d['entries']}")


# ══════════════════════════════════════════════════════════════════════════
# ANALYSIS 4: Remaining unmatched variant-name merges from v5 "possible" list
# ══════════════════════════════════════════════════════════════════════════
print("\n=== ANALYSIS 4: v5 'possible' merges (party match, 15-40yr gap) ===")
v5_analysis = json.load(open("_tmp_v5_analysis.json", encoding="utf-8"))
possible = v5_analysis.get("possible", [])
# Check which are still unmerged
still_possible = []
for m in possible:
    p1, p2 = m["p1"], m["p2"]
    # Check if still separate PIDs in current data
    if p1 in profiles and p2 in profiles:
        still_possible.append(m)

print(f"Still-unmerged 'possible' merges: {len(still_possible)}")
for m in still_possible:
    print(f"  {m['n1']} (PID {m['p1']}) vs {m['n2']} (PID {m['p2']})")
    print(f"    parties: {m['parties1']} vs {m['parties2']}, gap={m['gap']}yr")
    # Check ARK
    af1 = ark_lookup.get(norm(m["n1"]), [])
    af2 = ark_lookup.get(norm(m["n2"]), [])
    shared = set(norm(f) for f in af1) & set(norm(f) for f in af2)
    if shared:
        print(f"    ARK CONFIRMS same person: {shared}")
    elif af1 and af2:
        print(f"    ARK shows different: {af1[:2]} vs {af2[:2]}")


# Save results
json.dump({
    "ark_splits": ark_splits,
    "ark_merges": ark_merges,
    "duplicates": [{"key": str(d["key"]), "pid": d["pid"], "count": d["count"]}
                   for d in dups],
    "possible_merges": still_possible,
}, open("_tmp_v7_analysis.json", "w", encoding="utf-8"),
    indent=2, default=str, ensure_ascii=False)

print(f"\nSaved analysis to _tmp_v7_analysis.json")
