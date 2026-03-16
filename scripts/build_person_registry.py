#!/usr/bin/env python3
"""Build person_registry.json from the Full election tables workbook."""

import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


def normalize_space(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_name_for_match(value: str) -> str:
    text = normalize_space(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().replace("\u2019", "'").replace("\u2018", "'")
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def main():
    xlsx_path = Path(
        "C:/Users/scomo/boundaries-website-archive/backups/"
        "requested-identity-fixes-round2-20260301-134629/"
        "full-workbook/Full election tables.xlsx"
    )
    if not xlsx_path.exists():
        print(f"ERROR: Workbook not found at {xlsx_path}")
        sys.exit(1)

    print(f"Loading workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # --- Read Names sheet ---
    ws_names = wb["Names"]
    nh = [c.value for c in next(ws_names.iter_rows(max_row=1))]
    ni = {h: nh.index(h) for h in nh if h is not None}

    persons = {}
    max_id = 0
    for row in ws_names.iter_rows(min_row=2, values_only=True):
        pid = row[ni["PersonID"]]
        if pid is None:
            continue
        pid = int(pid)
        full = normalize_space(row[ni["Full Name usually known by"]])
        first = normalize_space(row[ni["First Name"]])
        last = normalize_space(row[ni["Last Name"]])
        gender = normalize_space(row[ni.get("Gender", 0)]) if "Gender" in ni else ""
        if not full:
            continue
        if pid in persons:
            # Duplicate in Names sheet — add as variant
            existing = persons[pid]
            if full not in existing["nameVariants"]:
                existing["nameVariants"].append(full)
            continue
        persons[pid] = {
            "personId": pid,
            "canonicalName": full,
            "firstName": first,
            "lastName": last,
            "gender": gender,
            "nameVariants": [full],
            "matchKeys": [normalize_name_for_match(full)],
            "history": [],
        }
        max_id = max(max_id, pid)

    print(f"  Names sheet: {len(persons)} unique persons (max ID: {max_id})")

    # --- Read ElectionResults sheet to build history ---
    ws_er = wb["ElectionResults"]
    eh = [c.value for c in next(ws_er.iter_rows(max_row=1))]
    ei = {h: eh.index(h) for h in eh if h is not None}

    body_short = {
        "House of Commons of the United Kingdom": "westminster",
        "Northern Ireland Assembly": "assembly",
        "European Parliament": "european",
        "Northern Ireland Constitutional Convention": "convention",
        "Northern Ireland Forum for Political Dialogue": "forum",
    }

    row_count = 0
    for row in ws_er.iter_rows(min_row=2, values_only=True):
        rt = normalize_space(row[ei["ResultType"]])
        if rt != "Candidate" and not rt.startswith("ListCandidate"):
            continue
        pid = row[ei["PersonID"]]
        if pid is None:
            continue
        pid = int(pid)
        if pid not in persons:
            continue

        date_val = row[ei["Date"]]
        date_str = (
            date_val.strftime("%Y-%m-%d")
            if hasattr(date_val, "strftime")
            else str(date_val or "")[:10]
        )
        body = normalize_space(row[ei["ElectedBody"]])
        party = normalize_space(row[ei["Party Name"]])
        const = normalize_space(row[ei["Constituency"]])
        name = normalize_space(row[ei["Name usually known by"]])

        persons[pid]["history"].append(
            {
                "date": date_str,
                "body": body_short.get(body, body),
                "party": party,
                "constituency": const,
            }
        )

        # Add name variant if different from canonical
        if name and name not in persons[pid]["nameVariants"]:
            persons[pid]["nameVariants"].append(name)
            mk = normalize_name_for_match(name)
            if mk not in persons[pid]["matchKeys"]:
                persons[pid]["matchKeys"].append(mk)

        row_count += 1

    wb.close()
    print(f"  ElectionResults: {row_count} candidate rows processed")

    # Deduplicate history entries
    for p in persons.values():
        seen = set()
        deduped = []
        for h in p["history"]:
            key = (h["date"], h["body"], h["constituency"])
            if key not in seen:
                seen.add(key)
                deduped.append(h)
        p["history"] = sorted(deduped, key=lambda x: x["date"])

    registry = {
        "meta": {
            "version": 1,
            "nextId": max_id + 1,
            "totalPersons": len(persons),
            "source": "Full election tables.xlsx Names + ElectionResults",
        },
        "persons": {str(pid): data for pid, data in sorted(persons.items())},
    }

    out_path = Path("C:/Users/scomo/boundaries-website/scripts/person_registry.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(registry, f, ensure_ascii=False, indent=2)

    print(f"\nSaved: {out_path}")
    print(f"  {len(persons)} persons, next ID: {max_id + 1}")

    # Summary stats
    with_history = sum(1 for p in persons.values() if p["history"])
    orphans = sum(1 for p in persons.values() if not p["history"])
    print(f"  {with_history} with election history, {orphans} orphans (no elections)")


if __name__ == "__main__":
    main()
