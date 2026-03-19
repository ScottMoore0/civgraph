#!/usr/bin/env python
"""Apply v5 safe variant merges to the PersonID workbook.

Reads the analysis from analyze_v5.py and applies only safe merges
(party match + <=15yr temporal gap). Updates both ElectionResults and
Transfers sheets.
"""

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl


def norm(name):
    if not name: return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def main():
    input_path = Path("Full election tables - comprehensive - personid-v5.xlsx")
    analysis = json.load(open("_tmp_v5_analysis.json", encoding="utf-8"))
    safe = analysis["safe"]

    print(f"Loading {input_path}...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb["ElectionResults"]

    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))

    pid_col = col["PersonID"]
    rt_col = col["ResultType"]
    name_col = col["Name usually known by"]

    # Build PID merge map from safe merges
    # For each safe pair, merge p2 into p1 (keep lower PID)
    # But we need to be careful about chains: if A→B and B→C, then A→C
    merge_map: dict[str, str] = {}

    for m in safe:
        p1, p2 = m["p1"], m["p2"]
        # Pick the "primary" — prefer lower numeric value (older/more established)
        try:
            v1, v2 = int(p1), int(p2)
            primary, secondary = (p1, p2) if v1 < v2 else (p2, p1)
        except ValueError:
            primary, secondary = p1, p2
        # Follow chains
        while primary in merge_map:
            primary = merge_map[primary]
        while secondary in merge_map:
            secondary = merge_map[secondary]
        if primary != secondary:
            merge_map[secondary] = primary

    # Resolve all chains
    def resolve(pid):
        visited = set()
        while pid in merge_map and pid not in visited:
            visited.add(pid)
            pid = merge_map[pid]
        return pid

    final_map = {k: resolve(k) for k in merge_map}

    print(f"\nMerge map: {len(final_map)} PIDs will be merged")

    # Apply merges
    merged_rows = 0
    for i, row in enumerate(rows):
        if row[rt_col] not in ("Candidate", "NonTransferable"):
            continue
        pid = str(row[pid_col])
        if pid in final_map:
            rows[i][pid_col] = final_map[pid]
            merged_rows += 1

    print(f"Updated {merged_rows} rows in ElectionResults")

    # Write back ElectionResults
    for i, rd in enumerate(rows):
        for j, val in enumerate(rd):
            ws.cell(row=i + 2, column=j + 1, value=val)

    # Update Transfers
    print("Updating Transfers sheet...")
    name_to_pid: dict[str, str] = {}
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        name = row[name_col]
        pid = str(row[pid_col] or "")
        if name and pid:
            n = norm(name)
            if n:
                name_to_pid[n] = pid

    ws_t = wb["Transfers"]
    t_h = [cell.value for cell in ws_t[1]]
    t_pid_idx = t_h.index("PersonID")
    t_name_idx = t_h.index("Name")
    t_upd = 0
    for t_row in ws_t.iter_rows(min_row=2):
        nv = t_row[t_name_idx].value
        if nv:
            n = norm(nv)
            if n in name_to_pid:
                new = name_to_pid[n]
                if str(t_row[t_pid_idx].value or "") != new:
                    t_row[t_pid_idx].value = new
                    t_upd += 1
    print(f"Updated {t_upd} Transfers rows")

    # Save
    print(f"\nSaving {input_path}...")
    wb.save(input_path)

    # Final stats
    pid_names = defaultdict(set)
    pid_bodies = defaultdict(set)
    total = 0
    for row in rows:
        if row[rt_col] != "Candidate":
            continue
        total += 1
        pid = str(row[pid_col])
        pid_names[pid].add(row[name_col] or "")
        pid_bodies[pid].add(row[col["ElectedBody"]] or "")

    cross = sum(1 for bs in pid_bodies.values() if len(bs) > 1)
    multi = sum(1 for names in pid_names.values() if len(names) > 1)

    body_counts = defaultdict(int)
    for row in rows:
        if row[rt_col] != "Candidate": continue
        body_counts[row[col["ElectedBody"]]] += 1

    print(f"\n{'='*60}")
    print(f"  V5 FINAL STATE")
    print(f"{'='*60}")
    print(f"  Candidate rows: {total}")
    print(f"  Unique PersonIDs: {len(pid_names)}")
    print(f"  Cross-body people: {cross}")
    print(f"  PIDs with name variants: {multi}")
    print()
    for body in sorted(body_counts):
        bp = len({p for p, bs in pid_bodies.items() if body in bs})
        print(f"  {body}: {body_counts[body]} candidacies, {bp} people")

    # Log
    log = {
        "merges_applied": len(final_map),
        "rows_updated": merged_rows,
        "transfers_updated": t_upd,
        "final_pids": len(pid_names),
        "cross_body": cross,
        "multi_name": multi,
    }
    Path("personid_v5_log.json").write_text(
        json.dumps(log, indent=2, default=str), encoding="utf-8")
    print(f"\n  Log: personid_v5_log.json")


if __name__ == "__main__":
    main()
