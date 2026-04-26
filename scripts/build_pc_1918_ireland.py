#!/usr/bin/env python
"""Dissolve parlconst's numbered uMap features by 1918 constituency name
(per the Area Tables in each section's XLSX) for ROI sections 105-111,
combine with the parlconst NI 1918 layer (29 features), and emit
data/maps/parliamentary/PC_1918_Ireland.geojson (then convert to FGB
via ogr2ogr).
"""
import json, sys, io, subprocess, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook
from shapely.geometry import shape, mapping
from shapely.ops import unary_union

REPO = Path(__file__).resolve().parent.parent
TMP = REPO / "_tmp_parlconst"
UMAP_TMP = REPO / "_tmp_umap_parlconst" / "maps"
OUT_GEOJSON = REPO / "data" / "maps" / "parliamentary" / "PC_1918_Ireland.geojson"
OUT_FGB = OUT_GEOJSON.with_suffix(".fgb")
OGR2OGR = "C:/Program Files/GDAL/ogr2ogr.exe"

ROI_SECTIONS = {
    "105": "1154760_105-leinster-east.umap.json",
    "106": "1155329_106-leinster-north.umap.json",
    "107": "1155518_107-leinster-south-and-munster-east.umap.json",
    "108": "1155724_108-munster-limerick-and-kerry.umap.json",
    "109": "1156056_109-munster-cork.umap.json",
    "110": "1156494_110-connaught-south-and-munster-clare.umap.json",
    "111": "1158347_111-connaught-north-and-ulster.umap.json",
}
NI_UMAP = "918768_1918-constituencies-northern-ireland.umap.json"


def normalize(s):
    if s is None: return None
    s = str(s).strip()
    if not s: return None
    try:
        float(s); return None
    except ValueError:
        return s


def load_section_lookup():
    """For each ROI section: feature# -> 1918 constituency name."""
    lookup = {}
    for section in ROI_SECTIONS:
        xlsx = next(Path('_tmp_parlconst/files').glob(f'{section}_other_dl_*.xlsx'))
        wb = load_workbook(xlsx, data_only=True)
        ws = wb['Area Table']
        feat = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            try: f = int(row[1])
            except (TypeError, ValueError): continue
            if f < 1: continue
            c1885 = normalize(row[2])
            c1918 = normalize(row[3])
            if not c1918 and c1885: c1918 = c1885
            if c1918: feat[f] = c1918
        lookup[section] = feat
    return lookup


def collect_features_by_1918(lookup):
    """Walk each section's uMap export. For each feature with Label='N',
    tag with its 1918 constituency name from the lookup."""
    by_name_geoms = defaultdict(list)
    section_unmatched = defaultdict(list)
    for section, fname in ROI_SECTIONS.items():
        umap = json.loads((UMAP_TMP / fname).read_text(encoding='utf-8'))
        for layer in umap.get('layers', []):
            for feat in layer.get('features', []):
                p = feat.get('properties') or {}
                lab = p.get('Label')
                if lab is None: continue
                # Strip alpha suffixes like '15a' → 15
                m = ''
                for ch in str(lab):
                    if ch.isdigit(): m += ch
                    else: break
                if not m: continue
                fnum = int(m)
                name = lookup[section].get(fnum)
                if not name:
                    section_unmatched[section].append(lab)
                    continue
                by_name_geoms[name].append(feat['geometry'])
    return by_name_geoms, section_unmatched


def collect_ni_features():
    """Take the parlconst NI 1918 layer (29 features named directly)."""
    umap = json.loads((UMAP_TMP / NI_UMAP).read_text(encoding='utf-8'))
    by_name_geoms = defaultdict(list)
    for layer in umap.get('layers', []):
        for feat in layer.get('features', []):
            name = (feat.get('properties') or {}).get('Name')
            if not name: continue
            by_name_geoms[name.strip()].append(feat['geometry'])
    return by_name_geoms


def dissolve_to_features(by_name_geoms):
    out = []
    for name, geoms in sorted(by_name_geoms.items()):
        polys = [shape(g) for g in geoms]
        merged = unary_union(polys)
        if merged.is_empty: continue
        out.append({
            "type": "Feature",
            "properties": {"Name": name},
            "geometry": mapping(merged),
        })
    return out


def main():
    print("Loading per-section feature# → 1918-name lookup ...")
    lookup = load_section_lookup()
    total_lookups = sum(len(v) for v in lookup.values())
    print(f"  {total_lookups} feature mappings across {len(lookup)} ROI sections")

    print("\nCollecting ROI features from parlconst uMap exports ...")
    roi_geoms, unmatched = collect_features_by_1918(lookup)
    n_roi_feat = sum(len(v) for v in roi_geoms.values())
    print(f"  {n_roi_feat} numbered features grouped into {len(roi_geoms)} 1918 constituencies")
    for s, labs in unmatched.items():
        if labs: print(f"    {s}: {len(labs)} unmatched labels: {labs[:10]}")

    print("\nCollecting NI features from parlconst NI 1918 export ...")
    ni_geoms = collect_ni_features()
    print(f"  {sum(len(v) for v in ni_geoms.values())} features → {len(ni_geoms)} NI constituencies")

    # Combine. Names should not collide (NI names like 'Belfast Falls' vs ROI 'Dublin Falls'-style).
    overlap = set(roi_geoms) & set(ni_geoms)
    if overlap:
        print(f"\nWARNING: {len(overlap)} name overlaps between ROI and NI: {overlap}")
    combined = defaultdict(list)
    for d in (roi_geoms, ni_geoms):
        for k, v in d.items():
            combined[k].extend(v)

    print("\nDissolving by 1918 constituency name ...")
    feats = dissolve_to_features(combined)
    print(f"  {len(feats)} dissolved 1918 constituencies")

    fc = {"type": "FeatureCollection", "features": feats}
    OUT_GEOJSON.write_text(json.dumps(fc, ensure_ascii=False), encoding='utf-8')
    print(f"\nwrote {OUT_GEOJSON}  ({OUT_GEOJSON.stat().st_size/1e6:.1f} MB)")

    # Convert to FGB
    if OUT_FGB.exists(): OUT_FGB.unlink()
    r = subprocess.run([OGR2OGR, "-f", "FlatGeobuf",
                        "-nlt", "PROMOTE_TO_MULTI",
                        "-skipfailures",
                        str(OUT_FGB), str(OUT_GEOJSON)],
                       capture_output=True, text=True)
    if r.returncode != 0:
        print(f"ogr2ogr failed: {r.stderr[:300]}")
        sys.exit(1)
    print(f"wrote {OUT_FGB}  ({OUT_FGB.stat().st_size/1e6:.1f} MB)")
    # gzip
    import gzip
    gz = OUT_FGB.with_suffix(OUT_FGB.suffix + ".gz")
    with OUT_FGB.open('rb') as src, gzip.open(gz, 'wb') as dst:
        shutil.copyfileobj(src, dst)
    print(f"wrote {gz}  ({gz.stat().st_size/1e6:.1f} MB)")


if __name__ == "__main__":
    main()
