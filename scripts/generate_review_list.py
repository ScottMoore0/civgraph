#!/usr/bin/env python
"""Generate a ranked review list of PersonID decisions most likely to be erroneous.

Outputs an Excel workbook with one row per review case, sorted from
most-likely-wrong to least-likely-wrong, with context and suggested action.
"""

import json
import re
import unicodedata
from collections import defaultdict, Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def norm(name):
    if not name: return ""
    nfkd = unicodedata.normalize("NFKD", name)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    lowered = stripped.lower()
    lowered = re.sub(r"[^a-z0-9 '\-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def split_fn_sn(name):
    parts = name.strip().split()
    if not parts: return ("", "")
    if len(parts) == 1: return ("", norm(parts[0]))
    return (norm(" ".join(parts[:-1])), norm(parts[-1]))


def normalize_party(p):
    if not p: return ""
    l = p.lower()
    for needle, code in [("sdlp","sdlp"),("democratic unionist","dup"),("dup","dup"),
        ("ulster unionist","uup"),("uup","uup"),("sinn","sf"),("alliance","all"),
        ("independent","ind"),("green","grn"),("tuv","tuv"),("workers","wp"),
        ("pup","pup"),("labour","lab"),("conservative","con"),("ukip","ukip"),
        ("ukup","ukup"),("vanguard","vupp"),("socialist","soc"),("people before","pbp")]:
        if needle in l: return code
    return l[:15]


VGROUPS = [
    ["james","jim","jimmy","seamus"],["william","will","willie","bill","billy","liam"],
    ["robert","rob","robbie","bob","bobby"],["john","johnny","jack","sean","shane"],
    ["thomas","tom","tommy"],["edward","ed","eddie","ted","eamonn","eamon","edmund"],
    ["richard","dick","dickie","ricky"],["charles","charlie","cathal"],
    ["patrick","paddy","pat","patsy","padraig"],["michael","mike","mick","mickey"],
    ["christopher","chris"],["daniel","dan","danny"],["andrew","andy","drew"],
    ["anthony","tony"],["joseph","joe","joey"],["samuel","sam","sammy"],
    ["alexander","alex","alec","alistair","alastair"],["frederick","fred","freddie"],
    ["kenneth","ken","kenny"],["ronald","ron","ronnie"],["raymond","ray"],
    ["stephen","steve","steven"],["philip","phil"],["lawrence","larry","laurence"],
    ["gerald","gerry","gerard","gearoid"],["peter","pete"],["david","dave","davy"],
    ["henry","harry","hal"],["albert","bert","bertie"],["alfred","alf","alfie"],
    ["francis","frank","frankie","proinsias"],["desmond","des"],["stanley","stan"],
    ["ernest","ernie"],["norman","norm"],["donald","don","donnie"],["dennis","denis"],
    ["terence","terry","terrence"],["vincent","vince"],["douglas","doug"],
    ["timothy","tim"],["nicholas","nick"],["geoffrey","geoff","jeff","jeffrey"],
    ["elizabeth","liz","lizzie","beth","betty","bess"],
    ["margaret","maggie","meg","peggy","madge"],
    ["catherine","kate","katie","cathy","kathleen"],
    ["patricia","pat","tricia","trish"],["dorothy","dot","dolly"],
    ["jacqueline","jackie"],["anne","ann","annie"],
    ["mary","molly","may","maire"],["bridget","brid","bridie"],
    ["niall","neal","neil"],["ciaran","kieran"],
    ["archibald","archie"],["reginald","reg","reggie"],
    ["leonard","len","lenny"],["bernard","bernie"],
    ["pamela","pam"],["susan","sue","susie"],["deborah","deb","debbie"],
    ["christine","chris","christina","tina"],["rosemary","rosie"],["caroline","carol"],
    ["joanna","jo","joanne","joan"],["jennifer","jenny","jen"],
    ["cornelius","con"],["roderick","roddy","rod"],["wallace","wally"],
    ["cecil","cec"],["clifford","cliff"],["sydney","sid"],["percy","percival"],
    ["owen","eugene"],["malachy","malachi"],
]
_VM = {}
for gi, grp in enumerate(VGROUPS):
    for n in grp: _VM[n.lower()] = gi

def are_fn_var(f1, f2):
    n1, n2 = f1.lower().strip(), f2.lower().strip()
    if n1 == n2: return True
    g1, g2 = _VM.get(n1), _VM.get(n2)
    if g1 is not None and g2 is not None and g1 == g2: return True
    if len(n1) >= 3 and len(n2) >= 3 and (n1.startswith(n2) or n2.startswith(n1)): return True
    return False


def main():
    print("Loading data...")
    ark_candidates = json.load(open("_tmp_ark_candidates.json", encoding="utf-8"))
    ark_lookup = json.load(open("_tmp_ark_name_lookup.json", encoding="utf-8"))

    ark_names = set()
    for c in ark_candidates:
        ark_names.add(norm(c["full_name"]))
        parts = c["full_name"].split()
        if len(parts) >= 2:
            ark_names.add(norm(f"{parts[0]} {parts[-1]}"))

    # ARK context: (norm_short, norm_const, year) -> norm_full
    ark_ctx = {}
    for c in ark_candidates:
        full = c["full_name"]
        parts = full.split()
        if len(parts) < 2: continue
        short = f"{parts[0]} {parts[-1]}"
        ark_ctx[(norm(short), norm(c["constituency"]), c["year"])] = full

    wb = openpyxl.load_workbook("Full election tables - comprehensive - personid-v8.xlsx", read_only=True)
    ws = wb["ElectionResults"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(headers)}

    # Read all candidate records
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col["ResultType"]] != "Candidate":
            continue
        date = str(row[col["Date"]] or "")
        year = date[:4] if date[:4].isdigit() else ""
        records.append({
            "pid": str(row[col["PersonID"]]),
            "name": row[col["Name usually known by"]] or "",
            "party": row[col["Party Name"]] or "",
            "date": date,
            "year": int(year) if year else 0,
            "body": row[col["ElectedBody"]] or "",
            "const": row[col["Constituency"]] or "",
        })

    # Build PID profiles
    profiles = defaultdict(lambda: {
        "names": set(), "parties": set(), "years": set(),
        "bodies": set(), "consts": set(), "records": [],
    })
    for r in records:
        p = profiles[r["pid"]]
        p["names"].add(r["name"])
        p["parties"].add(r["party"])
        if r["year"]: p["years"].add(r["year"])
        p["bodies"].add(r["body"])
        p["consts"].add(r["const"])
        p["records"].append(r)

    # Build name indexes
    norm_to_pids = defaultdict(set)
    sn_to_pids = defaultdict(set)
    for pid, p in profiles.items():
        for name in p["names"]:
            n = norm(name)
            if n:
                norm_to_pids[n].add(pid)
            fn, sn = split_fn_sn(name)
            if sn:
                sn_to_pids[sn].add(pid)

    review_cases = []

    # ══════════════════════════════════════════════════════════════════════
    # CATEGORY 1: Same name, same constituency, same year, different PID
    # ══════════════════════════════════════════════════════════════════════
    print("Category 1: Same name, same constituency, same year, different PID...")
    election_key_names = defaultdict(list)
    for r in records:
        key = (r["date"], r["body"], r["const"])
        election_key_names[key].append((r["pid"], r["name"], r["party"]))

    for key, entries in election_key_names.items():
        norm_groups = defaultdict(list)
        for pid, name, party in entries:
            norm_groups[norm(name)].append((pid, name, party))
        for n, group in norm_groups.items():
            pids = set(e[0] for e in group)
            if len(pids) > 1:
                date, body, const = key
                for pid, name, party in group:
                    other_pids = pids - {pid}
                    review_cases.append({
                        "category": 1,
                        "priority": 100,
                        "pid": pid,
                        "name": name,
                        "party": party,
                        "date": date,
                        "body": body,
                        "const": const,
                        "issue": f"Same name+constituency+date as PID(s) {','.join(sorted(other_pids))}",
                        "suggested_action": "Merge if same person, or investigate data error",
                        "other_pids": ",".join(sorted(other_pids)),
                        "ark_full": ark_ctx.get((norm(name), norm(const), date[:4]), ""),
                    })

    cat1 = sum(1 for r in review_cases if r["category"] == 1)
    print(f"  Found: {cat1}")

    # ══════════════════════════════════════════════════════════════════════
    # CATEGORY 2: Same name, same year, different PID, different constituency
    # ══════════════════════════════════════════════════════════════════════
    print("Category 2: Same name, same year, different PID, same party...")
    year_name_pids = defaultdict(lambda: defaultdict(set))
    for r in records:
        if r["year"]:
            year_name_pids[(r["year"], norm(r["name"]))][r["pid"]].add(
                (r["party"], r["const"], r["body"])
            )

    for (year, n), pid_details in year_name_pids.items():
        if len(pid_details) <= 1:
            continue
        pid_list = sorted(pid_details.keys())
        for i in range(len(pid_list)):
            for j in range(i + 1, len(pid_list)):
                p1, p2 = pid_list[i], pid_list[j]
                d1, d2 = pid_details[p1], pid_details[p2]
                parties1 = {normalize_party(party) for party, _, _ in d1}
                parties2 = {normalize_party(party) for party, _, _ in d2}
                consts1 = {const for _, const, _ in d1}
                consts2 = {const for _, const, _ in d2}
                # Only flag if same party (different party = probably different people)
                if parties1 & parties2 - {"ind", ""}:
                    # Skip if already caught in category 1
                    if consts1 & consts2:
                        continue
                    prof1 = profiles[p1]
                    prof2 = profiles[p2]
                    review_cases.append({
                        "category": 2,
                        "priority": 90,
                        "pid": p1,
                        "name": sorted(prof1["names"])[0],
                        "party": ",".join(sorted(parties1)),
                        "date": str(year),
                        "body": "",
                        "const": ",".join(sorted(consts1))[:60],
                        "issue": f"Same name+year+party as PID {p2} in {','.join(sorted(consts2))[:60]}",
                        "suggested_action": "Likely different people (different constituency) — confirm",
                        "other_pids": p2,
                        "ark_full": "",
                    })

    cat2 = sum(1 for r in review_cases if r["category"] == 2)
    print(f"  Found: {cat2}")

    # ══════════════════════════════════════════════════════════════════════
    # CATEGORY 3: PIDs with 15-25yr gap between consecutive appearances
    # ══════════════════════════════════════════════════════════════════════
    print("Category 3: PIDs with 15-25yr gap (grey zone era splits)...")
    for pid, prof in profiles.items():
        years = sorted(prof["years"])
        if len(years) < 2:
            continue
        for k in range(len(years) - 1):
            gap = years[k + 1] - years[k]
            if 15 <= gap <= 25:
                # Get records before and after gap
                before = [r for r in prof["records"] if r["year"] <= years[k]]
                after = [r for r in prof["records"] if r["year"] >= years[k + 1]]
                bp = {r["party"] for r in before}
                ap = {r["party"] for r in after}
                bc = {r["const"] for r in before}
                ac = {r["const"] for r in after}
                party_cont = bool({normalize_party(p) for p in bp} & {normalize_party(p) for p in ap} - {""})

                # Check ARK
                ark_fulls_before = set()
                ark_fulls_after = set()
                for r in before:
                    af = ark_ctx.get((norm(r["name"]), norm(r["const"]), str(r["year"])), "")
                    if af: ark_fulls_before.add(af)
                for r in after:
                    af = ark_ctx.get((norm(r["name"]), norm(r["const"]), str(r["year"])), "")
                    if af: ark_fulls_after.add(af)
                ark_match = bool(set(norm(f) for f in ark_fulls_before) &
                                 set(norm(f) for f in ark_fulls_after)) if ark_fulls_before and ark_fulls_after else None

                priority = 80
                if not party_cont:
                    priority = 85  # No party continuity = more suspicious
                if ark_match is False:
                    priority = 88  # ARK says different = very suspicious
                elif ark_match is True:
                    priority = 60  # ARK confirms = less suspicious

                review_cases.append({
                    "category": 3,
                    "priority": priority,
                    "pid": pid,
                    "name": sorted(prof["names"])[0],
                    "party": f"Before: {','.join(sorted(bp)[:3])} | After: {','.join(sorted(ap)[:3])}",
                    "date": f"{years[k]}-{years[k+1]}",
                    "body": ",".join(sorted(prof["bodies"])),
                    "const": f"Before: {','.join(sorted(bc))[:40]} | After: {','.join(sorted(ac))[:40]}",
                    "issue": f"{gap}yr gap ({years[k]} to {years[k+1]}), party_cont={party_cont}",
                    "suggested_action": "Split if different person, keep if same person with long break",
                    "other_pids": "",
                    "ark_full": f"Before: {','.join(sorted(ark_fulls_before))[:50]} | After: {','.join(sorted(ark_fulls_after))[:50]}" if ark_fulls_before or ark_fulls_after else "",
                })
                break  # Only flag the worst gap per PID

    cat3 = sum(1 for r in review_cases if r["category"] == 3)
    print(f"  Found: {cat3}")

    # ══════════════════════════════════════════════════════════════════════
    # CATEGORY 4: Uncorroborated single-appearance PIDs sharing a name
    #             with a different PID (potential missed merges)
    # ══════════════════════════════════════════════════════════════════════
    print("Category 4: Uncorroborated candidates with name-sharing PIDs...")
    for pid, prof in profiles.items():
        pid_val = int(pid) if pid.isdigit() else 0
        if pid_val <= 100019:
            continue  # Skip curated PIDs
        if len(prof["years"]) > 1:
            continue  # Skip multi-appearance
        if not prof["names"]:
            continue

        name = sorted(prof["names"])[0]
        n = norm(name)
        fn, sn = split_fn_sn(name)

        # Find other PIDs with same or variant name
        candidate_pids = set()
        # Exact norm match
        for other_pid in norm_to_pids.get(n, set()):
            if other_pid != pid:
                candidate_pids.add(other_pid)
        # Variant match (same surname, variant first name)
        for other_pid in sn_to_pids.get(sn, set()):
            if other_pid == pid:
                continue
            other_prof = profiles[other_pid]
            for other_name in other_prof["names"]:
                other_fn, other_sn = split_fn_sn(other_name)
                if other_sn == sn and are_fn_var(fn, other_fn):
                    candidate_pids.add(other_pid)

        if not candidate_pids:
            continue

        # Score each candidate match
        best_score = 0
        best_pid = ""
        best_detail = ""
        for cp in candidate_pids:
            cp_prof = profiles[cp]
            np1 = {normalize_party(p) for p in prof["parties"]}
            np2 = {normalize_party(p) for p in cp_prof["parties"]}
            party_match = bool(np1 & np2 - {"", "ind"})
            y1 = sorted(prof["years"])
            y2 = sorted(cp_prof["years"])
            gap = min(abs(a - b) for a in y1 for b in y2) if y1 and y2 else 999
            geo = bool(prof["consts"] & cp_prof["consts"])

            score = 0
            if party_match: score += 40
            elif bool(np1 & np2 - {""}): score += 20  # ind match
            if gap <= 5: score += 20
            elif gap <= 10: score += 15
            elif gap <= 20: score += 10
            if geo: score += 30
            score += 10  # name base

            if score > best_score:
                best_score = score
                best_pid = cp
                best_detail = f"PID {cp} ({sorted(cp_prof['names'])[0]}, {','.join(sorted(cp_prof['parties'])[:2])}, {sorted(cp_prof['years'])[:2]})"

        if best_score < 40:
            continue  # Too weak

        # Check ARK
        ark_full = ""
        for r in prof["records"]:
            af = ark_ctx.get((norm(r["name"]), norm(r["const"]), str(r["year"])), "")
            if af:
                ark_full = af
                break

        priority = max(10, 70 - best_score // 2)  # Higher score = lower priority (less likely wrong)

        review_cases.append({
            "category": 4,
            "priority": priority,
            "pid": pid,
            "name": name,
            "party": ",".join(sorted(prof["parties"])),
            "date": ",".join(str(y) for y in sorted(prof["years"])),
            "body": ",".join(sorted(prof["bodies"])),
            "const": ",".join(sorted(prof["consts"]))[:60],
            "issue": f"Single-appearance, name matches {best_detail} (score={best_score})",
            "suggested_action": "Merge if same person" if best_score >= 60 else "Probably different people — confirm",
            "other_pids": best_pid,
            "ark_full": ark_full,
        })

    cat4 = sum(1 for r in review_cases if r["category"] == 4)
    print(f"  Found: {cat4}")

    # ══════════════════════════════════════════════════════════════════════
    # CATEGORY 5: PIDs with multiple name variants not in dictionary
    # ══════════════════════════════════════════════════════════════════════
    print("Category 5: PIDs with unusual name variants...")
    for pid, prof in profiles.items():
        if len(prof["names"]) <= 1:
            continue
        names = sorted(prof["names"])
        # Check all pairs
        for i in range(len(names)):
            for j in range(i + 1, len(names)):
                n1, n2 = names[i], names[j]
                if norm(n1) == norm(n2):
                    continue
                fn1, sn1 = split_fn_sn(n1)
                fn2, sn2 = split_fn_sn(n2)
                if sn1 != sn2:
                    review_cases.append({
                        "category": 5,
                        "priority": 75,
                        "pid": pid,
                        "name": f"{n1} / {n2}",
                        "party": ",".join(sorted(prof["parties"])[:3]),
                        "date": f"{min(prof['years'])}-{max(prof['years'])}" if prof["years"] else "",
                        "body": ",".join(sorted(prof["bodies"])),
                        "const": "",
                        "issue": f"Different surnames in same PID: '{n1}' vs '{n2}'",
                        "suggested_action": "Split if different people",
                        "other_pids": "",
                        "ark_full": "",
                    })
                    break
            else:
                continue
            break

    cat5 = sum(1 for r in review_cases if r["category"] == 5)
    print(f"  Found: {cat5}")

    # ══════════════════════════════════════════════════════════════════════
    # CATEGORY 6: ARK shows different middle names within same PID
    # ══════════════════════════════════════════════════════════════════════
    print("Category 6: ARK middle-name conflicts within same PID...")
    for pid, prof in profiles.items():
        ark_fulls = set()
        for r in prof["records"]:
            af = ark_ctx.get((norm(r["name"]), norm(r["const"]), str(r["year"])))
            if af:
                ark_fulls.add(af)
        if len(ark_fulls) <= 1:
            continue
        # Check for genuinely different middle names
        nf_list = sorted(ark_fulls)
        for i in range(len(nf_list)):
            for j in range(i + 1, len(nf_list)):
                p1 = nf_list[i].split()
                p2 = nf_list[j].split()
                if len(p1) < 3 or len(p2) < 3:
                    continue
                if p1[-1] != p2[-1] or p1[0] != p2[0]:
                    continue
                m1 = " ".join(p1[1:-1])
                m2 = " ".join(p2[1:-1])
                if m1 and m2 and m1 != m2:
                    if not m1.startswith(m2[:3]) and not m2.startswith(m1[:3]):
                        if m1.replace(".", "").replace(" ", "") != m2.replace(".", "").replace(" ", ""):
                            review_cases.append({
                                "category": 6,
                                "priority": 70,
                                "pid": pid,
                                "name": sorted(prof["names"])[0],
                                "party": ",".join(sorted(prof["parties"])[:3]),
                                "date": f"{min(prof['years'])}-{max(prof['years'])}" if prof["years"] else "",
                                "body": ",".join(sorted(prof["bodies"])),
                                "const": "",
                                "issue": f"ARK shows '{nf_list[i]}' vs '{nf_list[j]}'",
                                "suggested_action": "Split if ARK middle names indicate different people",
                                "other_pids": "",
                                "ark_full": f"{nf_list[i]} vs {nf_list[j]}",
                            })
                            break
            else:
                continue
            break

    cat6 = sum(1 for r in review_cases if r["category"] == 6)
    print(f"  Found: {cat6}")

    # ══════════════════════════════════════════════════════════════════════
    # Sort and write output
    # ══════════════════════════════════════════════════════════════════════

    # Deduplicate (same PID can appear in multiple categories)
    seen_keys = set()
    deduped = []
    for r in review_cases:
        key = (r["category"], r["pid"], r.get("other_pids", ""))
        if key not in seen_keys:
            seen_keys.add(key)
            deduped.append(r)

    deduped.sort(key=lambda r: (-r["priority"], r["category"], r["pid"]))

    print(f"\nTotal review cases: {len(deduped)}")
    print(f"  Cat 1 (same name/const/date, diff PID): {sum(1 for r in deduped if r['category'] == 1)}")
    print(f"  Cat 2 (same name/year/party, diff const): {sum(1 for r in deduped if r['category'] == 2)}")
    print(f"  Cat 3 (15-25yr gap): {sum(1 for r in deduped if r['category'] == 3)}")
    print(f"  Cat 4 (uncorroborated, name matches another PID): {sum(1 for r in deduped if r['category'] == 4)}")
    print(f"  Cat 5 (unusual name variants): {sum(1 for r in deduped if r['category'] == 5)}")
    print(f"  Cat 6 (ARK middle-name conflict): {sum(1 for r in deduped if r['category'] == 6)}")

    # Write Excel
    out_path = Path("PersonID review list.xlsx")
    out_wb = openpyxl.Workbook()
    ws_out = out_wb.active
    ws_out.title = "Review"

    out_headers = [
        "Priority", "Category", "PID", "Name", "Party", "Date/Year",
        "Body", "Constituency", "Issue", "ARK Full Name",
        "Other PID(s)", "Suggested Action", "Decision", "Notes",
    ]
    ws_out.append(out_headers)

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for c in range(1, len(out_headers) + 1):
        cell = ws_out.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font

    cat_names = {
        1: "Same name+const+date",
        2: "Same name+year+party",
        3: "15-25yr gap",
        4: "Uncorroborated name match",
        5: "Unusual name variant",
        6: "ARK middle-name conflict",
    }

    for r in deduped:
        ws_out.append([
            r["priority"],
            f"Cat {r['category']}: {cat_names[r['category']]}",
            r["pid"],
            r["name"],
            r["party"],
            r["date"],
            r["body"],
            r["const"],
            r["issue"],
            r["ark_full"],
            r.get("other_pids", ""),
            r["suggested_action"],
            "",  # Decision (blank for user)
            "",  # Notes (blank for user)
        ])

    # Auto-width columns
    for c in range(1, len(out_headers) + 1):
        max_len = len(str(out_headers[c - 1]))
        for row in ws_out.iter_rows(min_row=2, min_col=c, max_col=c):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max_len + 2

    out_wb.save(out_path)
    print(f"\nSaved to: {out_path}")


if __name__ == "__main__":
    main()
